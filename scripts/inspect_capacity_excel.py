import json
import sys
from collections import Counter

from openpyxl import load_workbook


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: inspect_capacity_excel.py <xlsx-path>")
        return 1

    path = sys.argv[1]
    wb = load_workbook(path, data_only=False, read_only=True)
    report = {"sheets": []}

    for ws in wb.worksheets:
        formulas = []
        labels = []
        non_empty = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value in (None, ""):
                    continue
                non_empty += 1
                if is_formula(value):
                    formulas.append({"cell": cell.coordinate, "formula": value})
                elif isinstance(value, str):
                    text = value.strip()
                    if text:
                        labels.append({"cell": cell.coordinate, "text": text})

        report["sheets"].append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "non_empty_cells": non_empty,
                "formula_count": len(formulas),
                "formula_samples": formulas[:60],
                "top_labels": labels[:80],
                "formula_prefixes": Counter(
                    formula["formula"].split("(")[0] for formula in formulas
                ).most_common(20),
            }
        )

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
