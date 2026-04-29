import csv
import threading
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

import customtkinter as ctk

from core.api_client import ApiClientError, deep_research_ai_lead, delete_ai_lead, enrich_ai_lead_from_apollo, enrich_ai_lead_from_hunter, generate_ai_email_sequence, get_ai_lead_provider_status, import_ai_leads_csv, list_ai_leads, search_apollo_ai_leads, search_hunter_companies, search_serpapi_domains
from core.session import get_app_token
from core.utils import apply_bomaksan_table_style, apply_zebra_striping
from lead_otomasyonu.lead_detail_screen import lead_detay_ekrani
from lead_otomasyonu.segment_settings_screen import segment_ayarlari_ekrani
from lead_otomasyonu.strategy_constants import (
    MOCK_LEADS,
    PRIORITY_OPTIONS,
    PRODUCT_CATEGORIES,
    SALES_CHANNELS,
    TARGET_COUNTRIES,
    build_segment_name,
    get_sequence_code,
    priority_for_segment,
)


def lead_otomasyonu_ekrani(parent=None, kullanici_rolu=None):
    win = ctk.CTkToplevel(parent) if parent else ctk.CTkToplevel()
    win.title("Lead Otomasyonu")
    win.geometry("1440x860")
    win.minsize(1120, 720)
    win.configure(fg_color="#f5f5f5")

    try:
        win.state("zoomed")
        win.lift()
        win.focus_force()
        win.attributes("-topmost", True)
        win.after(350, lambda: win.attributes("-topmost", False))
    except Exception:
        pass

    state = {
        "leads": [dict(item) for item in MOCK_LEADS],
        "filtered": [],
        "api_mode": False,
    }

    root = ctk.CTkFrame(win, fg_color="transparent")
    root.pack(fill="both", expand=True, padx=20, pady=20)
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(3, weight=1)

    header = ctk.CTkFrame(root, fg_color="#ffffff", corner_radius=16, border_width=1, border_color="#e5e7eb")
    header.grid(row=0, column=0, sticky="ew", pady=(0, 14))
    header.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        header,
        text="Lead Otomasyonu",
        font=ctk.CTkFont(family="Inter", size=28, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, sticky="w", padx=22, pady=(18, 4))
    status_var = ctk.StringVar(value="MVP iskeleti mock veriyle hazır. API bağlandığında aynı ekran canlı veriyi kullanacak.")
    ctk.CTkLabel(
        header,
        textvariable=status_var,
        font=ctk.CTkFont(size=13),
        text_color="#64748b",
    ).grid(row=1, column=0, sticky="w", padx=22, pady=(0, 18))

    actions = ctk.CTkFrame(header, fg_color="transparent")
    actions.grid(row=0, column=1, rowspan=2, sticky="e", padx=22, pady=18)

    summary = ctk.CTkFrame(root, fg_color="transparent")
    summary.grid(row=1, column=0, sticky="ew", pady=(0, 14))
    summary.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)
    metric_vars = {
        "total": ctk.StringVar(value="0"),
        "pending": ctk.StringVar(value="0"),
        "high": ctk.StringVar(value="0"),
        "excluded": ctk.StringVar(value="0"),
        "drafts": ctk.StringVar(value="0"),
        "approval": ctk.StringVar(value="0"),
    }
    metric_defs = [
        ("Toplam Lead", "total", "#2563eb"),
        ("AI Bekleyen", "pending", "#7c3aed"),
        ("High / Very High", "high", "#15803d"),
        ("Excluded", "excluded", "#dc2626"),
        ("Taslak", "drafts", "#b45309"),
        ("Onay Bekleyen", "approval", "#0f766e"),
    ]
    for index, (title, key, color) in enumerate(metric_defs):
        _metric_card(summary, index, title, metric_vars[key], color)

    filters = ctk.CTkFrame(root, fg_color="#ffffff", corner_radius=14, border_width=1, border_color="#e5e7eb")
    filters.grid(row=2, column=0, sticky="ew", pady=(0, 14))
    filters.grid_columnconfigure(6, weight=1)

    search_var = ctk.StringVar()
    channel_var = ctk.StringVar(value="Tüm Kanallar")
    product_var = ctk.StringVar(value="Tüm Ürünler")
    priority_var = ctk.StringVar(value="Tüm Öncelikler")

    _filter_label(filters, "Arama", 0)
    search_entry = ctk.CTkEntry(filters, textvariable=search_var, width=220, placeholder_text="Firma, ülke veya segment")
    search_entry.grid(row=1, column=0, padx=(16, 8), pady=(0, 14), sticky="ew")

    _filter_label(filters, "Satış Kanalı", 1)
    channel_combo = ctk.CTkComboBox(filters, values=["Tüm Kanallar"] + SALES_CHANNELS, variable=channel_var, width=220)
    channel_combo.grid(row=1, column=1, padx=8, pady=(0, 14), sticky="ew")

    _filter_label(filters, "Ürün / Hizmet", 2)
    product_combo = ctk.CTkComboBox(filters, values=["Tüm Ürünler"] + PRODUCT_CATEGORIES, variable=product_var, width=200)
    product_combo.grid(row=1, column=2, padx=8, pady=(0, 14), sticky="ew")

    _filter_label(filters, "Öncelik", 3)
    priority_combo = ctk.CTkComboBox(filters, values=["Tüm Öncelikler"] + PRIORITY_OPTIONS, variable=priority_var, width=170)
    priority_combo.grid(row=1, column=3, padx=8, pady=(0, 14), sticky="ew")

    filter_actions = ctk.CTkFrame(filters, fg_color="transparent")
    filter_actions.grid(row=1, column=4, columnspan=2, sticky="w", padx=(10, 8), pady=(0, 14))
    bulk_research_frame = ctk.CTkFrame(filters, fg_color="transparent")
    bulk_research_frame.grid(row=2, column=0, columnspan=7, sticky="ew", padx=16, pady=(0, 12))
    bulk_research_frame.grid_columnconfigure(0, weight=1)
    bulk_research_progress = ctk.CTkProgressBar(bulk_research_frame, mode="determinate")
    bulk_research_progress.grid(row=0, column=0, sticky="ew", padx=(0, 12))
    bulk_research_progress.set(0)
    bulk_research_status_var = ctk.StringVar(value="")
    ctk.CTkLabel(bulk_research_frame, textvariable=bulk_research_status_var, text_color="#64748b", width=260, anchor="w").grid(row=0, column=1, sticky="e")
    bulk_research_frame.grid_remove()
    bulk_enrich_frame = ctk.CTkFrame(filters, fg_color="transparent")
    bulk_enrich_frame.grid(row=3, column=0, columnspan=7, sticky="ew", padx=16, pady=(0, 12))
    bulk_enrich_frame.grid_columnconfigure(0, weight=1)
    bulk_enrich_progress = ctk.CTkProgressBar(bulk_enrich_frame, mode="determinate")
    bulk_enrich_progress.grid(row=0, column=0, sticky="ew", padx=(0, 12))
    bulk_enrich_progress.set(0)
    bulk_enrich_status_var = ctk.StringVar(value="")
    ctk.CTkLabel(bulk_enrich_frame, textvariable=bulk_enrich_status_var, text_color="#64748b", width=260, anchor="w").grid(row=0, column=1, sticky="e")
    bulk_enrich_frame.grid_remove()

    table_card = ctk.CTkFrame(root, fg_color="#ffffff", corner_radius=14, border_width=1, border_color="#e5e7eb")
    table_card.grid(row=3, column=0, sticky="nsew")
    table_card.grid_rowconfigure(0, weight=1)
    table_card.grid_columnconfigure(0, weight=1)

    table_wrap = ctk.CTkFrame(table_card, fg_color="transparent")
    table_wrap.grid(row=0, column=0, sticky="nsew", padx=16, pady=16)
    table_wrap.grid_rowconfigure(0, weight=1)
    table_wrap.grid_columnconfigure(0, weight=1)

    y_scroll = ttk.Scrollbar(table_wrap, orient="vertical")
    x_scroll = ttk.Scrollbar(table_wrap, orient="horizontal")
    columns = (
        "company_name",
        "contact_name",
        "contact_email",
        "country",
        "sales_channel",
        "product_category",
        "priority",
        "email_sequence_stage",
    )
    tree = ttk.Treeview(
        table_wrap,
        columns=columns,
        show="headings",
        yscrollcommand=y_scroll.set,
        xscrollcommand=x_scroll.set,
        selectmode="extended",
    )
    y_scroll.config(command=tree.yview)
    x_scroll.config(command=tree.xview)
    apply_bomaksan_table_style(tree)

    headings = {
        "company_name": "Firma",
        "contact_name": "Kişi",
        "contact_email": "Email",
        "country": "Ülke",
        "sales_channel": "Satış Kanalı",
        "product_category": "Ürün / Hizmet",
        "priority": "Öncelik",
        "email_sequence_stage": "Email Sekans Aşaması",
    }
    widths = {
        "company_name": 240,
        "contact_name": 170,
        "contact_email": 220,
        "country": 120,
        "sales_channel": 210,
        "product_category": 160,
        "priority": 110,
        "email_sequence_stage": 190,
    }
    for col in columns:
        tree.heading(col, text=headings[col])
        tree.column(col, width=widths[col], minwidth=70, anchor="w")

    tree.grid(row=0, column=0, sticky="nsew")
    y_scroll.grid(row=0, column=1, sticky="ns")
    x_scroll.grid(row=1, column=0, sticky="ew")

    def refresh_metrics():
        leads = state["leads"]
        metric_vars["total"].set(str(len(leads)))
        metric_vars["pending"].set(str(sum(1 for item in leads if item.get("ai_status") in {"New", "Pending AI Analysis"})))
        metric_vars["high"].set(str(sum(1 for item in leads if item.get("priority") in {"High", "Very High"})))
        metric_vars["excluded"].set(str(sum(1 for item in leads if item.get("priority") == "Excluded" or item.get("ai_status") == "Excluded")))
        metric_vars["drafts"].set(str(sum(1 for item in leads if item.get("ai_status") == "Draft Generated")))
        metric_vars["approval"].set(str(sum(1 for item in leads if item.get("approval_status") in {"Awaiting Approval", "Review Needed"})))

    def apply_filters(*_args):
        query = search_var.get().strip().casefold()
        channel = channel_var.get()
        product = product_var.get()
        priority = priority_var.get()
        filtered = []
        for item in state["leads"]:
            haystack = " ".join(str(item.get(key, "")) for key in ("company_name", "contact_name", "contact_email", "email_status", "enrichment_note", "country", "segment_name", "sales_channel", "product_category")).casefold()
            if query and query not in haystack:
                continue
            if channel != "Tüm Kanallar" and item.get("sales_channel") != channel:
                continue
            if product != "Tüm Ürünler" and item.get("product_category") != product:
                continue
            if priority != "Tüm Öncelikler" and item.get("priority") != priority:
                continue
            filtered.append(item)
        state["filtered"] = filtered
        render_table()

    def render_table():
        children = tree.get_children()
        if children:
            tree.delete(*children)
        for item in state["filtered"]:
            values = [_table_value(item, col) for col in columns]
            tree.insert("", "end", iid=str(item.get("id")), values=values)
        apply_zebra_striping(tree, tree.get_children())
        refresh_metrics()

    def selected_lead():
        leads = selected_leads()
        return leads[0] if leads else None

    def has_completed_research(lead):
        status = str(lead.get("research_status") or "").strip().casefold()
        summary = str(lead.get("research_summary") or "").strip()
        return bool(summary) or status == "completed"

    def selected_leads():
        selected_ids = {str(item_id) for item_id in tree.selection()}
        if not selected_ids:
            return []
        leads = []
        for item in state["leads"]:
            if str(item.get("id")) in selected_ids:
                leads.append(item)
        return leads

    def open_detail(_event=None):
        lead = selected_lead()
        if not lead:
            messagebox.showwarning("Lead Otomasyonu", "Lütfen bir lead seçin.", parent=win)
            return
        lead_detay_ekrani(win, lead, on_update=apply_filters)

    def add_manual_lead():
        dialog = ctk.CTkToplevel(win)
        dialog.title("Lead Ekle")
        dialog.geometry("560x660")
        dialog.configure(fg_color="#f5f5f5")
        dialog.transient(win)
        dialog.grab_set()

        form = ctk.CTkFrame(dialog, fg_color="#ffffff", corner_radius=14)
        form.pack(fill="both", expand=True, padx=18, pady=18)
        vars_ = {
            "company_name": ctk.StringVar(),
            "contact_name": ctk.StringVar(),
            "contact_title": ctk.StringVar(),
            "contact_email": ctk.StringVar(),
            "country": ctk.StringVar(),
            "sales_channel": ctk.StringVar(value="White Label / Resellers"),
            "product_category": ctk.StringVar(value="Dust Collection"),
            "detected_activity": ctk.StringVar(),
        }
        _form_entry(form, "Firma", vars_["company_name"], 0)
        _form_entry(form, "Kişi", vars_["contact_name"], 1)
        _form_entry(form, "Unvan", vars_["contact_title"], 2)
        _form_entry(form, "Email", vars_["contact_email"], 3)
        _form_entry(form, "Ülke", vars_["country"], 4)
        _form_combo(form, "Satış Kanalı", vars_["sales_channel"], SALES_CHANNELS, 5)
        _form_combo(form, "Ürün / Hizmet", vars_["product_category"], PRODUCT_CATEGORIES, 6)
        _form_entry(form, "Aktivite", vars_["detected_activity"], 7)

        def save():
            company = vars_["company_name"].get().strip()
            if not company:
                messagebox.showwarning("Eksik Bilgi", "Firma adı zorunludur.", parent=dialog)
                return
            product = vars_["product_category"].get()
            channel = vars_["sales_channel"].get()
            priority = priority_for_segment(product, channel)
            lead = {
                "id": _next_id(state["leads"]),
                "company_name": company,
                "contact_name": vars_["contact_name"].get().strip(),
                "contact_title": vars_["contact_title"].get().strip(),
                "contact_email": vars_["contact_email"].get().strip(),
                "email_status": "user managed" if vars_["contact_email"].get().strip() else "missing",
                "country": vars_["country"].get().strip(),
                "source": "Manual",
                "sales_channel": channel,
                "product_category": product,
                "segment_name": build_segment_name(product, channel),
                "priority": priority,
                "ai_score": 0,
                "suggested_sequence": get_sequence_code(channel),
                "ai_status": "Segmented",
                "approval_status": "Awaiting Approval",
                "last_action": "Lead eklendi ve segment önerildi",
                "website": "",
                "detected_activity": vars_["detected_activity"].get().strip(),
                "short_reasoning": "MVP kural setiyle otomatik segment önerisi oluşturuldu.",
                "personalization_angle": "Partner adayının faaliyet alanına göre kısa kişiselleştirme.",
            }
            state["leads"].append(lead)
            dialog.destroy()
            apply_filters()

        ctk.CTkButton(form, text="Kaydet", command=save, fg_color="#d32f2f", hover_color="#b91c1c").grid(row=9, column=1, sticky="e", padx=16, pady=18)

    def import_csv():
        token = get_app_token()
        if not token:
            messagebox.showerror("Apollo Import", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        path = filedialog.askopenfilename(
            parent=win,
            title="Apollo CSV/Excel Seç",
            filetypes=[("Apollo Export", "*.csv *.xlsx"), ("CSV Dosyası", "*.csv"), ("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
        )
        if not path:
            return
        try:
            rows = _read_import_rows(path)
            summary = _summarize_import_rows(rows)
            if not rows:
                messagebox.showwarning("Apollo Import", "Dosyada içe aktarılacak satır bulunamadı.", parent=win)
                return
            prompt = (
                f"{summary['rows']} satır bulundu.\n"
                f"{summary['companies']} benzersiz firma görünüyor.\n"
                f"{summary['emails']} email alanı dolu.\n"
                f"{summary['people']} kişi adı/ünvanı bulunan satır var.\n\n"
                "Bu dosya Lead Otomasyonu veritabanına aktarılsın mı?"
            )
            if not messagebox.askyesno("Apollo Import", prompt, parent=win):
                return
        except Exception as exc:
            messagebox.showerror("Apollo Import", f"Dosya okunamadı: {exc}", parent=win)
            return

        status_var.set("Apollo export içe aktarılıyor...")

        def worker():
            try:
                result = import_ai_leads_csv(token, rows) or {}
                created = int(result.get("created_leads") or result.get("created") or 0)
                updated = int(result.get("updated_leads") or 0)
                contacts = int(result.get("created_contacts") or 0)
                contact_updates = int(result.get("updated_contacts") or 0)
                duplicates = int(result.get("skipped_duplicates") or 0)
                invalid = int(result.get("skipped_invalid") or 0)
                message = (
                    f"{created} yeni firma lead'i oluşturuldu.\n"
                    f"{updated} mevcut firma güncellendi.\n"
                    f"{contacts} yeni kişi/email eklendi.\n"
                    f"{contact_updates} kişi/email güncellendi.\n"
                    f"{duplicates} tekrar firma mevcut kayıtla eşleştirildi."
                )
                if invalid:
                    message += f"\n{invalid} satır firma adı olmadığı için atlandı."
                win.after(0, load_from_api)
                win.after(0, lambda msg=message: messagebox.showinfo("Apollo Import", msg, parent=win))
            except Exception as exc:
                win.after(0, lambda err=str(exc): messagebox.showerror("Apollo Import", f"İçe aktarma başarısız: {err}", parent=win))
            finally:
                win.after(0, lambda: status_var.set(""))

        threading.Thread(target=worker, daemon=True).start()

    def load_from_api():
        token = get_app_token()
        if not token:
            status_var.set("API oturumu bulunamadı; mock veri gösteriliyor.")
            return

        def worker():
            try:
                leads = list_ai_leads(token)
                if leads:
                    state["leads"] = leads
                    state["api_mode"] = True
                    win.after(0, lambda: status_var.set("Canlı API verisi yüklendi."))
                    win.after(0, apply_filters)
            except ApiClientError as exc:
                message = f"API hazır değil veya erişilemiyor; mock veri kullanılıyor. Detay: {exc}"
                win.after(0, lambda msg=message: status_var.set(msg))
            except Exception as exc:
                message = f"Lead verisi yüklenemedi; mock veri kullanılıyor. Detay: {exc}"
                win.after(0, lambda msg=message: status_var.set(msg))

        threading.Thread(target=worker, daemon=True).start()

    def apollo_search():
        token = get_app_token()
        if not token:
            messagebox.showerror("Apollo", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return

        dialog = ctk.CTkToplevel(win)
        dialog.title("Apollo Search")
        dialog.geometry("560x460")
        dialog.configure(fg_color="#f5f5f5")
        dialog.transient(win)
        dialog.grab_set()

        form = ctk.CTkFrame(dialog, fg_color="#ffffff", corner_radius=14)
        form.pack(fill="both", expand=True, padx=18, pady=18)
        vars_ = {
            "country": ctk.StringVar(value="Germany"),
            "titles": ctk.StringVar(value="Business Development Manager, Sales Manager, Managing Director"),
            "keywords": ctk.StringVar(value="industrial ventilation, dust collection, fume extraction"),
            "per_page": ctk.StringVar(value="25"),
        }
        _form_entry(form, "Ülke", vars_["country"], 0)
        _form_entry(form, "Unvanlar", vars_["titles"], 1)
        _form_entry(form, "Anahtar Kelimeler", vars_["keywords"], 2)
        _form_entry(form, "Limit", vars_["per_page"], 3)
        note = ctk.CTkLabel(
            form,
            text="Not: Apollo People Search email döndürmez. Email için enrichment çalıştırılır ve ai_lead_contacts.email alanına yazılır.",
            text_color="#64748b",
            wraplength=460,
            justify="left",
        )
        note.grid(row=4, column=0, columnspan=2, sticky="w", padx=16, pady=(10, 4))

        def run_search():
            payload = {
                "country": vars_["country"].get().strip(),
                "person_titles": [item.strip() for item in vars_["titles"].get().split(",") if item.strip()],
                "keywords": [item.strip() for item in vars_["keywords"].get().split(",") if item.strip()],
                "per_page": int(vars_["per_page"].get() or 25),
            }

            def worker():
                try:
                    result = search_apollo_ai_leads(token, payload)
                    created = int(result.get("created") or 0)
                    win.after(0, lambda: messagebox.showinfo("Apollo", f"{created} lead eklendi. Email için enrichment çalıştırın.", parent=win))
                    win.after(0, load_from_api)
                    win.after(0, dialog.destroy)
                except Exception as exc:
                    win.after(0, lambda err=str(exc): messagebox.showerror("Apollo", f"Apollo search başarısız: {err}", parent=dialog))

            threading.Thread(target=worker, daemon=True).start()

        ctk.CTkButton(form, text="Apollo'dan Lead Çek", command=run_search, fg_color="#d32f2f", hover_color="#b91c1c").grid(row=5, column=1, sticky="e", padx=16, pady=18)

    def segment_search():
        token = get_app_token()
        if not token:
            messagebox.showerror("SerpAPI Firma Bul", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        dialog = ctk.CTkToplevel(win)
        dialog.title("SerpAPI Firma Bul")
        dialog.geometry("600x500")
        dialog.configure(fg_color="#f5f5f5")
        dialog.transient(win)
        dialog.grab_set()

        form = ctk.CTkFrame(dialog, fg_color="#ffffff", corner_radius=14)
        form.pack(fill="both", expand=True, padx=18, pady=18)

        vars_ = {
            "country": ctk.StringVar(value=TARGET_COUNTRIES[0] if TARGET_COUNTRIES else "Germany"),
            "pages": ctk.StringVar(value="1"),
            "search_mode": ctk.StringVar(value="Geniş Arama"),
            "search_engine": ctk.StringVar(value="Google"),
        }
        keyword_tags = [
            "dust collection distributor",
            "industrial filtration company",
            "fume extraction supplier",
        ]
        _form_country_selector(form, "Ülke", vars_["country"], TARGET_COUNTRIES, 0, dialog)
        keyword_entry = _keyword_tag_editor(form, "Anahtar Kelimeler", keyword_tags, 1)
        _form_entry(form, "Arama Sayfası", vars_["pages"], 2)

        ctk.CTkLabel(form, text="Arama Motoru", text_color="#475569").grid(row=3, column=0, sticky="w", padx=16, pady=8)
        ctk.CTkSegmentedButton(
            form,
            values=["Google", "Bing", "Google + Bing"],
            variable=vars_["search_engine"],
        ).grid(row=3, column=1, sticky="ew", padx=16, pady=8)

        ctk.CTkLabel(form, text="Arama Esnekliği", text_color="#475569").grid(row=4, column=0, sticky="w", padx=16, pady=8)
        ctk.CTkSegmentedButton(
            form,
            values=["Geniş Arama", "Dar Arama"],
            variable=vars_["search_mode"],
        ).grid(row=4, column=1, sticky="ew", padx=16, pady=8)

        note = ctk.CTkLabel(
            form,
            text="Geniş arama keyword kelimelerini esnek eşleştirir. Dar arama keywordü tırnak içinde exact phrase olarak arar. Bu ekran sadece SerpAPI ile firma/domain adayı bulur; Apollo karar verici/email araması ayrıca Email Enrich ile çalışır.",
            text_color="#64748b",
            wraplength=500,
            justify="left",
        )
        note.grid(row=5, column=0, columnspan=2, sticky="w", padx=16, pady=(12, 4))

        progress = ctk.CTkProgressBar(form, mode="indeterminate")
        progress.grid(row=6, column=0, columnspan=2, sticky="ew", padx=16, pady=(12, 2))
        progress.set(0)
        progress.grid_remove()
        status_label = ctk.CTkLabel(form, text="", text_color="#64748b")
        status_label.grid(row=7, column=0, columnspan=2, sticky="w", padx=16, pady=(2, 4))
        search_state = {"running": False, "message_index": 0}
        progress_messages = [
            "Sihir konuşuyor, keywordler yola çıktı...",
            "İlgili firmalar bulunuyor...",
            "Domainler ayıklanıyor, alakasız kalabalık dışarıda kalıyor...",
            "Bulunan adaylar dashboard'a hazırlanıyor...",
        ]

        def tick_progress():
            if not search_state["running"]:
                return
            status_label.configure(text=progress_messages[search_state["message_index"] % len(progress_messages)])
            search_state["message_index"] += 1
            dialog.after(1400, tick_progress)

        def run_segment_search():
            _add_keyword_tags_from_entry(keyword_entry, keyword_tags)
            keywords = list(keyword_tags)
            if not keywords:
                messagebox.showwarning("Eksik Bilgi", "Lütfen en az bir keyword girin.", parent=dialog)
                return
            try:
                pages = max(1, min(int(vars_["pages"].get() or 1), 10))
            except Exception:
                messagebox.showwarning("Eksik Bilgi", "Arama sayfası sayı olmalı.", parent=dialog)
                return

            payload = {
                "country": vars_["country"].get().strip(),
                "keywords": keywords,
                "pages": pages,
                "search_mode": "exact" if vars_["search_mode"].get() == "Dar Arama" else "broad",
                "search_engine": "google+bing" if vars_["search_engine"].get() == "Google + Bing" else vars_["search_engine"].get().casefold(),
            }
            search_state["running"] = True
            search_state["message_index"] = 0
            progress.grid()
            progress.start()
            status_label.configure(text=progress_messages[0])
            search_button.configure(state="disabled", text="Aranıyor...")
            tick_progress()

            def finish_search():
                search_state["running"] = False
                progress.stop()
                progress.grid_remove()
                status_label.configure(text="")
                search_button.configure(state="normal", text="SerpAPI Firma Bul")

            def worker():
                try:
                    result = search_serpapi_domains(token, payload)
                    created = int(result.get("created") or 0)
                    skipped = int(result.get("skipped_duplicates") or 0)
                    domains = int(result.get("found_domains") or 0)
                    warning_count = int(result.get("warning_count") or 0)
                    warning_note = f"\n\n{warning_count} SerpAPI sorgusu hata verdi ve atlandı." if warning_count else ""
                    win.after(0, finish_search)
                    win.after(0, lambda: messagebox.showinfo("SerpAPI Firma Bul", f"{domains} firma/domain adayı bulundu. {created} aday eklendi. {skipped} tekrar kayıt atlandı.{warning_note}", parent=win))
                    win.after(0, load_from_api)
                    win.after(0, dialog.destroy)
                except Exception as exc:
                    win.after(0, finish_search)
                    win.after(0, lambda err=str(exc): messagebox.showerror("SerpAPI Firma Bul", f"Arama başarısız: {err}", parent=dialog))

            threading.Thread(target=worker, daemon=True).start()

        search_button = ctk.CTkButton(
            form,
            text="SerpAPI Firma Bul",
            command=run_segment_search,
            fg_color="#d32f2f",
            hover_color="#b91c1c",
        )
        search_button.grid(row=8, column=1, sticky="e", padx=16, pady=18)

    def hunter_company_search():
        token = get_app_token()
        if not token:
            messagebox.showerror("Hunter Firma Bul", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        dialog = ctk.CTkToplevel(win)
        dialog.title("Hunter Firma Bul")
        dialog.geometry("600x430")
        dialog.configure(fg_color="#f5f5f5")
        dialog.transient(win)
        dialog.grab_set()

        form = ctk.CTkFrame(dialog, fg_color="#ffffff", corner_radius=14)
        form.pack(fill="both", expand=True, padx=18, pady=18)
        vars_ = {
            "country": ctk.StringVar(value="Germany"),
            "limit": ctk.StringVar(value="25"),
        }
        keyword_tags = [
            "industrial filtration distributor",
            "dust collection equipment supplier",
            "fume extraction supplier",
        ]
        _form_country_selector(form, "Ülke", vars_["country"], TARGET_COUNTRIES, 0, dialog)
        keyword_entry = _keyword_tag_editor(form, "Anahtar Kelimeler", keyword_tags, 1)
        _form_entry(form, "Kayıt Limiti", vars_["limit"], 2)
        note = ctk.CTkLabel(
            form,
            text="Hunter Discover firma/domain adayı bulur. Ürün x segment etiketi daha sonra manuel atanır; email kişileri için Hunter Email Bulucu çalıştırılır.",
            text_color="#64748b",
            wraplength=500,
            justify="left",
        )
        note.grid(row=3, column=0, columnspan=2, sticky="w", padx=16, pady=(12, 4))
        progress = ctk.CTkProgressBar(form, mode="indeterminate")
        progress.grid(row=4, column=0, columnspan=2, sticky="ew", padx=16, pady=(12, 2))
        progress.grid_remove()
        status_label = ctk.CTkLabel(form, text="", text_color="#64748b")
        status_label.grid(row=5, column=0, columnspan=2, sticky="w", padx=16, pady=(2, 4))

        def run_search():
            _add_keyword_tags_from_entry(keyword_entry, keyword_tags)
            keywords = list(keyword_tags)
            if not keywords:
                messagebox.showwarning("Eksik Bilgi", "Lütfen en az bir keyword girin.", parent=dialog)
                return
            try:
                limit = max(1, min(int(vars_["limit"].get() or 25), 100))
            except Exception:
                messagebox.showwarning("Eksik Bilgi", "Kayıt limiti sayı olmalı.", parent=dialog)
                return
            payload = {
                "country": vars_["country"].get().strip(),
                "keywords": keywords,
                "limit": limit,
            }
            progress.grid()
            progress.start()
            status_label.configure(text="Hunter firma/domain adaylarını arıyor...")
            search_button.configure(state="disabled", text="Aranıyor...")

            def finish_search():
                progress.stop()
                progress.grid_remove()
                status_label.configure(text="")
                search_button.configure(state="normal", text="Hunter Firma Bul")

            def worker():
                try:
                    result = search_hunter_companies(token, payload)
                    created = int(result.get("created") or 0)
                    found = int(result.get("found_companies") or 0)
                    skipped = int(result.get("skipped_duplicates") or 0)
                    win.after(0, finish_search)
                    win.after(0, lambda: messagebox.showinfo("Hunter Firma Bul", f"{found} firma/domain adayı bulundu. {created} aday eklendi. {skipped} tekrar kayıt atlandı.", parent=win))
                    win.after(0, load_from_api)
                    win.after(0, dialog.destroy)
                except Exception as exc:
                    win.after(0, finish_search)
                    win.after(0, lambda err=str(exc): messagebox.showerror("Hunter Firma Bul", f"Arama başarısız: {err}", parent=dialog))

            threading.Thread(target=worker, daemon=True).start()

        search_button = ctk.CTkButton(
            form,
            text="Hunter Firma Bul",
            command=run_search,
            fg_color="#d32f2f",
            hover_color="#b91c1c",
        )
        search_button.grid(row=6, column=1, sticky="e", padx=16, pady=18)

    def open_segment_settings():
        segment_ayarlari_ekrani(win)

    def show_provider_status():
        token = get_app_token()
        if not token:
            messagebox.showerror("Entegrasyon Durumu", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        status_var.set("Entegrasyon durumları kontrol ediliyor...")

        def format_provider(label, data):
            status = data.get("status") or "-"
            message = data.get("message") or ""
            key_hint = data.get("key_hint") or ""
            extra = []
            if data.get("account_email"):
                extra.append(f"Hesap: {data.get('account_email')}")
            if data.get("plan_name"):
                extra.append(f"Plan: {data.get('plan_name')}")
            if data.get("credits_available") is not None:
                extra.append(f"Kalan kredi: {data.get('credits_available')}")
            if data.get("total_searches_left") is not None:
                extra.append(f"Kalan arama: {data.get('total_searches_left')}")
            if data.get("this_month_usage") is not None:
                extra.append(f"Bu ay kullanılan: {data.get('this_month_usage')}")
            if data.get("account_rate_limit_per_hour") is not None:
                extra.append(f"Saatlik limit: {data.get('account_rate_limit_per_hour')}")
            suffix = f" ({key_hint})" if key_hint else ""
            extra_text = f"\n  " + "\n  ".join(extra) if extra else ""
            return f"{label}: {status}{suffix}\n  {message}{extra_text}"

        def worker():
            try:
                result = get_ai_lead_provider_status(token)
                lines = [
                    format_provider("Apollo", result.get("apollo") or {}),
                    format_provider("SerpAPI", result.get("serpapi") or {}),
                    format_provider("Hunter", result.get("hunter") or {}),
                ]
                message = "\n\n".join(lines)
                win.after(0, lambda: status_var.set("Entegrasyon durumu kontrol edildi."))
                win.after(0, lambda msg=message: messagebox.showinfo("Entegrasyon Durumu", msg, parent=win))
            except Exception as exc:
                win.after(0, lambda: status_var.set("Entegrasyon durumu alınamadı."))
                win.after(0, lambda err=str(exc): messagebox.showerror("Entegrasyon Durumu", f"Kontrol başarısız: {err}", parent=win))

        threading.Thread(target=worker, daemon=True).start()

    def enrich_selected():
        token = get_app_token()
        leads = selected_leads()
        if not token:
            messagebox.showerror("Apollo", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        if not leads:
            messagebox.showwarning("Apollo", "Lütfen enrichment yapılacak leadleri seçin.", parent=win)
            return
        total = len(leads)
        bulk_enrich_progress.set(0)
        bulk_enrich_status_var.set(f"Apollo email araması başlıyor: 0/{total}")
        bulk_enrich_frame.grid()
        apollo_email_button.configure(state="disabled", text="Aranıyor...")
        hunter_email_button.configure(state="disabled")

        def update_enrich_progress(done, lead_name=""):
            progress = done / total if total else 0
            bulk_enrich_progress.set(progress)
            name_suffix = f" | {lead_name}" if lead_name else ""
            bulk_enrich_status_var.set(f"Apollo email arıyor: {done}/{total}{name_suffix}")

        def finish_enrich_progress():
            apollo_email_button.configure(state="normal", text="Apollo Email Bulucu")
            hunter_email_button.configure(state="normal")
            bulk_enrich_frame.grid_remove()
            bulk_enrich_status_var.set("")
            bulk_enrich_progress.set(0)

        def worker():
            success = 0
            failed = 0
            last_note = ""
            try:
                for index, lead in enumerate(leads, start=1):
                    lead_name = str(lead.get("company_name") or "Lead")
                    win.after(0, lambda done=index - 1, name=lead_name: update_enrich_progress(done, name))
                    try:
                        result = enrich_ai_lead_from_apollo(token, lead.get("id"))
                        contact = result.get("contact") or {}
                        if contact:
                            lead["contact_email"] = contact.get("email") or lead.get("contact_email") or ""
                            lead["email_status"] = contact.get("email_status") or lead.get("email_status") or "enriched"
                            lead["enrichment_note"] = contact.get("enrichment_note") or lead.get("enrichment_note") or ""
                            lead["contact_name"] = contact.get("name") or lead.get("contact_name") or ""
                            lead["contact_title"] = contact.get("title") or lead.get("contact_title") or ""
                        note = lead.get("enrichment_note") or "Apollo enrichment tamamlandı."
                        lead["last_action"] = note
                        last_note = note
                        success += 1
                    except Exception:
                        failed += 1
                    win.after(0, lambda done=index, name=lead_name: update_enrich_progress(done, name))
                win.after(0, apply_filters)
                win.after(0, finish_enrich_progress)
                message = f"{success} lead için enrichment tamamlandı."
                if failed:
                    message += f" {failed} lead başarısız oldu."
                if success == 1 and last_note:
                    message += f"\n\n{last_note}"
                win.after(0, lambda msg=message: messagebox.showinfo("Apollo", msg, parent=win))
            except Exception as exc:
                win.after(0, finish_enrich_progress)
                win.after(0, lambda err=str(exc): messagebox.showerror("Apollo", f"Apollo enrichment başarısız: {err}", parent=win))

        threading.Thread(target=worker, daemon=True).start()

    def hunter_enrich_selected():
        token = get_app_token()
        leads = selected_leads()
        if not token:
            messagebox.showerror("Hunter", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        if not leads:
            messagebox.showwarning("Hunter", "Lütfen email aranacak leadleri seçin.", parent=win)
            return
        total = len(leads)
        bulk_enrich_progress.set(0)
        bulk_enrich_status_var.set(f"Hunter email araması başlıyor: 0/{total}")
        bulk_enrich_frame.grid()
        hunter_email_button.configure(state="disabled", text="Aranıyor...")
        apollo_email_button.configure(state="disabled")

        def update_enrich_progress(done, lead_name=""):
            progress = done / total if total else 0
            bulk_enrich_progress.set(progress)
            name_suffix = f" | {lead_name}" if lead_name else ""
            bulk_enrich_status_var.set(f"Hunter email arıyor: {done}/{total}{name_suffix}")

        def finish_enrich_progress():
            hunter_email_button.configure(state="normal", text="Hunter Email Bulucu")
            apollo_email_button.configure(state="normal")
            bulk_enrich_frame.grid_remove()
            bulk_enrich_status_var.set("")
            bulk_enrich_progress.set(0)

        def worker():
            success = 0
            failed = 0
            last_note = ""
            try:
                for index, lead in enumerate(leads, start=1):
                    lead_name = str(lead.get("company_name") or "Lead")
                    win.after(0, lambda done=index - 1, name=lead_name: update_enrich_progress(done, name))
                    try:
                        result = enrich_ai_lead_from_hunter(token, lead.get("id"))
                        contact = result.get("contact") or {}
                        refreshed_lead = result.get("lead") or {}
                        if refreshed_lead:
                            lead.update(refreshed_lead)
                        if contact:
                            lead["contact_email"] = contact.get("email") or lead.get("contact_email") or ""
                            lead["email_status"] = contact.get("email_status") or lead.get("email_status") or "hunter_found"
                            lead["enrichment_note"] = contact.get("enrichment_note") or lead.get("enrichment_note") or ""
                            lead["contact_name"] = contact.get("name") or lead.get("contact_name") or ""
                            lead["contact_title"] = contact.get("title") or lead.get("contact_title") or ""
                        note = lead.get("enrichment_note") or "Hunter Domain Search tamamlandı."
                        lead["last_action"] = note
                        last_note = note
                        success += 1
                    except Exception:
                        failed += 1
                    win.after(0, lambda done=index, name=lead_name: update_enrich_progress(done, name))
                win.after(0, apply_filters)
                win.after(0, finish_enrich_progress)
                message = f"{success} lead için Hunter email araması tamamlandı."
                if failed:
                    message += f" {failed} lead başarısız oldu."
                if success == 1 and last_note:
                    message += f"\n\n{last_note}"
                win.after(0, lambda msg=message: messagebox.showinfo("Hunter", msg, parent=win))
            except Exception as exc:
                win.after(0, finish_enrich_progress)
                win.after(0, lambda err=str(exc): messagebox.showerror("Hunter", f"Hunter email araması başarısız: {err}", parent=win))

        threading.Thread(target=worker, daemon=True).start()

    def create_sequence_for_selected():
        token = get_app_token()
        lead = selected_lead()
        if not token:
            messagebox.showerror("Email Sekansı", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        if not lead:
            messagebox.showwarning("Email Sekansı", "Lütfen sekans oluşturulacak lead'i seçin.", parent=win)
            return

        def worker():
            try:
                result = generate_ai_email_sequence(token, lead.get("id"))
                drafts = result.get("drafts") or []
                lead["draft_count"] = len(drafts)
                lead["ai_status"] = "Draft Generated"
                lead["approval_status"] = "Awaiting Approval"
                lead["last_action"] = "3 adımlı email sekansı taslakları oluşturuldu."
                win.after(0, apply_filters)
                win.after(0, lambda: messagebox.showinfo("Email Sekansı", f"{len(drafts)} email taslağı oluşturuldu. Gönderim yapılmadı; detay ekranından kontrol edip onaylayabilirsiniz.", parent=win))
            except Exception as exc:
                win.after(0, lambda err=str(exc): messagebox.showerror("Email Sekansı", f"Sekans oluşturulamadı: {err}", parent=win))

        threading.Thread(target=worker, daemon=True).start()

    def research_selected():
        token = get_app_token()
        leads = selected_leads()
        if not token:
            messagebox.showerror("AI Araştır", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        if not leads:
            messagebox.showwarning("AI Araştır", "Lütfen araştırılacak leadleri seçin.", parent=win)
            return
        skipped_existing = [
            lead
            for lead in leads
            if has_completed_research(lead)
        ]
        leads = [lead for lead in leads if lead not in skipped_existing]
        if not leads:
            messagebox.showinfo("AI Araştır", "AI Araştırması zaten yapıldı.", parent=win)
            return
        total = len(leads)
        bulk_research_progress.set(0)
        bulk_research_status_var.set(f"AI araştırma başlıyor: 0/{total}")
        bulk_research_frame.grid()
        ai_research_button.configure(state="disabled", text="Araştırılıyor...")

        def update_research_progress(done, lead_name=""):
            progress = done / total if total else 0
            bulk_research_progress.set(progress)
            name_suffix = f" | {lead_name}" if lead_name else ""
            bulk_research_status_var.set(f"AI araştırıyor: {done}/{total}{name_suffix}")

        def finish_research_progress():
            ai_research_button.configure(state="normal", text="AI Araştır")
            bulk_research_frame.grid_remove()
            bulk_research_status_var.set("")
            bulk_research_progress.set(0)

        def worker():
            completed = 0
            failed = 0
            last_error = ""
            for index, lead in enumerate(leads, start=1):
                lead_name = str(lead.get("company_name") or "Lead")
                win.after(0, lambda done=index - 1, name=lead_name: update_research_progress(done, name))
                try:
                    result = deep_research_ai_lead(token, lead.get("id"))
                    research = result.get("research") or {}
                    refreshed_lead = result.get("lead") or {}
                    if refreshed_lead:
                        lead.update(refreshed_lead)
                    lead["research_status"] = research.get("status") or "Completed"
                    lead["research_summary"] = research.get("company_overview") or ""
                    lead["last_action"] = "AI firma araştırması tamamlandı."
                    completed += 1
                except Exception as exc:
                    failed += 1
                    last_error = str(exc)
                win.after(0, lambda done=index, name=lead_name: update_research_progress(done, name))
            win.after(0, apply_filters)
            win.after(0, finish_research_progress)
            message = f"{completed} lead için AI araştırma tamamlandı."
            if skipped_existing:
                message += f" {len(skipped_existing)} lead zaten araştırıldığı için atlandı."
            if failed:
                message += f" {failed} lead araştırılamadı."
                if last_error:
                    message += f"\n\nSon hata: {last_error}"
            win.after(0, lambda msg=message: messagebox.showinfo("AI Araştır", msg, parent=win))

        threading.Thread(target=worker, daemon=True).start()

    def delete_selected_lead():
        token = get_app_token()
        leads = selected_leads()
        if not token:
            messagebox.showerror("Lead Sil", "API oturumu bulunamadı. Lütfen yeniden giriş yapın.", parent=win)
            return
        if not leads:
            messagebox.showwarning("Lead Sil", "Lütfen silinecek leadleri seçin.", parent=win)
            return
        if len(leads) == 1:
            prompt = f"{leads[0].get('company_name') or 'seçili lead'} tablodan silinsin mi?"
        else:
            prompt = f"Seçili {len(leads)} lead tablodan silinsin mi?"
        if not messagebox.askyesno("Lead Sil", prompt, parent=win):
            return

        def worker():
            deleted = 0
            failed = 0
            try:
                deleted_ids = set()
                for lead in leads:
                    try:
                        delete_ai_lead(token, lead.get("id"))
                        deleted_ids.add(str(lead.get("id")))
                        deleted += 1
                    except Exception:
                        failed += 1
                state["leads"] = [item for item in state["leads"] if str(item.get("id")) not in deleted_ids]
                win.after(0, apply_filters)
                message = f"{deleted} lead tablodan silindi."
                if failed:
                    message += f" {failed} lead silinemedi."
                win.after(0, lambda msg=message: messagebox.showinfo("Lead Sil", msg, parent=win))
            except Exception as exc:
                win.after(0, lambda err=str(exc): messagebox.showerror("Lead Sil", f"Lead silinemedi: {err}", parent=win))

        threading.Thread(target=worker, daemon=True).start()

    ctk.CTkButton(actions, text="Yenile", width=110, command=load_from_api, fg_color="#ffffff", text_color="#d32f2f", border_width=1, border_color="#d32f2f").pack(side="left", padx=(0, 8))
    ctk.CTkButton(actions, text="Apollo CSV İçe Aktar", width=165, command=import_csv, fg_color="#ffffff", text_color="#2563eb", border_width=1, border_color="#2563eb").pack(side="left", padx=8)
    ctk.CTkButton(actions, text="Segment Ayarları", width=145, command=open_segment_settings, fg_color="#ffffff", text_color="#2563eb", border_width=1, border_color="#2563eb").pack(side="left", padx=8)

    apollo_email_button = ctk.CTkButton(filter_actions, text="Apollo Email Bulucu", width=165, command=enrich_selected, fg_color="#ffffff", text_color="#0f766e", border_width=1, border_color="#0f766e")
    apollo_email_button.pack(side="left", padx=(0, 8))
    hunter_email_button = ctk.CTkButton(filter_actions, text="Hunter Email Bulucu", width=165, command=hunter_enrich_selected, fg_color="#ffffff", text_color="#b45309", border_width=1, border_color="#b45309")
    hunter_email_button.pack(side="left", padx=(0, 8))
    ai_research_button = ctk.CTkButton(filter_actions, text="AI Araştır", width=120, command=research_selected, fg_color="#ffffff", text_color="#7c3aed", border_width=1, border_color="#7c3aed")
    ai_research_button.pack(side="left", padx=(0, 8))
    ctk.CTkButton(filter_actions, text="Sekans Oluştur", width=145, command=create_sequence_for_selected, fg_color="#ffffff", text_color="#0f766e", border_width=1, border_color="#0f766e").pack(side="left")

    bottom_actions = ctk.CTkFrame(root, fg_color="transparent")
    bottom_actions.grid(row=4, column=0, sticky="e", pady=(14, 0))
    ctk.CTkButton(bottom_actions, text="Lead Ekle", width=125, command=add_manual_lead, fg_color="#d32f2f", hover_color="#b91c1c").pack(side="left", padx=8)
    ctk.CTkButton(bottom_actions, text="SerpAPI Firma Bul", width=165, command=segment_search, fg_color="#d32f2f", hover_color="#b91c1c").pack(side="left", padx=8)
    ctk.CTkButton(bottom_actions, text="Hunter Firma Bul", width=155, command=hunter_company_search, fg_color="#d32f2f", hover_color="#b91c1c").pack(side="left", padx=8)
    ctk.CTkButton(bottom_actions, text="Apollo Search", width=130, command=apollo_search, fg_color="#ffffff", text_color="#7c3aed", border_width=1, border_color="#7c3aed").pack(side="left", padx=8)
    ctk.CTkButton(bottom_actions, text="Entegrasyon Durumu", width=165, command=show_provider_status, fg_color="#ffffff", text_color="#334155", border_width=1, border_color="#94a3b8").pack(side="left", padx=8)
    ctk.CTkButton(bottom_actions, text="Sil", width=90, command=delete_selected_lead, fg_color="#ffffff", text_color="#dc2626", border_width=1, border_color="#dc2626").pack(side="left", padx=(8, 0))

    search_var.trace_add("write", apply_filters)
    channel_var.trace_add("write", apply_filters)
    product_var.trace_add("write", apply_filters)
    priority_var.trace_add("write", apply_filters)
    tree.bind("<Double-1>", open_detail)

    apply_filters()
    load_from_api()


def _metric_card(parent, column, title, value_var, color):
    card = ctk.CTkFrame(parent, fg_color="#ffffff", corner_radius=14, border_width=1, border_color="#e5e7eb")
    card.grid(row=0, column=column, sticky="ew", padx=6)
    ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=12), text_color="#64748b").pack(anchor="w", padx=14, pady=(12, 2))
    ctk.CTkLabel(card, textvariable=value_var, font=ctk.CTkFont(size=24, weight="bold"), text_color=color).pack(anchor="w", padx=14, pady=(0, 12))


def _filter_label(parent, text, column):
    ctk.CTkLabel(parent, text=text, font=ctk.CTkFont(size=12, weight="bold"), text_color="#475569").grid(row=0, column=column, sticky="w", padx=(16 if column == 0 else 8, 8), pady=(14, 5))


def _form_entry(parent, label, variable, row):
    ctk.CTkLabel(parent, text=label, text_color="#475569").grid(row=row, column=0, sticky="w", padx=16, pady=8)
    ctk.CTkEntry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=16, pady=8)
    parent.grid_columnconfigure(1, weight=1)


def _add_keyword_tags_from_entry(entry, tags):
    raw_value = entry.get().replace("\n", ",")
    new_tags = [item.strip() for item in raw_value.split(",") if item.strip()]
    for tag in new_tags:
        if tag.casefold() not in {existing.casefold() for existing in tags}:
            tags.append(tag)
    entry.delete(0, "end")


def _keyword_tag_editor(parent, label, tags, row):
    ctk.CTkLabel(parent, text=label, text_color="#475569").grid(row=row, column=0, sticky="nw", padx=16, pady=8)
    wrapper = ctk.CTkFrame(parent, fg_color="#f8fafc", corner_radius=8)
    wrapper.grid(row=row, column=1, sticky="ew", padx=16, pady=8)
    wrapper.grid_columnconfigure(0, weight=1)

    tag_frame = ctk.CTkFrame(wrapper, fg_color="transparent")
    tag_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 4))

    entry = ctk.CTkEntry(wrapper, placeholder_text="Keyword yazıp Enter'a basın. Virgüllü girişler ayrı etikete dönüşür.")
    entry.grid(row=1, column=0, sticky="ew", padx=(8, 6), pady=(4, 8))
    parent.grid_columnconfigure(1, weight=1)

    def render_tags():
        for child in tag_frame.winfo_children():
            child.destroy()
        if not tags:
            ctk.CTkLabel(tag_frame, text="Henüz keyword yok.", text_color="#94a3b8").grid(row=0, column=0, sticky="w", padx=4, pady=4)
            return
        column_count = 2
        for index, tag in enumerate(tags):
            tag_button = ctk.CTkButton(
                tag_frame,
                text=f"{tag}  x",
                command=lambda value=tag: remove_tag(value),
                fg_color="#e0f2fe",
                hover_color="#bae6fd",
                text_color="#075985",
                height=28,
                corner_radius=14,
            )
            tag_button.grid(row=index // column_count, column=index % column_count, sticky="w", padx=4, pady=4)

    def add_tags(_event=None):
        _add_keyword_tags_from_entry(entry, tags)
        render_tags()
        return "break"

    def remove_tag(value):
        tags[:] = [tag for tag in tags if tag != value]
        render_tags()

    ctk.CTkButton(
        wrapper,
        text="Ekle",
        command=add_tags,
        fg_color="#ffffff",
        hover_color="#f1f5f9",
        text_color="#2563eb",
        border_width=1,
        border_color="#2563eb",
        width=70,
    ).grid(row=1, column=1, sticky="e", padx=(0, 8), pady=(4, 8))
    entry.bind("<Return>", add_tags)
    render_tags()
    return entry


def _form_combo(parent, label, variable, values, row):
    ctk.CTkLabel(parent, text=label, text_color="#475569").grid(row=row, column=0, sticky="w", padx=16, pady=8)
    combo = ctk.CTkComboBox(parent, values=values, variable=variable)
    combo.grid(row=row, column=1, sticky="ew", padx=16, pady=8)
    parent.grid_columnconfigure(1, weight=1)
    return combo


def _form_country_selector(parent, label, variable, values, row, owner):
    ctk.CTkLabel(parent, text=label, text_color="#475569").grid(row=row, column=0, sticky="w", padx=16, pady=8)
    field = ctk.CTkFrame(parent, fg_color="transparent")
    field.grid(row=row, column=1, sticky="ew", padx=16, pady=8)
    field.grid_columnconfigure(0, weight=1)
    entry = ctk.CTkEntry(field, textvariable=variable)
    entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))

    def open_picker(_event=None):
        picker = ctk.CTkToplevel(owner)
        picker.title("Ülke Seç")
        picker.geometry("380x460")
        picker.minsize(320, 360)
        picker.configure(fg_color="#f5f5f5")
        picker.transient(owner)
        picker.grab_set()

        frame = ctk.CTkFrame(picker, fg_color="#ffffff", corner_radius=12)
        frame.pack(fill="both", expand=True, padx=14, pady=14)
        ctk.CTkLabel(
            frame,
            text="Baş harfe basarak liste içinde hızlıca gezinebilirsiniz.",
            text_color="#64748b",
            wraplength=330,
        ).pack(anchor="w", padx=12, pady=(12, 8))

        list_frame = ctk.CTkFrame(frame, fg_color="transparent")
        list_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical")
        listbox = tk.Listbox(
            list_frame,
            activestyle="dotbox",
            exportselection=False,
            height=18,
            font=("Segoe UI", 10),
        )
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for country in values:
            listbox.insert("end", country)

        def select_current(_event=None):
            selected = listbox.curselection()
            if not selected:
                return
            variable.set(listbox.get(selected[0]))
            picker.destroy()

        def jump_to_letter(event):
            char = str(event.char or "").casefold()
            if not char or not char.isalpha():
                return
            for index, country in enumerate(values):
                if str(country).casefold().startswith(char):
                    listbox.selection_clear(0, "end")
                    listbox.selection_set(index)
                    listbox.activate(index)
                    listbox.see(index)
                    return "break"
            return None

        current = variable.get()
        if current in values:
            index = values.index(current)
            listbox.selection_set(index)
            listbox.activate(index)
            listbox.see(index)
        listbox.bind("<Double-1>", select_current)
        listbox.bind("<Return>", select_current)
        listbox.bind("<Key>", jump_to_letter)
        picker.bind("<Escape>", lambda _event: picker.destroy())
        picker.after(100, listbox.focus_set)

    ctk.CTkButton(field, text="Seç", width=70, command=open_picker, fg_color="#ffffff", text_color="#2563eb", border_width=1, border_color="#2563eb").grid(row=0, column=1, sticky="e")
    entry.bind("<Button-1>", open_picker)
    parent.grid_columnconfigure(1, weight=1)
    return entry


def _next_id(leads):
    return max([int(item.get("id") or 0) for item in leads] or [0]) + 1


def _read_import_rows(path):
    lowered = str(path or "").casefold()
    if lowered.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("Excel içe aktarma için openpyxl bulunamadı. Dosyayı CSV olarak dışa aktarabilirsiniz.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        output = []
        for values in rows[1:]:
            item = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else ""
                item[header] = "" if value is None else str(value).strip()
            if any(item.values()):
                output.append(item)
        return output
    encodings = ("utf-8-sig", "cp1254", "latin-1")
    last_error = None
    for encoding in encodings:
        try:
            with open(path, newline="", encoding=encoding) as file:
                return [dict(row) for row in csv.DictReader(file) if any((value or "").strip() for value in row.values())]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"CSV encoding okunamadı: {last_error}")


def _summarize_import_rows(rows):
    companies = set()
    emails = 0
    people = 0
    for row in rows:
        company = _first_value(row, ["company_name", "Company", "Company Name", "Company Name for Emails", "company", "Firma"])
        country = _first_value(row, ["country", "Country", "Company Country", "Ülke"])
        if company:
            companies.add(f"{company.casefold()}|{country.casefold()}")
        email_values = [
            _first_value(row, ["email", "Email", "Email Address", "person_email", "contact_email"]),
            _first_value(row, ["Secondary Email"]),
            _first_value(row, ["Tertiary Email"]),
        ]
        emails += sum(1 for value in email_values if value)
        if any(_first_value(row, keys) for keys in [["first_name", "First Name"], ["last_name", "Last Name"], ["name", "Name", "Person Name"], ["title", "Title", "Job Title"]]):
            people += 1
    return {"rows": len(rows), "companies": len(companies), "emails": emails, "people": people}


def _first_value(row, keys):
    for key in keys:
        value = row.get(key)
        if value:
            return str(value).strip()
    return ""


def _guess_channel(text):
    haystack = str(text or "").casefold()
    if any(word in haystack for word in ["integrator", "integration", "epc", "engineering", "automation"]):
        return "System Integration Solution Partner"
    if any(word in haystack for word in ["hvac", "ventilation", "indoor air", "clean air"]):
        return "Clean Air Solution Partner"
    if any(word in haystack for word in ["reseller", "distributor", "supplier"]):
        return "White Label / Resellers"
    if any(word in haystack for word in ["manufacturer", "machine builder", "oem"]):
        return "OEM"
    return "White Label / Resellers"


def _channel_from_apollo_row(row, text):
    segment = _first_value(row, ["segment", "Segment"]).casefold()
    campaign = _first_value(row, ["recommended_campaign", "Recommended Campaign"]).casefold()
    haystack = f"{segment} {campaign} {text}".casefold()
    if any(word in haystack for word in ["system integrator", "integration", "integrator", "epc", "turnkey"]):
        return "System Integration Solution Partner"
    if any(word in haystack for word in ["hvac", "clean air", "ventilation", "industrial_hvac"]):
        return "Clean Air Solution Partner"
    if any(word in haystack for word in ["distributor", "reseller", "dealer", "supplier"]):
        return "White Label / Resellers"
    if any(word in haystack for word in ["oem", "manufacturer", "machine builder"]):
        return "OEM"
    return _guess_channel(text)


def _guess_product(text):
    haystack = str(text or "").casefold()
    if any(word in haystack for word in ["oil mist", "cnc", "machining", "die casting"]):
        return "Oil Mist Filtration"
    if any(word in haystack for word in ["dust", "bulk", "foundry", "smelter", "powder"]):
        return "Dust Collection"
    if any(word in haystack for word in ["fume", "welding", "laser", "plasma", "cutting"]):
        return "Fume Extraction"
    if any(word in haystack for word in ["turnkey", "epc", "plant"]):
        return "Turnkey Solutions"
    return "Hall Ventilation"


def _product_from_apollo_row(row, text):
    campaign = _first_value(row, ["recommended_campaign", "Recommended Campaign"]).casefold()
    haystack = f"{campaign} {text}".casefold()
    if any(word in haystack for word in ["oil_mist", "oil mist", "cnc", "machining"]):
        return "Oil Mist Filtration"
    if any(word in haystack for word in ["dust", "atex", "bulk", "foundry", "powder"]):
        return "Dust Collection"
    if any(word in haystack for word in ["fume", "welding", "cutting", "grinding"]):
        return "Fume Extraction"
    if any(word in haystack for word in ["turnkey", "epc", "plant"]):
        return "Turnkey Solutions"
    if any(word in haystack for word in ["hvac", "ventilation", "clean air"]):
        return "Hall Ventilation"
    return _guess_product(text)


def _priority_from_icp(row):
    return ""


def _score_from_icp(row, priority):
    return 0


def _table_value(item, column):
    if column == "email_sequence_stage":
        value = item.get("email_sequence_stage")
        if value:
            return value
        status = item.get("ai_status")
        email_status = str(item.get("email_status") or "").casefold()
        draft_count = int(item.get("draft_count") or 0)
        if status == "Excluded":
            return "Kapsam Dışı"
        if status == "Review Needed":
            return "Review Needed"
        if draft_count:
            return f"Taslak Oluşturuldu ({draft_count})"
        if not item.get("contact_email"):
            return "Email Bekliyor"
        if email_status not in {"verified", "user managed"}:
            return "Email Doğrulama Bekliyor"
        return "Sekans Hazır"
    return item.get(column, "")


def _sequence_from_campaign(campaign, channel):
    normalized = str(campaign or "").casefold()
    if "hvac" in normalized or "clean" in normalized:
        return "CASP"
    if "integr" in normalized or "turnkey" in normalized:
        return "SISP"
    if "oem" in normalized:
        return "OEM"
    if "reseller" in normalized or "distributor" in normalized:
        return "WL_RESELLER"
    return get_sequence_code(channel)


def _guess_language(country):
    mapping = {
        "Germany": "German",
        "Austria": "German",
        "France": "French",
        "Spain": "Spanish",
        "Italy": "Italian",
        "Netherlands": "Dutch",
        "Belgium": "French/Dutch",
        "Turkey": "Turkish",
    }
    return mapping.get(str(country or "").strip(), "English")
