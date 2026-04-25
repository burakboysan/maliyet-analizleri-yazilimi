# flans_toplu_import.py (Flanş Toplu İçe Aktarma Sistemi)

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from core.database import veritabani_baglanti
from maliyet.cost_calculator import maliyet_hesapla
from decimal import Decimal, InvalidOperation
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import threading

ISCIKLIK_TURLERI = ["Plazma/Lazer", "Makas", "Testere", "Abkant", "Silindir", "Delik Delme", "Kaynak", "Argon", "Montaj", "Boya", "Elektrik", "Ambalaj/Yükleme"]

def flans_sablon_indir():
    dosya_yolu = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")],
        title="Flanş İçe Aktarma Şablonunu Kaydet", initialfile="flans_import_sablonu.xlsx"
    )
    if not dosya_yolu: return

    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Flanş Veri Girişi"

        basliklar = ["Flanş Malzemesi Kodu", "Flanş Çapı (mm)", "Flanş Kalınlığı (mm)", "Flanş Alanı (m²)"]
        for tur in ISCIKLIK_TURLERI:
            basliklar.append(f"{tur} - Usta (saat)")
            basliklar.append(f"{tur} - Yardımcı (saat)")
        ws.append(basliklar)

        bold_font = Font(bold=True)
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = bold_font; cell.fill = fill

        ws2 = workbook.create_sheet(title="Gecerli Malzemeler")
        ws2.append(["Malzeme Kodu", "Malzeme Adı"])
        
        db = veritabani_baglanti()
        cursor = db.cursor()
        cursor.execute("SELECT malzeme_kodu, ad FROM malzemeler WHERE malzeme_tipi = 'Yarı Mamül'")
        for kod, ad in cursor.fetchall():
            ws2.append([kod, ad])
        db.close()
        
        for cell in ws2[1]:
            cell.font = bold_font; cell.fill = fill

        workbook.save(dosya_yolu)
        messagebox.showinfo("Başarılı", f"Flanş şablonu başarıyla kaydedildi:\n{dosya_yolu}")
    except Exception as e:
        messagebox.showerror("Hata", f"Şablon oluşturulurken bir hata oluştu:\n{e}")

class FlansImportEkrani:
    def __init__(self, parent_window, yenileme_fonksiyonu):
        self.pencere = ctk.CTkToplevel(parent_window)
        self.pencere.title("Excel'den Toplu Flanş Aktarımı")
        self.pencere.geometry("1100x700")
        self.pencere.transient(parent_window)
        self.pencere.grab_set()
        
        self.yenileme_fonksiyonu = yenileme_fonksiyonu
        self.onizleme_verisi_df = None
        self.gecerli_malzemeler = set()
        self._verileri_on_yukle()
        self._arayuzu_olustur()

    def _verileri_on_yukle(self):
        try:
            db = veritabani_baglanti()
            cursor = db.cursor()
            cursor.execute("SELECT malzeme_kodu FROM malzemeler WHERE malzeme_tipi = 'Yarı Mamül'")
            self.gecerli_malzemeler = {row[0] for row in cursor.fetchall()}
            db.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Malzeme kodları yüklenemedi: {e}", parent=self.pencere)

    def _arayuzu_olustur(self):
        ust_cerceve = ctk.CTkFrame(self.pencere)
        ust_cerceve.pack(pady=10, padx=10, fill="x")
        ctk.CTkButton(ust_cerceve, text="Flanş Şablonu İndir", command=flans_sablon_indir).pack(side="left", padx=10)
        ctk.CTkButton(ust_cerceve, text="Excel Dosyası Seç", command=self.dosya_sec_ve_onizle).pack(side="left", padx=10)
        self.dosya_yolu_etiketi = ctk.CTkLabel(ust_cerceve, text="Henüz dosya seçilmedi...")
        self.dosya_yolu_etiketi.pack(side="left", padx=10)

        tree_frame = ctk.CTkFrame(self.pencere)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        sutunlar = ("Durum", "Satır No", "Malzeme Kodu", "Çap", "Kalınlık", "Alan", "Hesaplanan Ağırlık", "Hata Mesajı")
        self.tree = ttk.Treeview(tree_frame, columns=sutunlar, show="headings")
        for col in sutunlar: self.tree.heading(col, text=col)
        self.tree.column("Durum", width=80, anchor="center")
        self.tree.column("Satır No", width=60, anchor="center")
        self.tree.column("Hesaplanan Ağırlık", width=120, anchor="center")
        self.tree.column("Hata Mesajı", width=300)
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure('hatali', background='#FF9999')
        self.tree.tag_configure('gecerli', background='#D4EDDA')

        alt_cerceve = ctk.CTkFrame(self.pencere)
        alt_cerceve.pack(pady=10, fill="x")
        
        # Progress bar ve durum etiketi için çerçeve
        progress_frame = ctk.CTkFrame(alt_cerceve, fg_color="transparent")
        progress_frame.pack(fill="x", pady=(0, 10))
        
        self.progress_label = ctk.CTkLabel(progress_frame, text="Hazır", font=ctk.CTkFont(size=12))
        self.progress_label.pack(pady=(0, 5))
        
        self.progress_bar = ctk.CTkProgressBar(progress_frame, mode='determinate')
        self.progress_bar.pack(fill="x", pady=(0, 5))
        self.progress_bar.set(0)
        
        self.kaydet_butonu = ctk.CTkButton(alt_cerceve, text="Doğrulanmış Flanşları İçe Aktar", command=self.veritabanina_kaydet, height=35)
        self.kaydet_butonu.pack()

    def dosya_sec_ve_onizle(self):
        dosya_yolu = filedialog.askopenfilename(filetypes=[("Excel Dosyaları", "*.xlsx")])
        if not dosya_yolu: return
        self.dosya_yolu_etiketi.configure(text=dosya_yolu)
        
        try:
            self.onizleme_verisi_df = pd.read_excel(dosya_yolu, dtype=str).fillna('')
        except Exception as e:
            messagebox.showerror("Okuma Hatası", f"Excel dosyası okunamadı: {e}", parent=self.pencere)
            return

        self.tree.delete(*self.tree.get_children())
        for index, row in self.onizleme_verisi_df.iterrows():
            hata_mesaji = self._satiri_dogrula(row)
            if hata_mesaji:
                self.tree.insert("", "end", values=(
                    "❌ Hatalı", index + 2, 
                    row.get('Flanş Malzemesi Kodu', ''), 
                    row.get('Flanş Çapı (mm)', ''), 
                    row.get('Flanş Kalınlığı (mm)', ''), 
                    row.get('Flanş Alanı (m²)', ''),
                    "0.00 kg",
                    hata_mesaji
                ), tags=('hatali',))
            else:
                # Ağırlık hesapla
                alan = self._to_decimal(row['Flanş Alanı (m²)'])
                kalinlik = self._to_decimal(row['Flanş Kalınlığı (mm)'])
                agirlik = alan * kalinlik * Decimal("8")
                
                self.tree.insert("", "end", values=(
                    "✅ Hazır", index + 2, 
                    row['Flanş Malzemesi Kodu'], 
                    row['Flanş Çapı (mm)'], 
                    row['Flanş Kalınlığı (mm)'], 
                    row['Flanş Alanı (m²)'],
                    f"{agirlik:.2f} kg",
                    "Veri geçerli"
                ), tags=('gecerli',))

    def _satiri_dogrula(self, satir):
        """Her bir Excel satırını kontrol eder ve hata varsa metnini döndürür."""
        
        # 1. Malzeme Kodu Kontrolü
        malzeme_kodu = satir.get('Flanş Malzemesi Kodu', '').strip()
        if not malzeme_kodu:
            return "Flanş Malzemesi Kodu sütunu boş olamaz."
        if malzeme_kodu not in self.gecerli_malzemeler:
            return f"Malzeme Kodu '{malzeme_kodu}' veritabanında bulunamadı."

        # 2. Zorunlu Sayısal Alanların Kontrolü (Çap, Kalınlık, Alan)
        zorunlu_alanlar = ['Flanş Çapı (mm)', 'Flanş Kalınlığı (mm)', 'Flanş Alanı (m²)']
        for alan in zorunlu_alanlar:
            deger_str = satir.get(alan, '').strip()
            if not deger_str:
                return f"'{alan}' sütunu boş bırakılamaz."
            try:
                # Virgülü noktaya çevirerek ondalıklı sayıları kabul et
                deger = Decimal(deger_str.replace(',', '.'))
                if deger <= 0:
                    return f"'{alan}' sütunundaki değer sıfırdan büyük olmalıdır."
            except InvalidOperation:
                return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."

        # 3. Opsiyonel Sayısal Alanların Kontrolü (İşçilik)
        for tur in ISCIKLIK_TURLERI:
            for kisi in ["Usta", "Yardımcı"]:
                alan = f'{tur} - {kisi} (saat)'
                deger_str = satir.get(alan, '0').strip()
                if not deger_str:  # Eğer hücre tamamen boşsa '0' kabul et
                    deger_str = '0'
                try:
                    deger = Decimal(deger_str.replace(',', '.'))
                    if deger < 0:
                        return f"'{alan}' sütunundaki değer negatif olamaz."
                except InvalidOperation:
                    return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."

        # Tüm kontrollerden geçtiyse hata yok demektir.
        return None
    
    def _to_decimal(self, value_str, default='0'):
        # Boşlukları temizler, virgülü noktaya çevirir, boşsa varsayılan değeri kullanır.
        clean_str = str(value_str).strip().replace(',', '.')
        if not clean_str:
            clean_str = default
        return Decimal(clean_str)

    def _kaydetme_islemi(self):      
        # Sadece geçerli olarak etiketlenmiş satırları al
        gecerli_satir_indeksleri = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            if values[0] == "✅ Hazır":
                gecerli_satir_indeksleri.append(int(values[1]) - 2)  # Satır numarasından index'e çevir
        
        if not gecerli_satir_indeksleri:
            messagebox.showwarning("Uyarı", "İçe aktarılacak geçerli veri bulunamadı.", parent=self.pencere)
            return

        gecerli_satirlar_df = self.onizleme_verisi_df.iloc[gecerli_satir_indeksleri]
        toplam_urun_sayisi = len(gecerli_satirlar_df)
        
        # Progress bar'ı başlat
        self.progress_bar.set(0)
        self.progress_label.configure(text=f"0 / {toplam_urun_sayisi} ürün işlendi...")

        db = None
        try:
            db = veritabani_baglanti()
            cursor = db.cursor(dictionary=True, buffered=True) 
            db.autocommit = False
            
            # Bulk insert için veri hazırlama
            urunler_data = []
            urun_agaci_data = []
            iscilik_data = []
            
            aktarilan_sayisi = 0
            for index, row in gecerli_satirlar_df.iterrows():
                # Progress bar'ı güncelle
                progress_orani = (aktarilan_sayisi / toplam_urun_sayisi) * 0.3  # Veri hazırlama %30
                self.progress_bar.set(progress_orani)
                self.progress_label.configure(text=f"Veri hazırlanıyor... {aktarilan_sayisi + 1} / {toplam_urun_sayisi}")
                
                # Güvenli _to_decimal fonksiyonunu kullanıyoruz
                cap_mm = self._to_decimal(row['Flanş Çapı (mm)'])
                kalinlik_mm = self._to_decimal(row['Flanş Kalınlığı (mm)'])
                alan_m2 = self._to_decimal(row['Flanş Alanı (m²)'])
                malzeme_kodu = row['Flanş Malzemesi Kodu']
                
                # Otomatik ürün kodu oluştur (her flanş için benzersiz)
                cursor.execute("SELECT COUNT(*) FROM urunler WHERE urun_kategorisi = 'FLANŞ'")
                mevcut_flans_sayisi = cursor.fetchone()['COUNT(*)']
                flans_sayisi = mevcut_flans_sayisi + aktarilan_sayisi + 1
                
                urun_kodu = f"FLANS-{flans_sayisi:04d}"
                urun_adi = f"Flanş {cap_mm}x{kalinlik_mm}mm ({malzeme_kodu})"
                
                # Ürün verilerini hazırla
                urunler_data.append((urun_kodu, urun_adi, cap_mm, kalinlik_mm))
                
                # Flanş ağırlığı hesapla (alan * kalınlık * 8)
                flans_agirligi = alan_m2 * kalinlik_mm * Decimal("8")
                
                # İşçilik verilerini hazırla
                for tur in ISCIKLIK_TURLERI:
                    usta_saat = self._to_decimal(row.get(f'{tur} - Usta (saat)'))
                    yardimci_saat = self._to_decimal(row.get(f'{tur} - Yardımcı (saat)'))
                    if usta_saat > 0 or yardimci_saat > 0:
                        iscilik_data.append((tur, usta_saat, yardimci_saat, aktarilan_sayisi))
                
                aktarilan_sayisi += 1
            
            # Bulk insert işlemleri
            if urunler_data:
                # Progress bar güncelleme - Veritabanı işlemleri başlıyor
                self.progress_label.configure(text="Veritabanına kaydediliyor...")
                self.progress_bar.set(0.35)  # %35
                
                # Ürünleri toplu ekle
                cursor.executemany(
                    "INSERT INTO urunler (urun_kodu, urun_adi, urun_kategorisi, urun_tipi, kanal_capi, kanal_et_kalinlik, maliyet) VALUES (%s, %s, 'FLANŞ', 'Yarı Mamül', %s, %s, 0)",
                    urunler_data
                )
                
                self.progress_bar.set(0.5)  # %50
                self.progress_label.configure(text="Ürün ağacı oluşturuluyor...")
                
                # Son eklenen ürünlerin ID'lerini al
                cursor.execute("SELECT LAST_INSERT_ID() as first_id")
                first_id = cursor.fetchone()['first_id']
                
                # Ürün ağacı verilerini hazırla ve ekle
                for i, (urun_kodu, urun_adi, cap_mm, kalinlik_mm) in enumerate(urunler_data):
                    urun_id = first_id + i
                    # Malzeme kodunu orijinal DataFrame'den al
                    malzeme_kodu = gecerli_satirlar_df.iloc[i]['Flanş Malzemesi Kodu']
                    
                    # Ağırlık hesaplama
                    alan_m2 = self._to_decimal(gecerli_satirlar_df.iloc[i]['Flanş Alanı (m²)'])
                    flans_agirligi = alan_m2 * kalinlik_mm * Decimal("8")
                    
                    urun_agaci_data.append((urun_id, malzeme_kodu, flans_agirligi))
                
                # Ürün ağacını toplu ekle
                cursor.executemany(
                    "INSERT INTO urun_agaci (urun_id, malzeme_kodu, miktar, malzeme_tipi) VALUES (%s, %s, %s, 'Yarı Mamül')",
                    urun_agaci_data
                )
                
                self.progress_bar.set(0.65)  # %65
                self.progress_label.configure(text="İşçilik verileri ekleniyor...")
                
                # İşçilik verilerini ekle
                for tur, usta_saat, yardimci_saat, index in iscilik_data:
                    urun_id = first_id + index
                    cursor.execute(
                        "INSERT INTO urun_iscilik (urun_id, iscilik_tipi, usta_saat, yardimci_saat) VALUES (%s, %s, %s, %s)",
                        (urun_id, tur, usta_saat, yardimci_saat)
                    )
                
                self.progress_bar.set(0.8)  # %80
                self.progress_label.configure(text="Maliyetler hesaplanıyor...")
                
                # Maliyetleri toplu hesapla
                for i in range(len(urunler_data)):
                    urun_id = first_id + i
                    maliyet_hesapla(urun_id, cursor)
                    
                    # Maliyet hesaplama progress'i
                    progress_orani = 0.8 + (i / len(urunler_data)) * 0.2  # %80-%100 arası
                    self.progress_bar.set(progress_orani)
                    self.progress_label.configure(text=f"Maliyet hesaplanıyor... {i + 1} / {len(urunler_data)}")
            
            db.commit()
            
            # İşlem tamamlandı
            self.progress_bar.set(1.0)  # %100
            self.progress_label.configure(text="İşlem tamamlandı!")
            
            messagebox.showinfo("Başarılı", f"{len(urunler_data)} adet flanş başarıyla içe aktarıldı ve maliyetlendirildi.", parent=self.pencere)
            if self.yenileme_fonksiyonu: self.yenileme_fonksiyonu()
            self.pencere.destroy()
        except Exception as e:
            if db: 
                try:
                    db.rollback()
                except:
                    pass
            messagebox.showerror("Kayıt Hatası", f"Kayıt sırasında bir hata oluştu: {e}", parent=self.pencere)
        finally:
            if db and db.is_connected(): 
                db.autocommit = True
                db.close()
            self.kaydet_butonu.configure(state="normal")
            self.progress_bar.pack_forget()
            self.progress_label.pack_forget()

    def veritabanina_kaydet(self):
        self.kaydet_butonu.configure(state="disabled")
        self.progress_bar.pack(pady=(0, 5))
        self.progress_label.pack(pady=(0, 5))
        # Kaydetme işlemini arayüzün donmaması için bir thread'de başlat
        threading.Thread(target=self._kaydetme_islemi, daemon=True).start()

def flans_import_ekrani(parent_window, yenileme_fonksiyonu):
    FlansImportEkrani(parent_window, yenileme_fonksiyonu) 
