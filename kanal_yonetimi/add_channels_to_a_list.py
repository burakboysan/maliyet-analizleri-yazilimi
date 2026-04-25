import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from decimal import Decimal
from core.database import veritabani_baglanti
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def open_add_channels_to_a_list(parent=None, list_id=None, list_name=None):
    """Belirli bir kanal listesine kanallar ekleme ekranı."""
    window = ctk.CTkToplevel() if parent else ctk.CTk()
    title_suffix = f" - {list_name}" if list_name else ""
    window.title(f"Kanal Listesine Kanal Ekle{title_suffix}")
    try:
        window.state('zoomed')
    except Exception:
        window.geometry("1200x800")
    window.configure(fg_color="#f5f5f5")

    # Ana kapsayıcı - iki frame: Sol | Sağ
    main = ctk.CTkFrame(window)
    main.pack(fill="both", expand=True, padx=20, pady=20)
    main.grid_columnconfigure(0, weight=1, uniform="half")
    main.grid_columnconfigure(1, weight=1, uniform="half")
    main.grid_rowconfigure(0, weight=1)

    # --- SOL FRAME ---
    left = ctk.CTkFrame(main, fg_color="#ffffff", corner_radius=10)
    left.grid(row=0, column=0, padx=(0, 10), pady=0, sticky="nsew")
    left.grid_rowconfigure(2, weight=1)
    left.grid_columnconfigure(0, weight=1)

    # Üst filtre çubuğu
    filter_bar = ctk.CTkFrame(left, fg_color="#f8f9fa")
    filter_bar.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
    ctk.CTkLabel(filter_bar, text="Tip:").pack(side="left", padx=(10, 6))
    type_var = ctk.StringVar(value="Kanal")
    type_combo = ctk.CTkComboBox(
        filter_bar,
        values=["Kanal", "Çatal TE", "Istavroz TE", "Dirsek", "Pantolon", "Adaptör", "Redüksiyon"],
        width=180,
        variable=type_var,
    )
    type_combo.pack(side="left", padx=(0, 10))
    ctk.CTkButton(
        filter_bar,
        text="Git",
        command=lambda: load_products(),
        fg_color="#2196f3",
        hover_color="#1976d2",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left", padx=(0, 8))
    ctk.CTkLabel(filter_bar, text="Ara:").pack(side="left", padx=(4, 6))
    keyword_var = ctk.StringVar()
    keyword_entry = ctk.CTkEntry(filter_bar, width=200, textvariable=keyword_var)
    keyword_entry.pack(side="left")
    try:
        keyword_entry.bind("<Return>", lambda e: load_products())
    except Exception:
        pass
    # Kalınlık filtresi
    ctk.CTkLabel(filter_bar, text="Kalınlık:").pack(side="left", padx=(10, 6))
    kalinlik_var = ctk.StringVar(value="Hepsi")
    kalinlik_combo = ctk.CTkComboBox(
        filter_bar,
        values=["Hepsi", "1", "1.5", "2.0", "3.0", "4.0", "5.0", "6.0", "8.0", "10.0"],
        width=100,
        variable=kalinlik_var,
    )
    kalinlik_combo.pack(side="left", padx=(0, 6))

    # Tablo
    table_wrap = ctk.CTkFrame(left)
    table_wrap.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="nsew")
    table_wrap.grid_rowconfigure(0, weight=1)
    table_wrap.grid_columnconfigure(0, weight=1)

    columns = ("id", "urun_tipi", "dirsek_aci", "kanal_capi", "kanal_et_kalinlik", "kanal_boyu", "flans_durumu", "maliyet")
    tree = ttk.Treeview(
        table_wrap,
        columns=columns,
        displaycolumns=("urun_tipi", "dirsek_aci", "kanal_capi", "kanal_et_kalinlik", "kanal_boyu", "flans_durumu", "maliyet"),
        show="headings",
        selectmode="browse",
    )
    vsb = ttk.Scrollbar(table_wrap, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")

    # Sıralama yardımcıları
    numeric_cols = {"kanal_capi", "kanal_et_kalinlik", "kanal_boyu", "maliyet", "dirsek_aci"}
    def _parse_value(col_name: str, value: str):
        try:
            if col_name in numeric_cols:
                if value is None:
                    return float("inf")
                s = str(value)
                # Para/ölçü birimlerini ve boşlukları temizle
                for ch in ["€", "kg", ","]:
                    s = s.replace(ch, "")
                s = s.strip()
                return float(s) if s != "" else float("inf")
            return (str(value) if value is not None else "").lower()
        except Exception:
            return float("inf") if col_name in numeric_cols else ""
    def _sort_left_tree(col: str, reverse: bool = False):
        try:
            data = [(tree.set(k, col), k) for k in tree.get_children("")]
            data.sort(key=lambda item: _parse_value(col, item[0]), reverse=reverse)
            for idx, (_, k) in enumerate(data):
                tree.move(k, "", idx)
            # Sonraki tıklamada yönü değiştir
            tree.heading(col, command=lambda: _sort_left_tree(col, not reverse))
        except Exception:
            pass
    tree.heading("urun_tipi", text="Tip", command=lambda: _sort_left_tree("urun_tipi", False))
    tree.heading("dirsek_aci", text="Dirsek Açısı", command=lambda: _sort_left_tree("dirsek_aci", False))
    tree.heading("kanal_capi", text="Çap", command=lambda: _sort_left_tree("kanal_capi", False))
    tree.heading("kanal_et_kalinlik", text="Et Kalınlığı", command=lambda: _sort_left_tree("kanal_et_kalinlik", False))
    tree.heading("kanal_boyu", text="Boy", command=lambda: _sort_left_tree("kanal_boyu", False))
    tree.heading("flans_durumu", text="Flanş Durumu", command=lambda: _sort_left_tree("flans_durumu", False))
    tree.heading("maliyet", text="Maliyet (€)", command=lambda: _sort_left_tree("maliyet", False))

    tree.column("urun_tipi", width=140, minwidth=120, anchor="center")
    tree.column("dirsek_aci", width=110, minwidth=90, anchor="center")
    tree.column("kanal_capi", width=90, minwidth=80, anchor="center")
    tree.column("kanal_et_kalinlik", width=110, minwidth=90, anchor="center")
    tree.column("kanal_boyu", width=90, minwidth=80, anchor="center")
    tree.column("flans_durumu", width=120, minwidth=100, anchor="center")
    tree.column("maliyet", width=120, minwidth=100, anchor="center")

    # Alt bar: sağ altta miktar ve ekle
    bottom_left = ctk.CTkFrame(left, fg_color="transparent")
    bottom_left.grid(row=3, column=0, padx=10, pady=(0, 10), sticky="ew")
    ctk.CTkLabel(bottom_left, text="Miktar:").pack(side="left", padx=(0, 6))
    qty_entry = ctk.CTkEntry(bottom_left, width=100)
    qty_entry.pack(side="left", padx=(0, 12))
    ctk.CTkButton(
        bottom_left,
        text="Seçili Kanalı Listeye Ekle",
        command=lambda: on_add_selected(),
        fg_color="#2e7d32",
        hover_color="#1b5e20",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left")

    # --- SAĞ FRAME: Seçilenler Tablosu ---
    right = ctk.CTkFrame(main, fg_color="#ffffff", corner_radius=10)
    right.grid(row=0, column=1, padx=(10, 0), pady=0, sticky="nsew")
    right.grid_rowconfigure(0, weight=1)
    right.grid_columnconfigure(0, weight=1)

    selected_wrap = ctk.CTkFrame(right)
    selected_wrap.grid(row=0, column=0, padx=12, pady=12, sticky="nsew")
    selected_wrap.grid_rowconfigure(0, weight=1)
    selected_wrap.grid_columnconfigure(0, weight=1)

    selected_columns = ("id", "urun_tipi", "dirsek_aci", "kanal_capi", "kanal_et_kalinlik", "kanal_boyu", "flans_durumu", "miktar", "toplam_agirlik", "toplam_maliyet")
    selected_tree = ttk.Treeview(
        selected_wrap,
        columns=selected_columns,
        displaycolumns=("urun_tipi", "dirsek_aci", "kanal_capi", "kanal_et_kalinlik", "kanal_boyu", "flans_durumu", "miktar", "toplam_agirlik", "toplam_maliyet"),
        show="headings",
        selectmode="browse",
    )
    selected_vsb = ttk.Scrollbar(selected_wrap, orient="vertical", command=selected_tree.yview)
    selected_tree.configure(yscrollcommand=selected_vsb.set)
    selected_tree.grid(row=0, column=0, sticky="nsew")
    selected_vsb.grid(row=0, column=1, sticky="ns")

    selected_tree.heading("urun_tipi", text="Tip")
    selected_tree.heading("dirsek_aci", text="Dirsek Açısı")
    selected_tree.heading("kanal_capi", text="Çap")
    selected_tree.heading("kanal_et_kalinlik", text="Et Kalınlığı")
    selected_tree.heading("kanal_boyu", text="Boy")
    selected_tree.heading("flans_durumu", text="Flanş Durumu")
    selected_tree.heading("miktar", text="Miktar")
    selected_tree.heading("toplam_agirlik", text="Toplam Ağırlık")
    selected_tree.heading("toplam_maliyet", text="Toplam Maliyet (€)")

    selected_tree.column("urun_tipi", width=140, minwidth=120, anchor="center")
    selected_tree.column("dirsek_aci", width=110, minwidth=90, anchor="center")
    selected_tree.column("kanal_capi", width=90, minwidth=80, anchor="center")
    selected_tree.column("kanal_et_kalinlik", width=110, minwidth=90, anchor="center")
    selected_tree.column("kanal_boyu", width=90, minwidth=80, anchor="center")
    selected_tree.column("flans_durumu", width=120, minwidth=100, anchor="center")
    selected_tree.column("miktar", width=90, minwidth=80, anchor="center")
    selected_tree.column("toplam_agirlik", width=140, minwidth=120, anchor="center")
    selected_tree.column("toplam_maliyet", width=160, minwidth=120, anchor="center")

    # Veri yükleme yardımcıları
    _columns_cache = None
    staged: dict[int, dict] = {}

    def _get_columns():
        nonlocal _columns_cache
        if _columns_cache is not None:
            return _columns_cache
        try:
            db = veritabani_baglanti()
            cur = db.cursor()
            cur.execute("SHOW COLUMNS FROM urunler")
            _columns_cache = {row[0] for row in cur.fetchall()}
            db.close()
        except Exception:
            _columns_cache = set()
        return _columns_cache

    def load_products():
        # Tablonun içini temizle
        for item in tree.get_children():
            tree.delete(item)
        try:
            selected_type = type_var.get().strip() if type_var.get() else None
            keyword = keyword_var.get().strip() if 'keyword_var' in locals() and keyword_var.get() else None
            selected_kalinlik = kalinlik_var.get().strip() if 'kalinlik_var' in locals() and kalinlik_var.get() else None
            columns = _get_columns()
            db = veritabani_baglanti()
            cur = db.cursor()

            base_select = "SELECT id, urun_adi"
            select_parts = []
            # tip ve dirsek
            if "urun_tipi" in columns:
                select_parts.append(", urun_tipi")
            else:
                select_parts.append(", '' as urun_tipi")
            if "dirsek_aci" in columns:
                select_parts.append(", dirsek_aci")
            else:
                select_parts.append(", '' as dirsek_aci")
            # ölçü ve diğerleri
            if "kanal_capi" in columns:
                select_parts.append(", kanal_capi")
            else:
                select_parts.append(", 0 as kanal_capi")
            if "kanal_et_kalinlik" in columns:
                select_parts.append(", kanal_et_kalinlik")
            else:
                select_parts.append(", 0 as kanal_et_kalinlik")
            if "kanal_boyu" in columns:
                select_parts.append(", kanal_boyu")
            else:
                select_parts.append(", 0 as kanal_boyu")
            if "flans_durumu" in columns:
                select_parts.append(", flans_durumu")
            else:
                select_parts.append(", 'Yok' as flans_durumu")
            if "maliyet" in columns:
                select_parts.append(", maliyet")
            else:
                select_parts.append(", 0 as maliyet")

            select_sql = base_select + "".join(select_parts) + " FROM urunler WHERE urun_kategorisi = %s"
            params = ["KANAL"]
            if selected_type and "urun_tipi" in columns:
                select_sql += " AND urun_tipi = %s"
                params.append(selected_type)
            if keyword:
                select_sql += " AND urun_adi LIKE %s"
                params.append(f"%{keyword}%")
            if selected_kalinlik and selected_kalinlik != "Hepsi" and "kanal_et_kalinlik" in columns:
                select_sql += " AND kanal_et_kalinlik = %s"
                try:
                    params.append(float(selected_kalinlik.replace(',', '.')))
                except Exception:
                    params.append(selected_kalinlik)

            select_sql += " ORDER BY urun_adi LIMIT 1000"

            cur.execute(select_sql, tuple(params))
            rows = cur.fetchall()
            db.close()

            for r in rows:
                # r tuple sırası select_sql ile aynı
                # Sıra: id, urun_adi, urun_tipi, dirsek_aci, cap, et, boy, flans, maliyet
                row_dict = {
                    "id": r[0],
                    "urun_adi": r[1],
                    "urun_tipi": r[2] if len(r) > 2 else "",
                    "dirsek_aci": r[3] if len(r) > 3 else "",
                    "kanal_capi": r[4] if len(r) > 4 else 0,
                    "kanal_et_kalinlik": r[5] if len(r) > 5 else 0,
                    "kanal_boyu": r[6] if len(r) > 6 else 0,
                    "flans_durumu": r[7] if len(r) > 7 else "Yok",
                    "maliyet": r[8] if len(r) > 8 else 0,
                }
                tree.insert(
                    "",
                    "end",
                    values=(
                        row_dict["id"],
                        row_dict["urun_tipi"],
                        row_dict["dirsek_aci"],
                        row_dict["kanal_capi"],
                        row_dict["kanal_et_kalinlik"],
                        row_dict["kanal_boyu"],
                        row_dict["flans_durumu"],
                        f"{float(row_dict['maliyet'] or 0):.2f} €",
                    ),
                    tags=(str(row_dict["id"]),),
                )
                # ID'yi gizli tutuyoruz; seçimde tag'lerden okuyacağız
        except Exception as e:
            messagebox.showerror("Hata", f"Ürün listesi yüklenirken hata: {e}", parent=window)

    def _get_selected_row_id():
        sel = tree.selection()
        if not sel:
            return None
        tags = tree.item(sel[0]).get("tags") or []
        return int(tags[0]) if tags else None

    def on_add_selected():
        sel_id = _get_selected_row_id()
        if not sel_id:
            messagebox.showwarning("Uyarı", "Lütfen tablodan bir kanal seçin.", parent=window)
            return
        try:
            qty_text = (qty_entry.get() or "").strip().replace(",", ".")
            qty = Decimal(qty_text)
            if qty <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning("Uyarı", "Lütfen miktarı 0'dan büyük sayısal bir değer girin.", parent=window)
            return
        try:
            # Ürün detaylarını çek
            db = veritabani_baglanti()
            cur = db.cursor(dictionary=True)
            cur.execute(
                """
                SELECT id,
                       urun_adi,
                       COALESCE(urun_tipi, '') as urun_tipi,
                       COALESCE(dirsek_aci, '') as dirsek_aci,
                       COALESCE(kanal_capi, 0) as kanal_capi,
                       COALESCE(kanal_et_kalinlik, 0) as kanal_et_kalinlik,
                       COALESCE(kanal_boyu, 0) as kanal_boyu,
                       COALESCE(flans_durumu, 'Yok') as flans_durumu,
                       agirlik,
                       COALESCE(maliyet, 0) as maliyet
                FROM urunler
                WHERE id = %s AND urun_kategorisi = 'KANAL'
                """,
                (sel_id,),
            )
            row = cur.fetchone()
            db.close()
            if not row:
                messagebox.showerror("Hata", "Seçilen ürün bulunamadı veya kategori uyumsuz.", parent=window)
                return

            existing = staged.get(sel_id)
            if existing:
                existing["miktar"] = (existing.get("miktar") or Decimal("0")) + qty
            else:
                staged[sel_id] = {
                    "id": row["id"],
                    "urun_adi": row["urun_adi"],
                    "urun_tipi": row.get("urun_tipi"),
                    "dirsek_aci": row.get("dirsek_aci"),
                    "kanal_capi": row["kanal_capi"],
                    "kanal_et_kalinlik": row["kanal_et_kalinlik"],
                    "kanal_boyu": row["kanal_boyu"],
                    "flans_durumu": row["flans_durumu"],
                    "agirlik": row["agirlik"],
                    "maliyet": row["maliyet"],
                    "miktar": qty,
                }
            # Sağ tabloyu yenile
            refresh_selected_table()
        except Exception as e:
            messagebox.showerror("Hata", f"Listeye ekleme sırasında hata: {e}", parent=window)

    def refresh_selected_table():
        for item in selected_tree.get_children():
            selected_tree.delete(item)
        for item in staged.values():
            raw_agirlik = item.get("agirlik")
            try:
                miktar_f = float(item.get("miktar") or 0)
            except Exception:
                miktar_f = 0.0
            if raw_agirlik is None:
                toplam_agirlik_display = "N/A"
            else:
                try:
                    toplam_agirlik_val = float(raw_agirlik) * miktar_f
                except Exception:
                    toplam_agirlik_val = 0.0
                toplam_agirlik_display = f"{toplam_agirlik_val:.2f} kg"
            toplam_maliyet = float(item.get("maliyet") or 0) * float(item.get("miktar") or 0)
            selected_tree.insert(
                "",
                "end",
                values=(
                    item["id"],
                    item.get("urun_tipi") or "",
                    item.get("dirsek_aci") or "",
                    item["kanal_capi"],
                    item["kanal_et_kalinlik"],
                    item["kanal_boyu"],
                    item["flans_durumu"],
                    f"{float(item['miktar']):.2f}",
                    toplam_agirlik_display,
                    f"{toplam_maliyet:.2f} €",
                ),
                tags=(str(item["id"]),),
            )

    # Sağ alt bar: Sol tarafta Sil | Hesapla | Excel'e Aktar, Sağ tarafta Listeyi Kaydet | İptal
    bottom_right = ctk.CTkFrame(right, fg_color="transparent")
    bottom_right.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="ew")
    bottom_right.grid_columnconfigure(0, weight=1)

    left_btns = ctk.CTkFrame(bottom_right, fg_color="transparent")
    left_btns.grid(row=0, column=0, sticky="w")

    right_btns = ctk.CTkFrame(bottom_right, fg_color="transparent")
    right_btns.grid(row=0, column=1, sticky="e")

    def _selected_get_selected_row_id():
        sel = selected_tree.selection()
        if not sel:
            return None
        tags = selected_tree.item(sel[0]).get("tags") or []
        return int(tags[0]) if tags else None

    def on_delete_selected_right():
        sel_id = _selected_get_selected_row_id()
        if not sel_id:
            messagebox.showwarning("Uyarı", "Silmek için sağ listeden bir satır seçin.", parent=window)
            return
        if sel_id in staged:
            staged.pop(sel_id, None)
            refresh_selected_table()

    def on_calculate():
        # Toplamları ve kırılımları hesaplayıp özet göster
        toplam_agirlik = 0.0
        toplam_genel = 0.0
        toplam_malzeme = 0.0
        toplam_iscilik = 0.0
        toplam_uretim = 0.0
        toplam_yonetim = 0.0
        try:
            from maliyet.cost_calculator import maliyet_hesapla as _maliyet_hesapla
        except Exception:
            _maliyet_hesapla = None

        # Ağırlık ve basit toplam
        for it in staged.values():
            try:
                miktar_f = float(it.get("miktar") or 0)
            except Exception:
                miktar_f = 0.0
            toplam_agirlik += float(it.get("agirlik") or 0) * miktar_f
            toplam_genel += float(it.get("maliyet") or 0) * miktar_f

        # Kırılım gerekiyorsa ve hesaplayıcı varsa detayları topla
        if _maliyet_hesapla is not None:
            dbx = None
            curx = None
            try:
                dbx = veritabani_baglanti()
                curx = dbx.cursor(dictionary=True)
                for it in staged.values():
                    try:
                        miktar_f = float(it.get("miktar") or 0)
                        if miktar_f <= 0:
                            continue
                        sonuc = _maliyet_hesapla(int(it["id"]), curx)
                        if not sonuc:
                            continue
                        toplam_malzeme += float(sonuc.get("malzeme maliyeti", sonuc.get("malzeme_maliyeti", 0)) or 0) * miktar_f
                        toplam_iscilik += float(sonuc.get("iscilik_maliyeti", 0) or 0) * miktar_f
                        toplam_uretim += float(sonuc.get("uretim_gideri", 0) or 0) * miktar_f
                        toplam_yonetim += float(sonuc.get("yonetim_gideri", 0) or 0) * miktar_f
                    except Exception:
                        continue
            except Exception:
                pass
            finally:
                try:
                    if curx:
                        curx.close()
                    if dbx:
                        dbx.close()
                except Exception:
                    pass
        # Özet mesajı
        ozet = (
            f"Toplam Ağırlık: {toplam_agirlik:.2f} kg\n"
            f"Genel Toplam: {toplam_genel:.2f} €"
        )
        if _maliyet_hesapla is not None:
            ozet += (
                f"\n\nKırılımlar:\n"
                f"- Malzeme: {toplam_malzeme:.2f} €\n"
                f"- İşçilik: {toplam_iscilik:.2f} €\n"
                f"- Üretim Gideri: {toplam_uretim:.2f} €\n"
                f"- Yönetim Gideri: {toplam_yonetim:.2f} €"
            )
        messagebox.showinfo("Hesaplama", ozet, parent=window)

    def on_export_excel():
        if not staged:
            messagebox.showwarning("Uyarı", "Aktarılacak veri yok.", parent=window)
            return
        # Dosya yolu seç
        default_name = f"Kanal_Listesi{('_' + str(list_name)) if list_name else ''}.xlsx"
        path = filedialog.asksaveasfilename(
            parent=window,
            defaultextension=".xlsx",
            filetypes=[("Excel dosyaları", "*.xlsx")],
            initialfile=default_name,
            title="Excel Dosyasını Kaydet",
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kanal Listesi"

            # Başlık
            ws["A1"] = "KANAL LİSTESİ"
            ws.merge_cells("A1:I1")
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Üst bilgi
            ws["A2"] = "Liste Adı:"; ws["B2"] = list_name or "-"
            ws["A3"] = "Liste ID:"; ws["B3"] = str(list_id or "-")

            # Tablo başlıkları (sağ tablonun yapısı ile uyumlu)
            headers = [
                "Tip", "Dirsek Açısı", "Çap", "Et Kalınlığı", "Boy", "Flanş Durumu",
                "Miktar", "Toplam Ağırlık", "Toplam Maliyet (€)"
            ]
            start_row = 5
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
                )

            toplam_agirlik_sum = 0.0
            toplam_maliyet_sum = 0.0

            # Veriler
            for idx, it in enumerate(staged.values(), 1):
                row = start_row + idx
                try:
                    miktar_f = float(it.get("miktar") or 0)
                except Exception:
                    miktar_f = 0.0
                # Toplam ağırlık: DB'de agirlik yoksa N/A
                raw_agirlik = it.get("agirlik")
                if raw_agirlik is None:
                    toplam_agirlik_disp = "N/A"
                    toplam_agirlik_val = None
                else:
                    try:
                        toplam_agirlik_val = float(raw_agirlik) * miktar_f
                    except Exception:
                        toplam_agirlik_val = 0.0
                    toplam_agirlik_disp = toplam_agirlik_val
                    toplam_agirlik_sum += float(toplam_agirlik_val or 0)

                try:
                    toplam_maliyet_val = float(it.get("maliyet") or 0) * miktar_f
                except Exception:
                    toplam_maliyet_val = 0.0
                toplam_maliyet_sum += toplam_maliyet_val

                values = [
                    it.get("urun_tipi") or "",
                    it.get("dirsek_aci") or "",
                    it.get("kanal_capi") or 0,
                    it.get("kanal_et_kalinlik") or 0,
                    it.get("kanal_boyu") or 0,
                    it.get("flans_durumu") or "",
                    miktar_f,
                    toplam_agirlik_disp,
                    toplam_maliyet_val,
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if col in (7,):
                        cell.number_format = "#,##0.00"
                    if col == 8 and isinstance(v, (int, float)):
                        cell.number_format = "#,##0.00 \"kg\""
                    if col == 9:
                        cell.number_format = "#,##0.00 \"€\""
                    cell.alignment = Alignment(horizontal="center")

            # Toplam satırı
            total_row = start_row + len(staged) + 1
            ws.cell(row=total_row, column=6, value="TOPLAM").font = Font(bold=True)
            if toplam_agirlik_sum > 0:
                cell = ws.cell(row=total_row, column=8, value=toplam_agirlik_sum)
                cell.number_format = "#,##0.00 \"kg\""
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            cell = ws.cell(row=total_row, column=9, value=toplam_maliyet_sum)
            cell.number_format = "#,##0.00 \"€\""
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

            # Sütun genişlikleri
            for column in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column[0].column)
                for c in column:
                    try:
                        max_length = max(max_length, len(str(c.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

            wb.save(path)
            messagebox.showinfo("Başarılı", f"Excel dosyası kaydedildi:\n{path}", parent=window)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel aktarımı sırasında hata: {e}", parent=window)

    def on_save_list():
        if not list_id:
            messagebox.showerror("Hata", "Liste ID bulunamadı. Pencere yanlış çağrılmış olabilir.", parent=window)
            return
        if not staged:
            messagebox.showwarning("Uyarı", "Kaydedilecek öğe yok.", parent=window)
            return
        try:
            db = veritabani_baglanti()
            cur = db.cursor()
            # Mevcut ağaç kayıtlarını temizle
            cur.execute("DELETE FROM urun_agaci WHERE urun_id = %s", (list_id,))
            # Yeni kayıtları ekle
            insert_sql = "INSERT INTO urun_agaci (urun_id, alt_urun_id, miktar, malzeme_tipi) VALUES (%s, %s, %s, 'Yarı Mamül')"
            batch = []
            for it in staged.values():
                try:
                    miktar_f = float(it.get("miktar") or 0)
                except Exception:
                    miktar_f = 0.0
                if miktar_f <= 0:
                    continue
                batch.append((list_id, int(it["id"]), miktar_f))
            if batch:
                cur.executemany(insert_sql, batch)
            # Maliyet kırılımları ve toplamı hesapla
            try:
                from maliyet.cost_calculator import maliyet_hesapla as _maliyet_hesapla
            except Exception:
                _maliyet_hesapla = None

            toplam_malzeme = 0.0
            toplam_iscilik = 0.0
            toplam_uretim = 0.0
            toplam_yonetim = 0.0
            toplam_genel = 0.0

            if _maliyet_hesapla is not None:
                cur_dict = db.cursor(dictionary=True)
                for it in staged.values():
                    try:
                        miktar_f = float(it.get("miktar") or 0)
                        if miktar_f <= 0:
                            continue
                        sonuc = _maliyet_hesapla(int(it["id"]), cur_dict)
                        if not sonuc:
                            continue
                        genel = float(sonuc.get("genel_toplam", 0) or 0)
                        malz = float(sonuc.get("malzeme maliyeti", sonuc.get("malzeme_maliyeti", 0)) or 0)
                        iscilik = float(sonuc.get("iscilik_maliyeti", 0) or 0)
                        uretim = float(sonuc.get("uretim_gideri", 0) or 0)
                        yonetim = float(sonuc.get("yonetim_gideri", 0) or 0)
                        toplam_genel += genel * miktar_f
                        toplam_malzeme += malz * miktar_f
                        toplam_iscilik += iscilik * miktar_f
                        toplam_uretim += uretim * miktar_f
                        toplam_yonetim += yonetim * miktar_f
                    except Exception:
                        continue
                try:
                    cur_dict.close()
                except Exception:
                    pass

            # Uygun kolonlar varsa ürüne yaz
            try:
                cur_cols = db.cursor()
                cur_cols.execute("SHOW COLUMNS FROM urunler")
                cols = {row[0] for row in cur_cols.fetchall()}
                cur_cols.close()
                update_fields = []
                params = []
                # Her durumda maliyet toplamını yazmaya çalış
                update_fields.append("maliyet = %s")
                params.append(toplam_genel)
                if "malzeme_maliyeti" in cols:
                    update_fields.append("malzeme_maliyeti = %s")
                    params.append(toplam_malzeme)
                if "iscilik_maliyeti" in cols:
                    update_fields.append("iscilik_maliyeti = %s")
                    params.append(toplam_iscilik)
                if "uretim_gideri" in cols:
                    update_fields.append("uretim_gideri = %s")
                    params.append(toplam_uretim)
                if "yonetim_gideri" in cols:
                    update_fields.append("yonetim_gideri = %s")
                    params.append(toplam_yonetim)
                if update_fields:
                    sql = f"UPDATE urunler SET {', '.join(update_fields)} WHERE id = %s"
                    params.append(list_id)
                    cur.execute(sql, tuple(params))
            except Exception:
                pass

            db.commit()
            db.close()
            messagebox.showinfo("Başarılı", "Liste ürün ağacı ve maliyetler kaydedildi.", parent=window)
        except Exception as e:
            try:
                db.close()
            except Exception:
                pass
            messagebox.showerror("Hata", f"Kayıt sırasında hata: {e}", parent=window)

    def on_cancel():
        window.destroy()

    ctk.CTkButton(
        left_btns,
        text="Sil",
        command=on_delete_selected_right,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left", padx=(0, 8))
    ctk.CTkButton(
        left_btns,
        text="Hesapla",
        command=on_calculate,
        fg_color="#ff9800",
        hover_color="#f57c00",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left", padx=(0, 8))
    ctk.CTkButton(
        left_btns,
        text="Excel'e Aktar",
        command=on_export_excel,
        fg_color="#4caf50",
        hover_color="#43a047",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left", padx=(0, 8))
    # Liste içe aktarım iptal edildi

    ctk.CTkButton(
        right_btns,
        text="Listeyi Kaydet",
        command=on_save_list,
        fg_color="#2196f3",
        hover_color="#1976d2",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left", padx=(0, 8))
    ctk.CTkButton(
        right_btns,
        text="İptal",
        command=on_cancel,
        fg_color="#757575",
        hover_color="#616161",
        text_color="white",
        corner_radius=8,
        height=32,
    ).pack(side="left")


    # Eğer düzenleme için liste verildiyse sağ tabloyu DB'den doldur
    try:
        if list_id:
            db_init = veritabani_baglanti()
            cur_init = db_init.cursor(dictionary=True)
            cur_init.execute(
                """
                SELECT ua.alt_urun_id as id,
                       u.urun_adi,
                       COALESCE(u.urun_tipi, '') as urun_tipi,
                       COALESCE(u.dirsek_aci, '') as dirsek_aci,
                       COALESCE(u.kanal_capi, 0) as kanal_capi,
                       COALESCE(u.kanal_et_kalinlik, 0) as kanal_et_kalinlik,
                       COALESCE(u.kanal_boyu, 0) as kanal_boyu,
                       COALESCE(u.flans_durumu, 'Yok') as flans_durumu,
                       u.agirlik,
                       COALESCE(u.maliyet, 0) as maliyet,
                       ua.miktar
                FROM urun_agaci ua
                JOIN urunler u ON u.id = ua.alt_urun_id
                WHERE ua.urun_id = %s
                ORDER BY ua.id ASC
                """,
                (list_id,),
            )
            for row in cur_init.fetchall() or []:
                staged[int(row["id"])] = {
                    "id": int(row["id"]),
                    "urun_adi": row["urun_adi"],
                    "urun_tipi": row.get("urun_tipi"),
                    "dirsek_aci": row.get("dirsek_aci"),
                    "kanal_capi": row["kanal_capi"],
                    "kanal_et_kalinlik": row["kanal_et_kalinlik"],
                    "kanal_boyu": row["kanal_boyu"],
                    "flans_durumu": row["flans_durumu"],
                    "agirlik": row.get("agirlik"),
                    "maliyet": row["maliyet"],
                    "miktar": Decimal(str(row["miktar"])) if row.get("miktar") is not None else Decimal("0"),
                }
            try:
                db_init.close()
            except Exception:
                pass
            refresh_selected_table()
    except Exception:
        pass

    # Başlangıçta tabloyu boş bırak; kullanıcı Tip seçip Git'e basınca dolacak
    return window


