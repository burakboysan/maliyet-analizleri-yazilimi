# kanal_toplu_import.py (Tam ve İşlevsel Hali)

import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from core.database import veritabani_baglanti
from maliyet.cost_calculator import maliyet_hesapla
from decimal import Decimal, InvalidOperation
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import threading

ISCIKLIK_TURLERI = ["Plazma/Lazer", "Makas", "Testere", "Abkant", "Silindir", "Delik Delme", "Kaynak", "Argon", "Montaj", "Boya", "Elektrik", "Ambalaj/Yükleme"]

def sablon_indir():
    dosya_yolu = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")],
        title="İçe Aktarma Şablonunu Kaydet", initialfile="kanal_import_sablonu.xlsx"
    )
    if not dosya_yolu: return

    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Kanal Veri Girişi"

        # Yeni şablon başlıkları
        basliklar = [
            "Kanal Malzemesi Kodu",  # Malzeme Kodu (örn. YMM-001)
            "Çap",                   # mm
            "Boy (H)",               # mm
            "Kalınlık",              # mm
            "Alan",                  # m2 (formül)
            "Ağırlık"                # kg (formül)
        ]
        for tur in ISCIKLIK_TURLERI:
            basliklar.append(f"{tur} - Usta (saat)")
            basliklar.append(f"{tur} - Yardımcı (saat)")
        basliklar.append("Kategori")  # urun_kategorisi (KANAL/FLANŞ)
        basliklar.append("Tip")       # urun_tipi
        ws.append(basliklar)

        bold_font = Font(bold=True)
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = bold_font; cell.fill = fill

        # Örnek formüller (2. satır için) – kullanıcı aşağıya çoğaltabilir
        # Alan = PI() * (Çap/1000) * (Boy/1000)
        # Ağırlık = Alan * Kalınlık * 8
        try:
            ws["E2"] = "=PI()*(B2/1000)*(C2/1000)"
            ws["F2"] = "=E2*D2*8"
        except Exception:
            pass

        # Pantolon için ayrı sayfa
        ws_pantolon = workbook.create_sheet(title="Pantolon Veri Girişi")
        basliklar_pantolon = [
            "Kanal Malzemesi Kodu",
            "Çap",
            "Kalınlık",
            "Alan",
            "Ağırlık",
        ]
        for tur in ISCIKLIK_TURLERI:
            basliklar_pantolon.append(f"{tur} - Usta (saat)")
            basliklar_pantolon.append(f"{tur} - Yardımcı (saat)")
        basliklar_pantolon.append("Kategori")
        basliklar_pantolon.append("Tip")
        ws_pantolon.append(basliklar_pantolon)

        for cell in ws_pantolon[1]:
            cell.font = bold_font; cell.fill = fill

        ws2 = workbook.create_sheet(title="Gecerli Malzemeler")
        ws2.append(["Malzeme Kodu", "Malzeme Adı"])
        
        db = veritabani_baglanti()
        cursor = db.cursor()
        cursor.execute("SELECT malzeme_kodu, ad FROM malzemeler WHERE malzeme_tipi = 'Yarı Mamül'")
        for kod, ad in cursor.fetchall():
            ws2.append([kod, ad])
        db.close()
        
        for cell in ws2[1]:
            cell.font = bold_font; cell.fill = fill

        # Dirsek için ayrı sayfa
        ws_dirsek = workbook.create_sheet(title="Dirsek Veri Girişi")
        basliklar_dirsek = [
            "Kanal Malzemesi Kodu",
            "Çap",
            "Kalınlık",
            "Alan",
            "Ağırlık",
            "Dirsek Açısı",
        ]
        for tur in ISCIKLIK_TURLERI:
            basliklar_dirsek.append(f"{tur} - Usta (saat)")
            basliklar_dirsek.append(f"{tur} - Yardımcı (saat)")
        basliklar_dirsek.append("Kategori")
        basliklar_dirsek.append("Tip")
        ws_dirsek.append(basliklar_dirsek)
        for cell in ws_dirsek[1]:
            cell.font = bold_font; cell.fill = fill

        # Adaptör için ayrı sayfa
        ws_adaptor = workbook.create_sheet(title="Adaptör Veri Girişi")
        basliklar_adaptor = [
            "Kanal Malzemesi Kodu",
            "Çap",
            "Kalınlık",
            "Alan",
            "Ağırlık",
            "Boy (H)",
        ]
        for tur in ISCIKLIK_TURLERI:
            basliklar_adaptor.append(f"{tur} - Usta (saat)")
            basliklar_adaptor.append(f"{tur} - Yardımcı (saat)")
        basliklar_adaptor.append("Kategori")
        basliklar_adaptor.append("Tip")
        ws_adaptor.append(basliklar_adaptor)
        for cell in ws_adaptor[1]:
            cell.font = bold_font; cell.fill = fill
        # Örnek formüller (2. satır)
        try:
            ws_adaptor["D2"] = "=PI()*(B2/1000)*(F2/1000)"  # Alan
            ws_adaptor["E2"] = "=D2*C2*8"                    # Ağırlık
        except Exception:
            pass

        # Redüksiyon için ayrı sayfa
        ws_reduksiyon = workbook.create_sheet(title="Redüksiyon Veri Girişi")
        basliklar_reduksiyon = [
            "Kanal Malzemesi Kodu",
            "Çap",
            "Kalınlık",
            "Alan",
            "Ağırlık",
            "Boy (H)",
        ]
        for tur in ISCIKLIK_TURLERI:
            basliklar_reduksiyon.append(f"{tur} - Usta (saat)")
            basliklar_reduksiyon.append(f"{tur} - Yardımcı (saat)")
        basliklar_reduksiyon.append("Kategori")
        basliklar_reduksiyon.append("Tip")
        ws_reduksiyon.append(basliklar_reduksiyon)
        for cell in ws_reduksiyon[1]:
            cell.font = bold_font; cell.fill = fill
        # Örnek formüller (2. satır)
        try:
            ws_reduksiyon["D2"] = "=PI()*(B2/1000)*(F2/1000)"  # Alan
            ws_reduksiyon["E2"] = "=D2*C2*8"                    # Ağırlık
        except Exception:
            pass

        workbook.save(dosya_yolu)
        messagebox.showinfo("Başarılı", f"Şablon başarıyla kaydedildi:\n{dosya_yolu}")
    except Exception as e:
        messagebox.showerror("Hata", f"Şablon oluşturulurken bir hata oluştu:\n{e}")

class KanalImportEkrani:
    def __init__(self, parent_window, yenileme_fonksiyonu):
        self.pencere = ctk.CTkToplevel(parent_window)
        self.pencere.title("Excel'den Toplu Kanal Aktarımı")
        self.pencere.geometry("1100x700")
        self.pencere.transient(parent_window)
        self.pencere.grab_set()
        
        self.yenileme_fonksiyonu = yenileme_fonksiyonu
        self.onizleme_verisi_df = None
        self.gecerli_malzemeler = set()
        self._verileri_on_yukle()
        self._arayuzu_olustur()

    def _verileri_on_yukle(self):
        try:
            db = veritabani_baglanti()
            cursor = db.cursor()
            cursor.execute("SELECT malzeme_kodu FROM malzemeler WHERE malzeme_tipi = 'Yarı Mamül'")
            self.gecerli_malzemeler = {row[0] for row in cursor.fetchall()}
            db.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Malzeme kodları yüklenemedi: {e}", parent=self.pencere)

    def _arayuzu_olustur(self):
        ust_cerceve = ctk.CTkFrame(self.pencere)
        ust_cerceve.pack(pady=10, padx=10, fill="x")
        ctk.CTkButton(ust_cerceve, text="Şablon İndir", command=sablon_indir).pack(side="left", padx=10)
        ctk.CTkButton(ust_cerceve, text="Excel Dosyası Seç", command=self.dosya_sec_ve_onizle).pack(side="left", padx=10)
        self.dosya_yolu_etiketi = ctk.CTkLabel(ust_cerceve, text="Henüz dosya seçilmedi...")
        self.dosya_yolu_etiketi.pack(side="left", padx=10)

        tree_frame = ctk.CTkFrame(self.pencere)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        sutunlar = ("Durum", "Satır No", "Malzeme Kodu", "Çap", "Boy", "Kalınlık", "Alan", "Ağırlık", "Dirsek Açısı", "Kategori", "Tip", "Hata Mesajı")
        self.tree = ttk.Treeview(tree_frame, columns=sutunlar, show="headings")
        for col in sutunlar: self.tree.heading(col, text=col)
        self.tree.column("Durum", width=80, anchor="center"); self.tree.column("Satır No", width=60, anchor="center")
        self.tree.column("Hata Mesajı", width=480)
        self.tree.column("Alan", width=80, anchor="center"); self.tree.column("Ağırlık", width=90, anchor="center")
        self.tree.column("Dirsek Açısı", width=100, anchor="center")
        self.tree.column("Kategori", width=100, anchor="center")
        self.tree.column("Tip", width=120, anchor="center")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure('hatali', background='#FF9999'); self.tree.tag_configure('gecerli', background='#D4EDDA')

        alt_cerceve = ctk.CTkFrame(self.pencere)
        alt_cerceve.pack(pady=10, fill="x")
        self.progress_label = ctk.CTkLabel(alt_cerceve, text="")
        self.progress_bar = ctk.CTkProgressBar(alt_cerceve, mode='determinate')
        self.kaydet_butonu = ctk.CTkButton(alt_cerceve, text="Doğrulanmış Kanalları İçe Aktar", command=self.veritabanina_kaydet, height=35)
        self.kaydet_butonu.pack()

    def dosya_sec_ve_onizle(self):
        dosya_yolu = filedialog.askopenfilename(filetypes=[("Excel Dosyaları", "*.xlsx")])
        if not dosya_yolu: return
        self.dosya_yolu_etiketi.configure(text=dosya_yolu)
        
        try:
            # Çok sayfalı dosyalarda uygun şablonu otomatik seç
            uygun_df = None
            try:
                xls = pd.ExcelFile(dosya_yolu)
                for sheet in xls.sheet_names:
                    df_tmp = pd.read_excel(dosya_yolu, sheet_name=sheet, dtype=str).fillna('')
                    # Geçici atayıp standart kolon isimlerini uygula
                    self.onizleme_verisi_df = df_tmp.copy()
                    self._standart_kolon_isimleri_uygula()
                    kolonlar = set(self.onizleme_verisi_df.columns)
                    temel_kolonlar = {"Kanal Malzemesi Kodu", "Çap", "Kalınlık"}
                    if temel_kolonlar.issubset(kolonlar):
                        # Şart: ya Boy (H) var, ya da Alan ve Ağırlık var
                        if ("Boy (H)" in kolonlar) or ({"Alan", "Ağırlık"}.issubset(kolonlar)):
                            uygun_df = self.onizleme_verisi_df.copy()
                            break
                if uygun_df is None:
                    # İlk sayfayı alıp standartlaştır
                    self.onizleme_verisi_df = pd.read_excel(dosya_yolu, dtype=str).fillna('')
                    self._standart_kolon_isimleri_uygula()
                else:
                    self.onizleme_verisi_df = uygun_df
            except Exception:
                # Tek sayfa okuma fallback
                self.onizleme_verisi_df = pd.read_excel(dosya_yolu, dtype=str).fillna('')
                self._standart_kolon_isimleri_uygula()
        except Exception as e:
            messagebox.showerror("Okuma Hatası", f"Excel dosyası okunamadı: {e}", parent=self.pencere)
            return

        # Geçerlilik ve önizleme
        self.tree.delete(*self.tree.get_children())
        gecerli_listesi = []
        for index, row in self.onizleme_verisi_df.iterrows():
            hata_mesaji = self._satiri_dogrula(row)
            alan_str = row.get('Alan', '').strip()
            agirlik_str = row.get('Ağırlık', '').strip()
            kategori_str = row.get('Kategori', '').strip()
            if hata_mesaji:
                gecerli_listesi.append(False)
                self.tree.insert(
                    "", "end",
                    values=(
                        "❌ Hatalı", index + 2,
                        row.get('Kanal Malzemesi Kodu', ''),
                        row.get('Çap', ''),
                        row.get('Boy (H)', ''),
                        row.get('Kalınlık', ''),
                        alan_str, agirlik_str,
                        row.get('Dirsek Açısı', '').strip(),
                        kategori_str,
                        row.get('Tip', '').strip(),
                        hata_mesaji
                    ),
                    tags=('hatali',)
                )
            else:
                gecerli_listesi.append(True)
                # Alan/Ağırlık boşsa hesapla ve DataFrame'e yaz
                if not alan_str or not agirlik_str:
                    try:
                        cap = self._to_decimal(row.get('Çap', '0'))
                        boy = self._to_decimal(row.get('Boy (H)', '0'))
                        kal = self._to_decimal(row.get('Kalınlık', '0'))
                        alan, agirlik = self._alan_agirlik_hesapla(cap, boy, kal)
                        self.onizleme_verisi_df.at[index, 'Alan'] = f"{alan:.6f}"
                        self.onizleme_verisi_df.at[index, 'Ağırlık'] = f"{agirlik:.3f}"
                        alan_str = f"{alan:.6f}"; agirlik_str = f"{agirlik:.3f}"
                    except Exception:
                        pass
                self.tree.insert(
                    "", "end",
                    values=(
                        "✅ Hazır", index + 2,
                        row.get('Kanal Malzemesi Kodu', ''),
                        row.get('Çap', ''),
                        row.get('Boy (H)', ''),
                        row.get('Kalınlık', ''),
                        alan_str, agirlik_str,
                        row.get('Dirsek Açısı', '').strip(),
                        kategori_str or 'KANAL',
                        row.get('Tip', '').strip() or 'Kanal',
                        "Veri geçerli"
                    ),
                    tags=('gecerli',)
                )

        # DataFrame'e Geçerli kolonu ekle
        try:
            self.onizleme_verisi_df['Geçerli'] = gecerli_listesi
        except Exception:
            pass

    def _satiri_dogrula(self, satir):
        """Her bir Excel satırını kontrol eder ve hata varsa metnini döndürür."""
        
        # 1. Malzeme Kodu Kontrolü
        malzeme_kodu = satir.get('Kanal Malzemesi Kodu', '').strip()
        if not malzeme_kodu:
            return "Kanal Malzemesi Kodu sütunu boş olamaz."
        if malzeme_kodu not in self.gecerli_malzemeler:
            return f"Malzeme Kodu '{malzeme_kodu}' veritabanında bulunamadı."

        # 2. Zorunlu Sayısal Alanların Kontrolü
        # Şart A: Boy (H) var ise -> Çap, Boy (H), Kalınlık zorunlu
        # Şart B: Boy (H) yok veya boş ise -> Çap, Kalınlık, Alan, Ağırlık zorunlu
        cap_str = (satir.get('Çap', '') or '').strip()
        kal_str = (satir.get('Kalınlık', '') or '').strip()
        boy_str = (satir.get('Boy (H)', '') or '').strip()
        alan_str = (satir.get('Alan', '') or '').strip()
        agirlik_str = (satir.get('Ağırlık', '') or '').strip()

        if boy_str:
            # A senaryosu
            for alan, deger_str in [('Çap', cap_str), ('Boy (H)', boy_str), ('Kalınlık', kal_str)]:
                if not deger_str:
                    return f"'{alan}' sütunu boş bırakılamaz."
                try:
                    Decimal(deger_str.replace(',', '.'))
                except InvalidOperation:
                    return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."
            # Alan/Ağırlık doluysa numeric kontrol et (opsiyonel)
            for alan, deger_str in [('Alan', alan_str), ('Ağırlık', agirlik_str)]:
                if deger_str:
                    try:
                        Decimal(deger_str.replace(',', '.'))
                    except InvalidOperation:
                        return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."
        else:
            # B senaryosu
            for alan, deger_str in [('Çap', cap_str), ('Kalınlık', kal_str), ('Alan', alan_str), ('Ağırlık', agirlik_str)]:
                if not deger_str:
                    return f"'{alan}' sütunu boş bırakılamaz."
                try:
                    Decimal(deger_str.replace(',', '.'))
                except InvalidOperation:
                    return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."

        # 3. Dirsek Açısı kontrolü (opsiyonel ama sayı olmalıysa)
        dirsek_aci_str = (satir.get('Dirsek Açısı', '') or '').strip()
        if dirsek_aci_str:
            try:
                Decimal(dirsek_aci_str.replace(',', '.'))
            except InvalidOperation:
                return f"'Dirsek Açısı' sütunundaki değer ('{dirsek_aci_str}') geçerli bir sayı değil."

        # 4. Opsiyonel Sayısal Alanların Kontrolü (İşçilik)
        for tur in ISCIKLIK_TURLERI:
            for kisi in ["Usta", "Yardımcı"]:
                alan = f'{tur} - {kisi} (saat)'
                deger_str = satir.get(alan, '0').strip()
                if not deger_str:  # Eğer hücre tamamen boşsa '0' kabul et
                    deger_str = '0'
                try:
                    Decimal(deger_str.replace(',', '.'))
                except InvalidOperation:
                    return f"'{alan}' sütunundaki değer ('{deger_str}') geçerli bir sayı değil."

        # 5. Kategori kontrolü (varsayılan KANAL). Çatal/Istavroz girilirse KANAL'a eşlenecek
        kategori = (satir.get('Kategori', '') or '').strip()
        if kategori:
            kategori_norm = self._normalize_kategori(kategori)
            if kategori_norm not in ("KANAL", "FLANŞ"):
                return f"'Kategori' değeri geçersiz: '{kategori}'. İzin verilen: KANAL, FLANŞ, Çatal TE Saplama, Istavroz TE Saplama"

        # 6. Tip kontrolü (varsayılan 'Kanal')
        tip = (satir.get('Tip', '') or '').strip()
        if not tip:
            tip = 'Kanal'
        # Çok uzun/boşluklu tipleri sadeleştirme (temel kontrol)
        if len(tip) > 100:
            return "'Tip' değeri çok uzun. 100 karakterden kısa olmalı."

        # Tüm kontrollerden geçtiyse hata yok demektir.
        return None
    
    def _to_decimal(self, value_str, default='0'):
        # Boşlukları temizler, virgülü noktaya çevirir, boşsa varsayılan değeri kullanır.
        clean_str = str(value_str).strip().replace(',', '.')
        if not clean_str:
            clean_str = default
        return Decimal(clean_str)

    def _kaydetme_islemi(self):      
        # Sadece geçerli olarak etiketlenmiş satırları al
        gecerli_satir_indeksleri = self.onizleme_verisi_df[self.onizleme_verisi_df['Geçerli'] == True].index
        gecerli_satirlar_df = self.onizleme_verisi_df.iloc[gecerli_satir_indeksleri]

        db = None
        try:
            db = veritabani_baglanti()
            cursor = db.cursor(dictionary=True, buffered=True) 
            db.autocommit = False
            
            # Bulk insert için veri hazırlama
            urunler_data = []
            urun_agaci_data = []
            iscilik_data = []
            agirlik_listesi = []
            urun_kodu_kumesi = set()
            
            aktarilan_sayisi = 0
            toplam = len(gecerli_satirlar_df)
            if toplam <= 0:
                self.pencere.after(0, lambda: self.progress_label.configure(text="İçe aktarılacak satır bulunamadı."))
            else:
                self.pencere.after(0, lambda: self.progress_label.configure(text=f"0 / {toplam} satır işlendi"))
            for index, row in gecerli_satirlar_df.iterrows():
                # ### DÜZELTME: Güvenli _to_decimal fonksiyonunu kullanıyoruz ###
                cap_mm = self._to_decimal(row.get('Çap', '0'))
                boy_mm = self._to_decimal(row.get('Boy (H)', '0'))
                kalinlik_mm = self._to_decimal(row.get('Kalınlık', '0'))
                malzeme_kodu = row['Kanal Malzemesi Kodu']
                kategori_xls = (row.get('Kategori', '') or '').strip()
                urun_kategorisi = self._normalize_kategori(kategori_xls) or 'KANAL'
                urun_tipi_xls = (row.get('Tip', '') or '').strip() or 'Kanal'
                dirsek_aci = self._to_decimal(row.get('Dirsek Açısı', '0')) if row.get('Dirsek Açısı', '') else None
                
                temel_urun_kodu = f"KANAL-{malzeme_kodu}-{cap_mm}x{kalinlik_mm}x{boy_mm}"
                urun_kodu = self._benzersiz_urun_kodu_uret(temel_urun_kodu, cursor, urun_kodu_kumesi)
                if dirsek_aci is not None and dirsek_aci != 0:
                    urun_adi = f"Dirsek {cap_mm}x{kalinlik_mm} A={dirsek_aci}° ({malzeme_kodu})"
                else:
                    urun_adi = f"Kanal {cap_mm}x{kalinlik_mm} L={boy_mm}mm ({malzeme_kodu})"
                
                # Ürün verilerini hazırla
                # Dirsek ürünleri için 'dirsek_aci' kolonuna da yazacağız (diğerlerinde NULL bırakılır)
                urunler_data.append((urun_kodu, urun_adi, urun_kategorisi, urun_tipi_xls, cap_mm, boy_mm, kalinlik_mm, dirsek_aci))
                
                # Alan/Ağırlık (Excel'den ya da hesapla) + Tip'e bağlı ağırlık kuralı
                alan_df = (row.get('Alan', '') or '').strip()
                agirlik_df = (row.get('Ağırlık', '') or '').strip()
                # Alanı al: varsa Excel, yoksa hesap
                try:
                    kanal_alani = self._to_decimal(alan_df) if alan_df else None
                except Exception:
                    kanal_alani = None
                if kanal_alani is None:
                    # Boy varsa standart formüle göre alanı hesapla; yoksa 0
                    try:
                        kanal_alani, _tmp_agirlik = self._alan_agirlik_hesapla(cap_mm, boy_mm, kalinlik_mm)
                    except Exception:
                        kanal_alani = Decimal("0")

                # Tip bazlı zorunlu ağırlık hesabı
                tip_norm = (urun_tipi_xls or 'Kanal').strip().lower()
                tip_zorunlu_kume = {"çatal te", "catal te", "ıstavroz te", "istavroz te", "dirsek", "pantolon", "adaptör", "adaptor", "redüksiyon", "reduksiyon"}
                if tip_norm in tip_zorunlu_kume:
                    kanal_agirligi = kanal_alani * kalinlik_mm * Decimal("8")
                else:
                    # Diğer tiplerde Excel ağırlığı varsa kullan, yoksa formül
                    if agirlik_df:
                        try:
                            kanal_agirligi = self._to_decimal(agirlik_df)
                        except Exception:
                            kanal_agirligi = kanal_alani * kalinlik_mm * Decimal("8")
                    else:
                        kanal_agirligi = kanal_alani * kalinlik_mm * Decimal("8")
                
                # İşçilik verilerini hazırla
                for tur in ISCIKLIK_TURLERI:
                    usta_saat = self._to_decimal(row.get(f'{tur} - Usta (saat)'))
                    yardimci_saat = self._to_decimal(row.get(f'{tur} - Yardımcı (saat)'))
                    if usta_saat > 0 or yardimci_saat > 0:
                        iscilik_data.append((tur, usta_saat, yardimci_saat, aktarilan_sayisi))

                # İleride ürün ağacına yazmak için ağırlığı sakla
                agirlik_listesi.append(kanal_agirligi)
                urun_kodu_kumesi.add(urun_kodu)
                
                aktarilan_sayisi += 1
                # Progress güncelle
                oran = aktarilan_sayisi / toplam if toplam else 1
                # Faz 1: Satır hazırlama 0-60%
                self.pencere.after(0, lambda s=aktarilan_sayisi, t=toplam, o=oran: (
                    self.progress_bar.set(0.6 * o),
                    self.progress_label.configure(text=f"{s} / {t} satır işlendi (hazırlanıyor)")
                ))
            
            # Bulk insert işlemleri
            if urunler_data:
                # Ürünleri toplu ekle
                # Not: Dirsek ürünleri için 'dirsek_aci' değeri dolu, diğer tiplerde NULL geçilir
                cursor.executemany(
                    "INSERT INTO urunler (urun_kodu, urun_adi, urun_kategorisi, urun_tipi, kanal_capi, kanal_boyu, kanal_et_kalinlik, dirsek_aci, maliyet) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0)",
                    urunler_data
                )
                # Faz 2: Ürün ekleme 70%
                self.pencere.after(0, lambda: (
                    self.progress_bar.set(0.7),
                    self.progress_label.configure(text="Ürünler veritabanına yazıldı (70%)")
                ))
                
                # Son eklenen ürünlerin ID'lerini al
                cursor.execute("SELECT LAST_INSERT_ID() as first_id")
                first_id = cursor.fetchone()['first_id']
                
                # Ürün ağacı verilerini hazırla ve ekle
                for i, (urun_kodu, urun_adi, urun_kategorisi, urun_tipi, cap_mm, boy_mm, kalinlik_mm, _dirsek_aci) in enumerate(urunler_data):
                    urun_id = first_id + i
                    malzeme_kodu = urun_kodu.split('-')[1]  # KANAL-YMM-001-100x1x1000 -> YMM-001

                    # Önceden hesaplanmış ağırlığı kullan
                    kanal_agirligi_final = agirlik_listesi[i] if i < len(agirlik_listesi) else Decimal("0")
                    urun_agaci_data.append((urun_id, malzeme_kodu, kanal_agirligi_final))
                
                # Ürün ağacını toplu ekle
                cursor.executemany(
                    "INSERT INTO urun_agaci (urun_id, malzeme_kodu, miktar, malzeme_tipi) VALUES (%s, %s, %s, 'Yarı Mamül')",
                    urun_agaci_data
                )
                # Faz 3: Ürün ağacı 80%
                self.pencere.after(0, lambda: (
                    self.progress_bar.set(0.8),
                    self.progress_label.configure(text="Ürün ağacı oluşturuluyor (80%)")
                ))
                
                # İşçilik verilerini ekle
                is_toplam = len(iscilik_data) or 1
                for i_is, (tur, usta_saat, yardimci_saat, index) in enumerate(iscilik_data):
                    urun_id = first_id + index
                    cursor.execute(
                        "INSERT INTO urun_iscilik (urun_id, iscilik_tipi, usta_saat, yardimci_saat) VALUES (%s, %s, %s, %s)",
                        (urun_id, tur, usta_saat, yardimci_saat)
                    )
                    # Faz 4: İşçilik 80-90%
                    ilerleme_is = 0.8 + (0.1 * ((i_is + 1) / is_toplam))
                    if (i_is + 1) % 10 == 0 or (i_is + 1) == is_toplam:
                        self.pencere.after(0, lambda ii=i_is+1, it=is_toplam, ip=ilerleme_is: (
                            self.progress_bar.set(ip),
                            self.progress_label.configure(text=f"İşçilik yazılıyor ({ii}/{it})")
                        ))
                
                # Maliyetleri toplu hesapla
                u_toplam = len(urunler_data) or 1
                for i in range(len(urunler_data)):
                    urun_id = first_id + i
                    maliyet_hesapla(urun_id, cursor)
                    # Faz 5: Maliyet 90-99%
                    ilerleme_m = 0.9 + (0.09 * ((i + 1) / u_toplam))
                    if (i + 1) % 5 == 0 or (i + 1) == u_toplam:
                        self.pencere.after(0, lambda ii=i+1, ut=u_toplam, ip=ilerleme_m: (
                            self.progress_bar.set(ip),
                            self.progress_label.configure(text=f"Maliyet hesaplanıyor ({ii}/{ut})")
                        ))
            
            db.commit()
            # Faz 6: Tamamlandı 100%
            self.pencere.after(0, lambda: (
                self.progress_bar.set(1.0),
                self.progress_label.configure(text="İçe aktarma tamamlandı (100%)")
            ))
            messagebox.showinfo("Başarılı", f"{len(urunler_data)} adet kanal başarıyla içe aktarıldı ve maliyetlendirildi.", parent=self.pencere)
            if self.yenileme_fonksiyonu: self.yenileme_fonksiyonu()
            self.pencere.destroy()
        except Exception as e:
            if db: 
                try:
                    db.rollback()
                except:
                    pass
            messagebox.showerror("Kayıt Hatası", f"Kayıt sırasında bir hata oluştu: {e}", parent=self.pencere)
        finally:
            if db and db.is_connected(): 
                db.autocommit = True
                db.close()
            self.kaydet_butonu.configure(state="normal")
            self.progress_bar.pack_forget()
            self.progress_label.configure(text="")

    def _standart_kolon_isimleri_uygula(self):
        """Excel'den gelen değişken başlıkları standart başlıklara dönüştürür."""
        kolon_harita = {
            # Malzeme kodu
            'Malzeme Kodu': 'Kanal Malzemesi Kodu',
            'Kanal Malzemesi Kodu': 'Kanal Malzemesi Kodu',
            # Ölçüler
            'Kanal Çapı (mm)': 'Çap', 'Çap (mm)': 'Çap', 'Çap': 'Çap',
            'Kanal Boyu (mm)': 'Boy (H)', 'Boy': 'Boy (H)', 'Boy (H) (mm)': 'Boy (H)', 'Boy (H)': 'Boy (H)',
            'Kanal Kalınlığı (mm)': 'Kalınlık', 'Et Kalınlığı': 'Kalınlık', 'Kalınlık (mm)': 'Kalınlık', 'Kalınlık': 'Kalınlık',
            # Alan / Ağırlık
            'Alan (m2)': 'Alan', 'Kanal Alanı': 'Alan', 'Alan': 'Alan',
            'Ağırlık (kg)': 'Ağırlık', 'Kanal Ağırlığı': 'Ağırlık', 'Ağırlık': 'Ağırlık', 'Agirlik': 'Ağırlık', 'agirlik': 'Ağırlık',
            # Kategori
            'Kategori': 'Kategori', 'urun_kategorisi': 'Kategori',
            # Tip
            'Tip': 'Tip', 'urun_tipi': 'Tip'
        }
        mevcut_kolonlar = list(self.onizleme_verisi_df.columns)
        yeni_isimler = {}
        for k in mevcut_kolonlar:
            if k in kolon_harita:
                yeni_isimler[k] = kolon_harita[k]
        if yeni_isimler:
            self.onizleme_verisi_df.rename(columns=yeni_isimler, inplace=True)

        # Zorunlu temel kolonlar yoksa boş olarak ekle
        for zorunlu in ["Kanal Malzemesi Kodu", "Çap", "Boy (H)", "Kalınlık"]:
            if zorunlu not in self.onizleme_verisi_df.columns:
                self.onizleme_verisi_df[zorunlu] = ''
        # Opsiyoneller
        for ops in ["Alan", "Ağırlık", "Kategori"]:
            if ops not in self.onizleme_verisi_df.columns:
                self.onizleme_verisi_df[ops] = ''

    def _alan_agirlik_hesapla(self, cap_mm: Decimal, boy_mm: Decimal, kalinlik_mm: Decimal):
        """Alan (m2) ve ağırlık (kg) hesaplar. Projede kullanılan formül ile uyumlu."""
        alan = Decimal("3.14") * (cap_mm/Decimal("1000")) * (boy_mm/Decimal("1000"))
        agirlik = alan * kalinlik_mm * Decimal("8")
        return alan, agirlik

    def _normalize_kategori(self, kategori: str):
        if not kategori:
            return ''
        k = kategori.strip().upper()
        if k in ("KANAL", "FLANŞ", "FLANS"):
            return "FLANŞ" if k in ("FLANS",) else k
        # Yeni tablar için girilen değerleri KANAL'a eşle
        if k in ("ÇATAL TE SAPLAMA", "CATAL TE SAPLAMA", "ISTAVROZ TE SAPLAMA", "İSTAVROZ TE SAPLAMA"):
            return "KANAL"
        return kategori  # geçersiz ise orijinali döndür (doğrulama yakalar)

    def _benzersiz_urun_kodu_uret(self, temel_kod: str, cursor, batch_set: set):
        """Veritabanı ve batch içinde çakışmayacak benzersiz ürün kodu üretir."""
        aday = temel_kod
        ek_sayi = 1
        while True:
            if aday not in batch_set:
                # Veri tabanında var mı?
                try:
                    cursor.execute("SELECT COUNT(*) as c FROM urunler WHERE urun_kodu = %s", (aday,))
                    sonuc = cursor.fetchone()
                    varsa = sonuc['c'] if isinstance(sonuc, dict) else (sonuc[0] if sonuc else 0)
                except Exception:
                    varsa = 0
                if not varsa:
                    return aday
            aday = f"{temel_kod}-{ek_sayi}"
            ek_sayi += 1

    def veritabanina_kaydet(self):
        self.kaydet_butonu.configure(state="disabled")
        self.progress_label.configure(text="İçe aktarıma hazırlanıyor...")
        self.progress_label.pack()
        self.progress_bar.set(0)
        self.progress_bar.pack(pady=(0, 10))
        # Kaydetme işlemini arayüzün donmaması için bir thread'de başlat
        threading.Thread(target=self._kaydetme_islemi, daemon=True).start()

def kanal_import_ekrani(parent_window, yenileme_fonksiyonu):
    KanalImportEkrani(parent_window, yenileme_fonksiyonu)
