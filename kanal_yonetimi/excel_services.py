from __future__ import annotations

from typing import Callable, Dict, Any, List
from decimal import Decimal
import math
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.database import veritabani_baglanti


ProgressCallback = Callable[[float, str], None]


def export_teklif_kanal_listesi(
    teklif_kodu: str,
    save_path: str,
    progress_cb: ProgressCallback | None = None,
) -> None:
    """Verilen teklif kodu için kanal listesini Excel'e yazar.

    Ağır DB sorguları ve Excel oluşturma işlemlerini burada yapar.
    UI, dosya yolu seçimini ve progress penceresini yönetir.
    """

    def update_progress(value: float, text: str) -> None:
        if progress_cb is not None:
            try:
                progress_cb(value, text)
            except Exception:
                # UI progress callback'inde olası hataları sessiz geç
                pass

    update_progress(0.1, "Veritabanına bağlanılıyor...")

    db = veritabani_baglanti()
    cursor_dict = db.cursor(dictionary=True, buffered=True)

    update_progress(0.2, "Teklif bilgileri alınıyor...")

    # Teklif bilgileri
    cursor_dict.execute(
        """
        SELECT teklif_adi, proje_referans_no, proje_kodu, durumu, notlar
        FROM teklifler WHERE teklif_kodu = %s
        """,
        (teklif_kodu,),
    )
    teklif_bilgi: Dict[str, Any] | None = cursor_dict.fetchone()

    update_progress(0.3, "Kanal listesi alınıyor...")

    # Kolon varlık kontrolleri
    cursor_dict.execute("SHOW COLUMNS FROM urunler LIKE 'flans_durumu'")
    flans_durumu_kolon_var = cursor_dict.fetchone()
    cursor_dict.execute("SHOW COLUMNS FROM urunler LIKE 'maliyet'")
    maliyet_kolon_var = cursor_dict.fetchone()

    # Teklif kanal detayları
    if flans_durumu_kolon_var and maliyet_kolon_var:
        cursor_dict.execute(
            """
            SELECT 
                tkd.id,
                tkd.urun_id,
                u.urun_adi,
                u.kanal_capi,
                u.kanal_et_kalinlik,
                u.kanal_boyu,
                COALESCE(u.flans_durumu, 'Yok') as flans_durumu,
                tkd.miktar,
                tkd.birim_maliyet,
                tkd.toplam_maliyet,
                tkd.ekleme_tarihi
            FROM teklif_kanal_detaylari tkd
            JOIN urunler u ON tkd.urun_id = u.id
            WHERE tkd.teklif_kodu = %s
            ORDER BY tkd.ekleme_tarihi DESC
            """,
            (teklif_kodu,),
        )
    elif flans_durumu_kolon_var and not maliyet_kolon_var:
        cursor_dict.execute(
            """
            SELECT 
                tkd.id,
                tkd.urun_id,
                u.urun_adi,
                u.kanal_capi,
                u.kanal_et_kalinlik,
                u.kanal_boyu,
                COALESCE(u.flans_durumu, 'Yok') as flans_durumu,
                tkd.miktar,
                tkd.birim_maliyet,
                tkd.toplam_maliyet,
                tkd.ekleme_tarihi
            FROM teklif_kanal_detaylari tkd
            JOIN urunler u ON tkd.urun_id = u.id
            WHERE tkd.teklif_kodu = %s
            ORDER BY tkd.ekleme_tarihi DESC
            """,
            (teklif_kodu,),
        )
    elif not flans_durumu_kolon_var and maliyet_kolon_var:
        cursor_dict.execute(
            """
            SELECT 
                tkd.id,
                tkd.urun_id,
                u.urun_adi,
                u.kanal_capi,
                u.kanal_et_kalinlik,
                u.kanal_boyu,
                'Yok' as flans_durumu,
                tkd.miktar,
                tkd.birim_maliyet,
                tkd.toplam_maliyet,
                tkd.ekleme_tarihi
            FROM teklif_kanal_detaylari tkd
            JOIN urunler u ON tkd.urun_id = u.id
            WHERE tkd.teklif_kodu = %s
            ORDER BY tkd.ekleme_tarihi DESC
            """,
            (teklif_kodu,),
        )
    else:
        cursor_dict.execute(
            """
            SELECT 
                tkd.id,
                tkd.urun_id,
                u.urun_adi,
                u.kanal_capi,
                u.kanal_et_kalinlik,
                u.kanal_boyu,
                'Yok' as flans_durumu,
                tkd.miktar,
                tkd.birim_maliyet,
                tkd.toplam_maliyet,
                tkd.ekleme_tarihi
            FROM teklif_kanal_detaylari tkd
            JOIN urunler u ON tkd.urun_id = u.id
            WHERE tkd.teklif_kodu = %s
            ORDER BY tkd.ekleme_tarihi DESC
            """,
            (teklif_kodu,),
        )

    kanallar: List[Dict[str, Any]] = cursor_dict.fetchall() or []

    update_progress(0.4, "Excel dosyası oluşturuluyor...")

    # Excel çalışma kitabı
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teklif Kanal Listesi"

    # Stilller
    baslik_font = Font(bold=True, size=12, color="FFFFFF")
    baslik_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alt_baslik_font = Font(bold=True, size=11)
    alt_baslik_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
    )

    update_progress(0.5, "Teklif bilgileri yazılıyor...")

    # Teklif bilgileri
    ws['A1'] = "TEKLİF BİLGİLERİ"
    ws['A1'].font = baslik_font
    ws['A1'].fill = baslik_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')

    ws['A2'] = "Teklif Kodu:"
    ws['B2'] = teklif_kodu
    ws['A3'] = "Teklif Adı:"
    ws['B3'] = teklif_bilgi['teklif_adi'] if teklif_bilgi else "-"
    ws['A4'] = "Proje Referans No:"
    ws['B4'] = teklif_bilgi['proje_referans_no'] if teklif_bilgi else "-"
    ws['A5'] = "Proje Kodu:"
    ws['B5'] = teklif_bilgi['proje_kodu'] if teklif_bilgi else "-"
    ws['A6'] = "Durum:"
    ws['B6'] = teklif_bilgi['durumu'] if teklif_bilgi else "-"
    ws['A7'] = "Notlar:"
    ws['B7'] = teklif_bilgi['notlar'] if teklif_bilgi else "-"

    # Teklif bilgileri stil
    for row in range(2, 8):
        ws[f'A{row}'].font = alt_baslik_font
        ws[f'A{row}'].fill = alt_baslik_fill
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border

    update_progress(0.6, "Kanal listesi oluşturuluyor...")

    # Kanal başlıkları
    baslik_satiri = 10
    basliklar = [
        "Sıra No",
        "Kanal Adı",
        "Kanal Çapı (mm)",
        "Kanal Kalınlığı (mm)",
        "Kanal Boyu (mm)",
        "Flanş Durumu",
        "Miktar (adet)",
        "Birim Maliyet (€)",
        "Toplam Maliyet (€)",
        "Toplam Ağırlık (kg)",
        "Ekleme Tarihi",
    ]
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=baslik_satiri, column=col)
        cell.value = baslik
        cell.font = baslik_font
        cell.fill = baslik_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    toplam_maliyet = Decimal('0')
    toplam_agirlik = Decimal('0')

    for i, kanal in enumerate(kanallar):
        progress_value = 0.6 + (0.2 * (i / max(len(kanallar), 1)))
        update_progress(progress_value, f"Kanal verileri yazılıyor... ({i+1}/{len(kanallar)})")

        row = baslik_satiri + 1 + i

        if kanal['kanal_capi'] and kanal['kanal_et_kalinlik'] and kanal['kanal_boyu']:
            capi_mm = float(kanal['kanal_capi'])
            kalinlik_mm = float(kanal['kanal_et_kalinlik'])
            boy_mm = float(kanal['kanal_boyu'])
            if capi_mm > 0 and kalinlik_mm > 0 and boy_mm > 0:
                birim_agirlik_kg = (math.pi * capi_mm * kalinlik_mm * boy_mm * 0.00785) / 1000
                toplam_agirlik_kanal = birim_agirlik_kg * float(kanal['miktar'])
            else:
                toplam_agirlik_kanal = 0
        else:
            toplam_agirlik_kanal = 0

        ws.cell(row=row, column=1, value=row - baslik_satiri).border = border
        ws.cell(row=row, column=2, value=kanal['urun_adi']).border = border
        ws.cell(row=row, column=3, value=kanal['kanal_capi']).border = border
        ws.cell(row=row, column=4, value=kanal['kanal_et_kalinlik']).border = border
        ws.cell(row=row, column=5, value=kanal['kanal_boyu']).border = border
        ws.cell(row=row, column=6, value=kanal['flans_durumu']).border = border
        ws.cell(row=row, column=7, value=kanal['miktar']).border = border
        ws.cell(row=row, column=8, value=float(kanal['birim_maliyet'])).border = border
        ws.cell(row=row, column=9, value=float(kanal['toplam_maliyet'])).border = border
        ws.cell(row=row, column=10, value=float(toplam_agirlik_kanal)).border = border
        ws.cell(row=row, column=11, value=kanal['ekleme_tarihi']).border = border

        ws.cell(row=row, column=8).number_format = '#,##0.00 €'
        ws.cell(row=row, column=9).number_format = '#,##0.00 €'
        ws.cell(row=row, column=10).number_format = '#,##0.00 kg'

        toplam_maliyet += Decimal(str(kanal['toplam_maliyet']))
        toplam_agirlik += Decimal(str(toplam_agirlik_kanal))

    toplam_satiri = baslik_satiri + len(kanallar) + 1
    ws.cell(row=toplam_satiri, column=1, value="TOPLAM").font = alt_baslik_font
    ws.cell(row=toplam_satiri, column=1).fill = alt_baslik_fill
    ws.cell(row=toplam_satiri, column=1).border = border

    ws.cell(row=toplam_satiri, column=9, value=float(toplam_maliyet)).font = alt_baslik_font
    ws.cell(row=toplam_satiri, column=9).fill = alt_baslik_fill
    ws.cell(row=toplam_satiri, column=9).border = border
    ws.cell(row=toplam_satiri, column=9).number_format = '#,##0.00 €'

    ws.cell(row=toplam_satiri, column=10, value=float(toplam_agirlik)).font = alt_baslik_font
    ws.cell(row=toplam_satiri, column=10).fill = alt_baslik_fill
    ws.cell(row=toplam_satiri, column=10).border = border
    ws.cell(row=toplam_satiri, column=10).number_format = '#,##0.00 kg'

    update_progress(0.9, "Sütun genişlikleri ayarlanıyor...")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    update_progress(0.95, "Dosya kaydediliyor...")
    wb.save(save_path)
    db.close()
    update_progress(1.0, "✅ Aktarım tamamlandı!")


