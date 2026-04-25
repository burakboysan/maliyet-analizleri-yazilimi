import math
from pathlib import Path
from typing import Optional

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox, ttk

from core.utils import apply_bomaksan_table_style, apply_zebra_striping
from core.window_utils import open_window_zoomed


SABIT_TABLOSU: dict[int, dict[float, float]] = {
    20: {0.75: 0.1674, 1.0: 0.1302, 1.5: 0.1054, 2.0: 0.1023},
    30: {0.75: 0.2430, 1.0: 0.1890, 1.5: 0.1530, 2.0: 0.1485},
    45: {0.75: 0.3240, 1.0: 0.2520, 1.5: 0.2040, 2.0: 0.1980},
    60: {0.75: 0.4212, 1.0: 0.3276, 1.5: 0.2652, 2.0: 0.2574},
    75: {0.75: 0.4860, 1.0: 0.3780, 1.5: 0.3060, 2.0: 0.2970},
    90: {0.75: 0.5400, 1.0: 0.4200, 1.5: 0.3400, 2.0: 0.3300},
    110: {0.75: 0.6102, 1.0: 0.4746, 1.5: 0.3842, 2.0: 0.3729},
    130: {0.75: 0.6480, 1.0: 0.5040, 1.5: 0.4080, 2.0: 0.3960},
    150: {0.75: 0.6912, 1.0: 0.5376, 1.5: 0.4352, 2.0: 0.4224},
    180: {0.75: 0.7560, 1.0: 0.5880, 1.5: 0.4760, 2.0: 0.4620},
}

TEMPLATE_FILENAME = "basinc_kaybi_taslagi.xlsx"
TEMPLATE_MAX_ROWS = 200
TEMPLATE_PATH = Path(__file__).resolve().with_name(TEMPLATE_FILENAME)
JET_CAP_RHO = 1.2
JET_CAP_ZETA = 2.5


def parse_float(value: str) -> float:
    try:
        raw = str(value).strip()
        if not raw:
            return 0.0
        cleaned = raw.replace(" ", "")
        if cleaned.count(",") == 1 and "." not in cleaned:
            cleaned = cleaned.replace(",", ".")
        elif cleaned.count(",") > 1 and "." not in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
        return float(cleaned)
    except Exception:
        return 0.0


def normalize_tip(value: object) -> str:
    tip = str(value or "").strip().lower()
    normalized = (
        tip.replace("ı", "i")
        .replace("ş", "s")
        .replace("ç", "c")
        .replace("ğ", "g")
        .replace("ö", "o")
        .replace("ü", "u")
    )
    if normalized == "dirsek":
        return "dirsek"
    if normalized in {"kanal", "duz kanal", "duzkanal"}:
        return "kanal"
    if normalized in {"jet-cap", "jetcap", "jet cap"}:
        return "jet-cap"
    if normalized in {"ekipman", "equipment"}:
        return "ekipman"
    return ""


def calculate_pressure_metrics(
    hava_sic_c: float,
    rakim_m: float,
    aci: int,
    rd_val: float,
    cap_mm: float,
    debi_m3h: float,
) -> dict[str, float]:
    mutlak_t = hava_sic_c + 273.15
    rakim = max(rakim_m, 0.0)
    p0 = 101325.0
    scale_height = 8434.0

    try:
        pressure_pa = p0 * math.exp(-rakim / scale_height)
    except Exception:
        pressure_pa = p0

    rho = (pressure_pa / (mutlak_t * 287.05)) if mutlak_t > 0 else 0.0
    sabit = SABIT_TABLOSU.get(aci, {}).get(rd_val, 0.0)

    cap_m = cap_mm / 1000.0 if cap_mm > 0 else 0.0
    alan = math.pi * (cap_m**2) / 4.0 if cap_m > 0 else 0.0
    hiz = (debi_m3h / alan / 3600.0) if (alan > 0 and debi_m3h > 0) else 0.0

    if cap_mm > 0 and hiz > 0:
        try:
            term = (0.03 / cap_mm) + (68.0 / (66.4 * cap_mm * hiz))
            f_pp_val = 0.11 * (term**0.25)
        except Exception:
            f_pp_val = 0.0
    else:
        f_pp_val = 0.0

    f_val = 0.85 * f_pp_val + 0.0028
    vp = rho * hiz * hiz / 2.0
    dpf_val = ((1000.0 * f_val) / cap_mm) * vp * 1.15 if cap_mm > 0 else 0.0

    return {
        "rho": rho,
        "alan": alan,
        "hiz": hiz,
        "dirsek_kayip": vp * sabit,
        "duz_kanal": dpf_val * 1.18,
    }


def calculate_jet_cap_loss(cap_mm: float, debi_m3h: float) -> dict[str, float]:
    cap_m = cap_mm / 1000.0 if cap_mm > 0 else 0.0
    q_s = debi_m3h / 3600.0 if debi_m3h > 0 else 0.0
    alan = math.pi * (cap_m**2) / 4.0 if cap_m > 0 else 0.0
    hiz = q_s / alan if alan > 0 else 0.0
    dinamik_basinc = JET_CAP_RHO * hiz * hiz / 2.0
    kayip = JET_CAP_ZETA * dinamik_basinc
    return {
        "q_s": q_s,
        "alan": alan,
        "hiz": hiz,
        "dinamik_basinc": dinamik_basinc,
        "kayip": kayip,
    }


def build_pressure_loss_template(file_path: str | Path) -> Path:
    workbook = Workbook()
    ayarlar_ws = workbook.active
    ayarlar_ws.title = "Ayarlar"
    liste_ws = workbook.create_sheet("Kanal Listesi")
    sabitler_ws = workbook.create_sheet("Sabitler")

    header_fill = PatternFill("solid", fgColor="D32F2F")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="F3F3F3")

    ayarlar_ws["A1"] = "Parametre"
    ayarlar_ws["B1"] = "Deger"
    ayarlar_ws["A2"] = "Hava Sicakligi (C)"
    ayarlar_ws["B2"] = 20
    ayarlar_ws["A3"] = "Rakim (m)"
    ayarlar_ws["B3"] = 0
    ayarlar_ws["A5"] = "Not"
    ayarlar_ws["B5"] = "Kanal Listesi sayfasinda Tip alanina yalnizca Dirsek veya Kanal yazin."
    for cell in ayarlar_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell_ref in ("A2", "A3", "A5"):
        ayarlar_ws[cell_ref].fill = section_fill
        ayarlar_ws[cell_ref].font = Font(bold=True)
    ayarlar_ws.column_dimensions["A"].width = 28
    ayarlar_ws.column_dimensions["B"].width = 60

    headers = [
        "Tip",
        "Aci",
        "r/D",
        "Boru Capi (mm)",
        "Debi (m3/h)",
        "Dirsek Adedi",
        "Duz Kanal Uzunlugu (m)",
        "Hava Yogunlugu (kg/m3)",
        "Kesit Alani (m2)",
        "Tasima Hizi (m/sn)",
        "Dirsek Birim Kaybi (Pa)",
        "Duz Kanal Birim Kaybi (Pa/m)",
        "Jet-Cap Basinc Kaybi (Pa)",
        "Ekipman Tipi",
        "Ekipman Basinc Kaybi (Pa)",
        "Toplam Basinc Kaybi (Pa)",
    ]
    liste_ws.append(headers)
    for cell in liste_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    sabitler_ws.append(["Aci", 0.75, 1.0, 1.5, 2.0])
    for cell in sabitler_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for aci, rd_map in SABIT_TABLOSU.items():
        sabitler_ws.append([aci, rd_map[0.75], rd_map[1.0], rd_map[1.5], rd_map[2.0]])
    sabitler_ws["G1"] = "Tip Secenekleri"
    sabitler_ws["G2"] = "Kanal"
    sabitler_ws["G3"] = "Dirsek"
    sabitler_ws["G4"] = "Jet-Cap"
    sabitler_ws["G5"] = "Ekipman"
    sabitler_ws["H1"] = "rD Secenekleri"
    sabitler_ws["H2"] = "0,75"
    sabitler_ws["H3"] = "1"
    sabitler_ws["H4"] = "1,5"
    sabitler_ws["H5"] = "2"
    workbook.defined_names["TipSecenekleri"] = DefinedName("TipSecenekleri", attr_text="Sabitler!$G$2:$G$5")
    workbook.defined_names["rDSecenekleri"] = DefinedName("rDSecenekleri", attr_text="Sabitler!$H$2:$H$5")
    sabitler_ws.sheet_state = "hidden"

    tip_validation = DataValidation(
        type="list",
        formula1="=TipSecenekleri",
        allow_blank=True,
    )
    tip_validation.promptTitle = "Tip Secimi"
    tip_validation.prompt = "Kanal veya Dirsek secin."
    tip_validation.errorTitle = "Gecersiz Tip"
    tip_validation.error = "Yalnizca Kanal veya Dirsek secilebilir."
    liste_ws.add_data_validation(tip_validation)
    tip_validation.add(f"A2:A{TEMPLATE_MAX_ROWS + 1}")

    rd_validation = DataValidation(
        type="list",
        formula1="=rDSecenekleri",
        allow_blank=True,
    )
    rd_validation.promptTitle = "r/D Secimi"
    rd_validation.prompt = "0,75 / 1 / 1,5 / 2 degerlerinden birini secin."
    rd_validation.errorTitle = "Gecersiz r/D"
    rd_validation.error = "Yalnizca 0,75 / 1 / 1,5 / 2 secilebilir."
    liste_ws.add_data_validation(rd_validation)
    rd_validation.add(f"C2:C{TEMPLATE_MAX_ROWS + 1}")

    liste_ws["A2"] = "Dirsek"
    liste_ws["B2"] = 90
    liste_ws["C2"] = "1,5"
    liste_ws["D2"] = 250
    liste_ws["E2"] = 4500
    liste_ws["F2"] = 3
    liste_ws["G2"] = 0

    liste_ws["A3"] = "Kanal"
    liste_ws["B3"] = 90
    liste_ws["C3"] = "1,5"
    liste_ws["D3"] = 250
    liste_ws["E3"] = 4500
    liste_ws["F3"] = 0
    liste_ws["G3"] = 12

    liste_ws["A4"] = "Jet-Cap"
    liste_ws["B4"] = 90
    liste_ws["C4"] = "1,5"
    liste_ws["D4"] = 250
    liste_ws["E4"] = 4500
    liste_ws["F4"] = 0
    liste_ws["G4"] = 0

    liste_ws["A5"] = "Ekipman"
    liste_ws["N5"] = "Filtre"
    liste_ws["O5"] = 250

    for row_idx in range(2, TEMPLATE_MAX_ROWS + 2):
        rho_formula = "=(101325*EXP(-Ayarlar!$B$3/8434))/((Ayarlar!$B$2+273.15)*287.05)"
        alan_formula = f'=IF(D{row_idx}>0,PI()*((D{row_idx}/1000)^2)/4,0)'
        hiz_formula = f'=IF(AND(I{row_idx}>0,E{row_idx}>0),E{row_idx}/I{row_idx}/3600,0)'
        lookup_formula = (
            f'IFERROR(INDEX(Sabitler!$B$2:$E$11,'
            f'MATCH(B{row_idx},Sabitler!$A$2:$A$11,0),'
            f'MATCH(C{row_idx},Sabitler!$B$1:$E$1,0)),0)'
        )
        dirsek_formula = (
            f'=IF(UPPER(A{row_idx})="DIRSEK",'
            f'(H{row_idx}*J{row_idx}*J{row_idx}/2)*{lookup_formula},0)'
        )
        duz_formula = (
            f'=IF(UPPER(A{row_idx})="KANAL",'
            f'IF(AND(D{row_idx}>0,J{row_idx}>0),'
            f'(((1000*((0.85*(0.11*(((0.03/D{row_idx})+(68/(66.4*D{row_idx}*J{row_idx})))^0.25)))+0.0028))/D{row_idx})'
            f'*(H{row_idx}*J{row_idx}*J{row_idx}/2)*1.15)*1.18,0),0)'
        )
        jetcap_formula = f'=IF(UPPER(A{row_idx})="JET-CAP",2.5*(1.2*J{row_idx}*J{row_idx}/2),0)'
        toplam_formula = (
            f'=IF(UPPER(A{row_idx})="DIRSEK",K{row_idx}*MAX(F{row_idx},1),'
            f'IF(UPPER(A{row_idx})="KANAL",L{row_idx}*MAX(G{row_idx},0),'
            f'IF(UPPER(A{row_idx})="JET-CAP",M{row_idx},'
            f'IF(UPPER(A{row_idx})="EKIPMAN",O{row_idx},0))))'
        )

        liste_ws[f"H{row_idx}"] = rho_formula
        liste_ws[f"I{row_idx}"] = alan_formula
        liste_ws[f"J{row_idx}"] = hiz_formula
        liste_ws[f"K{row_idx}"] = dirsek_formula
        liste_ws[f"L{row_idx}"] = duz_formula
        liste_ws[f"M{row_idx}"] = jetcap_formula
        liste_ws[f"P{row_idx}"] = toplam_formula

    total_row = TEMPLATE_MAX_ROWS + 3
    liste_ws[f"O{total_row}"] = "Toplam Basinc Kaybi"
    liste_ws[f"O{total_row}"].fill = section_fill
    liste_ws[f"O{total_row}"].font = Font(bold=True)
    liste_ws[f"P{total_row}"] = f"=SUM(P2:P{TEMPLATE_MAX_ROWS + 1})"
    liste_ws[f"P{total_row}"].fill = section_fill
    liste_ws[f"P{total_row}"].font = Font(bold=True)
    liste_ws.freeze_panes = "A2"

    widths = {
        "A": 14,
        "B": 10,
        "C": 10,
        "D": 16,
        "E": 16,
        "F": 14,
        "G": 22,
        "H": 20,
        "I": 18,
        "J": 18,
        "K": 22,
        "L": 24,
        "M": 22,
        "N": 18,
        "O": 22,
        "P": 22,
    }
    for col_letter, width in widths.items():
        liste_ws.column_dimensions[col_letter].width = width

    output_path = Path(file_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def pressure_loss_calc_ekrani_ac(
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    pencere = ctk.CTkToplevel(parent)
    pencere.title("Basınç Kaybı Hesabı")
    pencere.geometry("980x820")

    screen_width = pencere.winfo_screenwidth()
    screen_height = pencere.winfo_screenheight()
    min_width = max(420, int(screen_width * 0.33))
    min_height = max(320, int(screen_height * 0.33))
    pencere.minsize(min_width, min_height)
    pencere.resizable(True, True)

    pencere.update_idletasks()
    pos_x = (screen_width // 2) - (980 // 2)
    pos_y = (screen_height // 2) - (820 // 2)
    pencere.geometry(f"980x820+{pos_x}+{pos_y}")
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=min_width, min_height=min_height)

    def set_readonly(entry: ctk.CTkEntry, text: str) -> None:
        entry.configure(state="normal")
        entry.delete(0, "end")
        entry.insert(0, text)
        entry.configure(state="disabled")

    def set_entry_value(entry: ctk.CTkEntry, text: str) -> None:
        entry.delete(0, "end")
        entry.insert(0, text)

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 16))
    ctk.CTkLabel(
        header,
        text="Hava Yoğunluğu ve Basınç Kaybı Hesabı",
        font=ctk.CTkFont(family="Inter", size=22, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text="Kayıp listesi Excel şablonundan içe aktarılabilir ve toplam basınç kaybı otomatik hesaplanır.",
        font=ctk.CTkFont(family="Inter", size=12),
        text_color="#666666",
    ).pack(anchor="w")

    split = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    split.pack(fill="both", expand=True)
    split.grid_columnconfigure(0, weight=3)
    split.grid_columnconfigure(1, weight=2)
    split.grid_rowconfigure(0, weight=1)

    content = ctk.CTkScrollableFrame(split, fg_color="#f5f5f5")
    content.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
    content.grid_columnconfigure(1, weight=1)

    right_panel = ctk.CTkFrame(split, fg_color="#f5f5f5")
    right_panel.grid(row=0, column=1, sticky="nsew")

    row = 0

    def add_section(title: str) -> None:
        nonlocal row
        ctk.CTkLabel(
            content,
            text=title,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#212121",
        ).grid(row=row, column=0, columnspan=3, sticky="w", pady=(6, 10))
        row += 1

    def add_entry_field(
        label: str,
        unit: str = "",
        placeholder: str = "0",
        readonly: bool = False,
    ) -> ctk.CTkEntry:
        nonlocal row
        ctk.CTkLabel(
            content,
            text=label,
            font=ctk.CTkFont(size=14),
            text_color="#212121",
        ).grid(row=row, column=0, sticky="w", padx=(0, 10), pady=(6, 6))
        entry = ctk.CTkEntry(
            content,
            placeholder_text=placeholder,
            height=34,
            fg_color="#eeeeee" if readonly else None,
            text_color="#333333" if readonly else None,
        )
        entry.grid(row=row, column=1, sticky="ew", pady=(6, 6))
        if readonly:
            entry.configure(state="disabled")
        if unit:
            ctk.CTkLabel(
                content,
                text=unit,
                font=ctk.CTkFont(size=13),
                text_color="#666666",
            ).grid(row=row, column=2, sticky="w", padx=(10, 0))
        row += 1
        return entry

    def add_option_field(label: str, values: list[str], default: str) -> ctk.CTkOptionMenu:
        nonlocal row
        ctk.CTkLabel(
            content,
            text=label,
            font=ctk.CTkFont(size=14),
            text_color="#212121",
        ).grid(row=row, column=0, sticky="w", padx=(0, 10), pady=(6, 6))
        option = ctk.CTkOptionMenu(content, values=values)
        option.set(default)
        option.grid(row=row, column=1, sticky="w", pady=(6, 6))
        row += 1
        return option

    add_section("HAVA YOĞUNLUĞU HESAPLAMA")
    entry_hava_sicak = add_entry_field("Hava Sıcaklığı", "°C")
    entry_rakim = add_entry_field("Rakım", "m")
    entry_hava_yogun = add_entry_field("Hava Yoğunluğu", "kg/m³", readonly=True)

    add_section("BASINÇ KAYBI HESABI")
    option_aci = add_option_field("Açı", ["20", "30", "45", "60", "75", "90", "110", "130", "150", "180"], "90")
    option_rd = add_option_field("r/D", ["0,75", "1", "1,5", "2"], "1,5")
    entry_cap_mm = add_entry_field("Boru Çapı", "mm")
    entry_kesit_alani = add_entry_field("Boru Kesit Alanı", "m²", readonly=True)
    entry_debi = add_entry_field("Debi", "m³/h")
    entry_tasima_hizi = add_entry_field("Taşıma Hızı", "m/sn", readonly=True)
    entry_dirsek_kayip = add_entry_field("Dirsekten Oluşacak Kayıp", "Pa", readonly=True)
    entry_dirsek_adet = add_entry_field("Dirsek Adedi", "adet", placeholder="1")
    btn_dirsek_ekle = ctk.CTkButton(
        content,
        text="Dirsek Ekle",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
    )
    btn_dirsek_ekle.grid(row=row, column=1, sticky="w", pady=(2, 12))
    row += 1

    entry_duz_kanal = add_entry_field("Düz Kanal Basınç Kaybı", "Pa/m", readonly=True)
    entry_duz_uzunluk = add_entry_field("Düz Kanal Uzunluğu", "m")
    btn_duz_kanal_ekle = ctk.CTkButton(
        content,
        text="Düz Kanal Ekle",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
    )
    btn_duz_kanal_ekle.grid(row=row, column=1, sticky="w", pady=(2, 12))
    row += 1

    entry_jet_cap_kayip = add_entry_field("Jet-Cap Basınç Kaybı", "Pa", readonly=True)
    btn_jet_cap_ekle = ctk.CTkButton(
        content,
        text="Jet-Cap Ekle",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
    )
    btn_jet_cap_ekle.grid(row=row, column=1, sticky="w", pady=(2, 12))
    row += 1

    add_section("EKIPMAN EKLEME")
    option_ekipman = add_option_field(
        "Ekipman Tipi",
        ["Filtre", "Akrobat Kol", "Davlumbaz", "Menfez", "Diger"],
        "Filtre",
    )
    entry_ekipman_kaybi = add_entry_field("Ekipman Basınç Kaybı", "Pa")
    btn_ekipman_ekle = ctk.CTkButton(
        content,
        text="Ekipman Ekle",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
    )
    btn_ekipman_ekle.grid(row=row, column=1, sticky="w", pady=(2, 12))
    row += 1

    right_toolbar = ctk.CTkFrame(right_panel, fg_color="#f5f5f5")
    right_toolbar.pack(fill="x", pady=(0, 8))

    btn_import = ctk.CTkButton(
        right_toolbar,
        text="Excel'den İçe Aktar",
        width=150,
        fg_color="#2e7d32",
        hover_color="#ffffff",
        text_color="white",
        border_color="#2e7d32",
        border_width=0,
        command=lambda: None,
    )
    btn_import.pack(side="left", padx=(0, 8))

    btn_template = ctk.CTkButton(
        right_toolbar,
        text="Taslak Excel İndir",
        width=150,
        fg_color="#6a1b9a",
        hover_color="#ffffff",
        text_color="white",
        border_color="#6a1b9a",
        border_width=0,
        command=lambda: None,
    )
    btn_template.pack(side="left", padx=(0, 8))

    btn_export = ctk.CTkButton(
        right_toolbar,
        text="Excel'e Aktar",
        width=120,
        fg_color="#1976d2",
        hover_color="#ffffff",
        text_color="white",
        border_color="#1976d2",
        border_width=0,
        command=lambda: None,
    )
    btn_export.pack(side="right", padx=(0, 8))

    btn_sil = ctk.CTkButton(
        right_toolbar,
        text="Seçili Ögeyi Sil",
        width=130,
        fg_color="#d32f2f",
        hover_color="#ffffff",
        text_color="white",
        border_color="#d32f2f",
        border_width=0,
        command=lambda: None,
    )
    btn_sil.pack(side="right", padx=(0, 8))

    def bind_hover(button: ctk.CTkButton, color: str) -> None:
        def on_enter(_event) -> None:
            button.configure(fg_color="white", text_color=color, border_color=color, border_width=2)

        def on_leave(_event) -> None:
            button.configure(fg_color=color, text_color="white", border_width=0)

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    bind_hover(btn_import, "#2e7d32")
    bind_hover(btn_template, "#6a1b9a")
    bind_hover(btn_export, "#1976d2")
    bind_hover(btn_sil, "#d32f2f")

    table_frame = ctk.CTkFrame(right_panel, fg_color="#f5f5f5")
    table_frame.pack(fill="both", expand=True)

    tree = ttk.Treeview(
        table_frame,
        columns=("id", "tip", "cap", "debi", "hiz", "miktar", "kayip"),
        show="headings",
        height=18,
    )
    tree.heading("id", text="ID")
    tree.heading("tip", text="Tip")
    tree.heading("cap", text="Kanal Çapı (mm)")
    tree.heading("debi", text="Debi (m³/h)")
    tree.heading("hiz", text="Kanal İçi Hız (m/sn)")
    tree.heading("miktar", text="Miktar")
    tree.heading("kayip", text="Basınç Kaybı (Pa)")
    tree.column("id", width=55, anchor="center")
    tree.column("tip", width=80, anchor="center")
    tree.column("cap", width=105, anchor="e")
    tree.column("debi", width=105, anchor="e")
    tree.column("hiz", width=120, anchor="e")
    tree.column("miktar", width=105, anchor="center")
    tree.column("kayip", width=130, anchor="e")
    tree.pack(fill="both", expand=True, padx=4, pady=(0, 8))

    try:
        apply_bomaksan_table_style(tree)
    except Exception:
        pass

    total_frame = ctk.CTkFrame(right_panel, fg_color="#f5f5f5")
    total_frame.pack(fill="x", pady=(6, 0))
    total_frame.grid_columnconfigure(1, weight=1)
    ctk.CTkLabel(
        total_frame,
        text="Toplam Basınç Kaybı",
        font=ctk.CTkFont(size=14),
        text_color="#212121",
    ).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=(6, 6))
    entry_toplam = ctk.CTkEntry(total_frame, height=34, fg_color="#eeeeee", text_color="#333333")
    entry_toplam.grid(row=0, column=1, sticky="ew", pady=(6, 6))
    entry_toplam.configure(state="disabled")
    ctk.CTkLabel(
        total_frame,
        text="Pa",
        font=ctk.CTkFont(size=13),
        text_color="#666666",
    ).grid(row=0, column=2, sticky="w", padx=(10, 0))

    next_id = {"val": 1}

    def tablo_toplam_guncelle() -> None:
        toplam = 0.0
        for item in tree.get_children():
            values = tree.item(item, "values")
            try:
                toplam += float(str(values[6]).replace(",", "."))
            except Exception:
                pass
        set_readonly(entry_toplam, f"{toplam:.1f}")

    def tablo_zebra() -> None:
        try:
            apply_zebra_striping(tree, tree.get_children())
        except Exception:
            pass

    def insert_tree_row(
        tip_text: str,
        cap_mm: float | str,
        debi: float | str,
        hiz: float | str,
        miktar_text: str,
        kayip: float,
    ) -> None:
        row_id = next_id["val"]
        next_id["val"] += 1
        cap_text = f"{cap_mm:.0f}" if isinstance(cap_mm, (int, float)) else str(cap_mm)
        debi_text = f"{debi:.0f}" if isinstance(debi, (int, float)) else str(debi)
        hiz_text = f"{hiz:.1f}" if isinstance(hiz, (int, float)) else str(hiz)
        tree.insert(
            "",
            "end",
            values=(
                row_id,
                tip_text,
                cap_text,
                debi_text,
                hiz_text,
                miktar_text,
                f"{kayip:.1f}",
            ),
        )

    def clear_tree() -> None:
        for item in tree.get_children():
            tree.delete(item)
        next_id["val"] = 1
        tablo_toplam_guncelle()

    def sil_secili() -> None:
        selected = tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("Onay", f"Seçili {len(selected)} ögeyi silmek istediğinize emin misiniz?"):
            return
        for item in selected:
            tree.delete(item)
        tablo_zebra()
        tablo_toplam_guncelle()

    def excel_aktar() -> None:
        items = tree.get_children()
        if not items:
            messagebox.showwarning("Uyarı", "Aktarılacak kayıt bulunamadı.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
            title="Excel'e Aktar",
        )
        if not file_path:
            return

        columns = list(tree.cget("columns"))
        headers = [tree.heading(col).get("text", col) for col in columns]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Basinc Kaybi Listesi"
        worksheet.append(headers)

        toplam_kayip = 0.0
        numeric_indices = {2, 3, 4, 6}
        for item_id in items:
            values = list(tree.item(item_id, "values"))
            row_data: list[object] = []
            for idx, value in enumerate(values):
                if idx in numeric_indices:
                    try:
                        numeric_value = float(str(value).replace(",", "."))
                        if idx == 6:
                            toplam_kayip += numeric_value
                        row_data.append(numeric_value)
                    except Exception:
                        row_data.append(value)
                else:
                    row_data.append(value)
            worksheet.append(row_data)

        worksheet.append(["", "", "", "", "", "Toplam", toplam_kayip])

        for col_idx, header_text in enumerate(headers, start=1):
            max_len = len(str(header_text))
            for row_cells in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_value = row_cells[0].value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(10, min(40, max_len + 2))

        workbook.save(file_path)
        messagebox.showinfo("Başarılı", f"Excel'e aktarıldı:\n{file_path}")

    def taslak_excel_indir() -> None:
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=TEMPLATE_FILENAME,
            filetypes=[("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")],
            title="Taslak Excel Kaydet",
        )
        if not file_path:
            return
        try:
            build_pressure_loss_template(file_path)
            messagebox.showinfo("Başarılı", f"Taslak Excel oluşturuldu:\n{file_path}")
        except Exception as exc:
            messagebox.showerror("Hata", f"Taslak Excel oluşturulamadı:\n{exc}")

    def excelden_ice_aktar() -> None:
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Dosyası", "*.xlsx;*.xlsm"), ("Tüm Dosyalar", "*.*")],
            title="Excel'den İçe Aktar",
        )
        if not file_path:
            return

        try:
            workbook = load_workbook(file_path, data_only=False)
        except Exception as exc:
            messagebox.showerror("Hata", f"Excel dosyası açılamadı:\n{exc}")
            return

        if "Ayarlar" not in workbook.sheetnames or "Kanal Listesi" not in workbook.sheetnames:
            messagebox.showwarning(
                "Uyarı",
                "Seçilen dosya beklenen taslak yapısına uymuyor.\n'Ayarlar' ve 'Kanal Listesi' sayfaları gerekli.",
            )
            return

        ayarlar_ws = workbook["Ayarlar"]
        liste_ws = workbook["Kanal Listesi"]

        hava_sic = parse_float(ayarlar_ws["B2"].value)
        rakim = parse_float(ayarlar_ws["B3"].value)
        set_entry_value(entry_hava_sicak, f"{hava_sic:g}")
        set_entry_value(entry_rakim, f"{rakim:g}")

        clear_tree()
        imported_count = 0
        skipped_count = 0

        for row_idx in range(2, liste_ws.max_row + 1):
            tip = normalize_tip(liste_ws[f"A{row_idx}"].value)
            if not tip:
                continue

            aci = int(parse_float(liste_ws[f"B{row_idx}"].value) or 90)
            rd_val = parse_float(liste_ws[f"C{row_idx}"].value)
            rd_val = rd_val if rd_val in {0.75, 1.0, 1.5, 2.0} else 1.5
            cap_mm = parse_float(liste_ws[f"D{row_idx}"].value)
            debi = parse_float(liste_ws[f"E{row_idx}"].value)
            dirsek_adedi = parse_float(liste_ws[f"F{row_idx}"].value)
            duz_uzunluk = parse_float(liste_ws[f"G{row_idx}"].value)
            ekipman_tipi = str(liste_ws[f"N{row_idx}"].value or "").strip() or "Ekipman"
            ekipman_kaybi = parse_float(liste_ws[f"O{row_idx}"].value)

            if tip == "ekipman":
                if ekipman_kaybi <= 0:
                    skipped_count += 1
                    continue
                insert_tree_row(ekipman_tipi, "-", "-", "-", "1 adet", ekipman_kaybi)
                imported_count += 1
                continue

            if cap_mm <= 0 or debi <= 0:
                skipped_count += 1
                continue

            metrics = calculate_pressure_metrics(hava_sic, rakim, aci, rd_val, cap_mm, debi)
            if tip == "dirsek":
                adet = max(dirsek_adedi, 1.0)
                kayip = metrics["dirsek_kayip"] * adet
                miktar_text = f"{adet:.0f} adet"
                tip_text = "Dirsek"
                hiz = metrics["hiz"]
            elif tip == "kanal":
                uzunluk = max(duz_uzunluk, 0.0)
                kayip = metrics["duz_kanal"] * uzunluk
                miktar_text = f"{uzunluk:.1f} mt"
                tip_text = "Kanal"
                hiz = metrics["hiz"]
            else:
                jet_cap = calculate_jet_cap_loss(cap_mm, debi)
                kayip = jet_cap["kayip"]
                miktar_text = "1 adet"
                tip_text = "Jet-Cap"
                hiz = jet_cap["hiz"]

            insert_tree_row(tip_text, cap_mm, debi, hiz, miktar_text, kayip)
            imported_count += 1

        tablo_zebra()
        tablo_toplam_guncelle()
        hesapla_ve_guncelle()

        if imported_count == 0:
            messagebox.showwarning("Uyarı", "Excel dosyasında içe aktarılabilir kayıt bulunamadı.")
            return

        summary = f"{imported_count} kayıt içe aktarıldı."
        if skipped_count:
            summary += f"\n{skipped_count} satır eksik veri nedeniyle atlandı."
        messagebox.showinfo("Başarılı", summary)

    def ekle_kayit(kayip_tipi: str) -> None:
        hava_sic = parse_float(entry_hava_sicak.get())
        rakim = parse_float(entry_rakim.get())
        try:
            aci = int(option_aci.get())
        except Exception:
            aci = 90
        try:
            rd_val = float(option_rd.get().replace(",", "."))
        except Exception:
            rd_val = 1.5

        cap = parse_float(entry_cap_mm.get())
        debi = parse_float(entry_debi.get())
        metrics = calculate_pressure_metrics(hava_sic, rakim, aci, rd_val, cap, debi)

        if kayip_tipi == "dirsek":
            adet = max(parse_float(entry_dirsek_adet.get()), 1.0)
            kayip = metrics["dirsek_kayip"] * adet
            miktar_text = f"{adet:.0f} adet"
            tip_text = "Dirsek"
        else:
            uzunluk = max(parse_float(entry_duz_uzunluk.get()), 0.0)
            kayip = metrics["duz_kanal"] * uzunluk
            miktar_text = f"{uzunluk:.1f} mt"
            tip_text = "Kanal"

        insert_tree_row(tip_text, cap, debi, metrics["hiz"], miktar_text, kayip)
        tablo_zebra()
        tablo_toplam_guncelle()

    def ekipman_ekle() -> None:
        ekipman_tipi = option_ekipman.get().strip() or "Ekipman"
        kayip = max(parse_float(entry_ekipman_kaybi.get()), 0.0)
        if kayip <= 0:
            messagebox.showwarning("Uyarı", "Lütfen ekipman için 0'dan büyük bir basınç kaybı girin.")
            return

        insert_tree_row(ekipman_tipi, "-", "-", "-", "1 adet", kayip)
        tablo_zebra()
        tablo_toplam_guncelle()
        set_entry_value(entry_ekipman_kaybi, "")

    def jet_cap_ekle() -> None:
        cap = parse_float(entry_cap_mm.get())
        debi = parse_float(entry_debi.get())
        if cap <= 0 or debi <= 0:
            messagebox.showwarning("Uyarı", "Jet-Cap için boru çapı ve debi değerleri 0'dan büyük olmalıdır.")
            return

        jet_cap = calculate_jet_cap_loss(cap, debi)
        insert_tree_row("Jet-Cap", cap, debi, jet_cap["hiz"], "1 adet", jet_cap["kayip"])
        tablo_zebra()
        tablo_toplam_guncelle()

    btn_dirsek_ekle.configure(command=lambda: ekle_kayit("dirsek"))
    btn_duz_kanal_ekle.configure(command=lambda: ekle_kayit("duz"))
    btn_jet_cap_ekle.configure(command=jet_cap_ekle)
    btn_ekipman_ekle.configure(command=ekipman_ekle)
    btn_sil.configure(command=sil_secili)
    btn_export.configure(command=excel_aktar)
    btn_import.configure(command=excelden_ice_aktar)
    btn_template.configure(command=taslak_excel_indir)

    def hesapla_ve_guncelle(_event=None) -> None:
        hava_sic = parse_float(entry_hava_sicak.get())
        rakim = parse_float(entry_rakim.get())
        try:
            aci = int(option_aci.get())
        except Exception:
            aci = 90
        try:
            rd_val = float(option_rd.get().replace(",", "."))
        except Exception:
            rd_val = 1.5

        cap_mm = parse_float(entry_cap_mm.get())
        debi_m3h = parse_float(entry_debi.get())
        metrics = calculate_pressure_metrics(hava_sic, rakim, aci, rd_val, cap_mm, debi_m3h)
        jet_cap = calculate_jet_cap_loss(cap_mm, debi_m3h)

        set_readonly(entry_hava_yogun, f"{metrics['rho']:.4f}")
        set_readonly(entry_kesit_alani, f"{metrics['alan']:.6f}")
        set_readonly(entry_tasima_hizi, f"{metrics['hiz']:.3f}")
        set_readonly(entry_dirsek_kayip, f"{metrics['dirsek_kayip']:.1f}")
        set_readonly(entry_duz_kanal, f"{metrics['duz_kanal']:.1f}")
        set_readonly(entry_jet_cap_kayip, f"{jet_cap['kayip']:.1f}")

    for entry in (entry_hava_sicak, entry_rakim, entry_cap_mm, entry_debi):
        entry.bind("<KeyRelease>", hesapla_ve_guncelle)
        entry.bind("<FocusOut>", hesapla_ve_guncelle)
    option_aci.configure(command=lambda _: hesapla_ve_guncelle())
    option_rd.configure(command=lambda _: hesapla_ve_guncelle())
    hesapla_ve_guncelle()

    buttons = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    buttons.pack(fill="x", pady=(10, 0))
    ctk.CTkButton(
        buttons,
        text="Kapat",
        width=100,
        fg_color="#9e9e9e",
        hover_color="#757575",
        text_color="white",
        command=pencere.destroy,
    ).pack(side="right")
