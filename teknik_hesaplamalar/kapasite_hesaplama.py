from typing import Optional

import customtkinter as ctk
from tkinter import filedialog, messagebox, simpledialog, ttk

from core.session import get_username
from core.window_utils import open_window_zoomed


MODULE_CONFIGS = [
    {
        "key": "isil_kesim",
        "icon": "🔥",
        "title": "Isil Kesim Kapasite Hesaplama",
        "description": "Thermal cutting secim aracindaki mantikla modül alanina gore gerekli emiş kapasitesini hesaplar.",
        "defaults": {
            "cutting_power": "12 kW (Laser)",
            "valve_tightness": "Kotu",
            "valve_count": "16",
            "module_width_mm": "1500",
            "module_length_mm": "2000",
            "reserve": "15",
        },
    },
    {
        "key": "cnc_torna",
        "icon": "🛢",
        "title": "CNC Torna Tezgah Kapasite Hesaplama",
        "description": "Oil mist filtre secim aracindaki mantikla gerekli hava debisini ve onerilen urunu hesaplar.",
        "defaults": {
            "coolant_type": "Water Soluble",
            "max_spindle_speed": "8000",
            "coolant_pump_pressure": "100",
            "processing_space": "6",
            "door_type": "Sliding Door, Open on Top",
            "door_opening_area": "1",
            "cycle_time": "300",
            "viscosity": "5",
        },
    },
    {
        "key": "davlumbaz",
        "icon": "🌬",
        "title": "Davlumbaz Kapasite Hesaplama",
        "description": "Teknik hesap dokumanindaki davlumbaz formulleri ile gerekli hava debisini hesaplar.",
        "defaults": {
            "hood_type": "Flansli Davlumbaz",
            "capture_velocity": "0,75",
            "distance": "0,5",
            "width": "1,2",
            "length": "0,8",
            "reserve": "15",
            "capture_profile": "Dusuk hizli yayilim (0,5 - 1,0 m/s)",
        },
    },
    {
        "key": "genel_hol",
        "icon": "GH",
        "title": "Kaynak Hol Havalandirmasi Kapasite Hesaplama",
        "description": "Kaynak prosesi kaynakli duman yukune gore minimum ACH ve gerekli kapasiteyi hesaplar.",
        "defaults": {
            "welding_type": "SC Ar/CO2",
            "manual_welder_count": "10",
            "manual_wire_consumption": "1",
            "robot_count": "4",
            "robot_wire_consumption": "5",
            "hall_depth": "50",
            "hall_width": "30",
            "height": "8",
            "particle_limit": "4",
            "requested_ach": "3",
        },
    },
]

THERMAL_CUTTING_POWER_OPTIONS = {
    "2 kW (Laser)": 3500,
    "4 kW (Laser)": 4000,
    "6 kW (Laser)": 4700,
    "8 kW (Laser)": 5000,
    "10 kW (Laser)": 5500,
    "12 kW (Laser)": 6000,
    "Up to 150 A (Plasma)": 3600,
    "Up to 300 A (Plasma)": 4100,
    "Up to 450 A (Plasma)": 4700,
    "Above 450 A (Plasma)": 5000,
    "Up to 100 mm (Oxy Cutting)": 3600,
    "Up to 200 mm (Oxy Cutting)": 4300,
    "Up to 300 mm (Oxy Cutting)": 5000,
}

VALVE_TIGHTNESS_OPTIONS = {
    "Kotu": 70,
    "Orta": 60,
    "Iyi": 50,
    "Cok Iyi": 40,
}

CNC_COOLANT_TYPE_OPTIONS = [
    "Straight Oil",
    "Water Soluble",
    "Semi-Synthetic",
    "Synthetic",
]

CNC_DEFAULT_VISCOSITY = {
    "Synthetic": "0,8",
    "Water Soluble": "0,8",
    "Semi-Synthetic": "1",
    "Straight Oil": "10",
}

CNC_DOOR_TYPE_OPTIONS = [
    "Sliding Door, Vertical",
    "Sliding Door, Horizontal",
    "Sliding Door, Open on Top",
]

CNC_HEPA_PRODUCTS = [
    ("H.YBFpro MINI", 500),
    ("H.YBFpro MIDI", 900),
    ("DT.HYBF.1.100.200.11.2.50.380.DS.LCD", 1000),
    ("DT.HYBF.1.150.200.15.2.50.380.DS.LCD", 1500),
    ("DT.HYBF.1.200.200.22.2.50.380.DS.LCD", 2000),
    ("DT.HYBF.2.300.200.30.2.50.380.DS.LCD", 3000),
    ("DT.HYBF.2.350.200.40.2.50.380.DS.LCD", 3500),
    ("DT.HYBF.2.400.200.40.2.50.380.DS.LCD", 4000),
    ("DT.HYBF.2.450.200.55.2.50.380.DS.LCD", 4500),
    ("DT.HYBF.2.500.200.55.2.50.380.DS.LCD", 5000),
]

CNC_STANDARD_PRODUCTS = [
    ("YBFpro MINI", 600),
    ("YBFpro MIDI", 1000),
    ("DT.YBF.1.100.170.7,5.3.50.380.DS.LCD", 1000),
    ("DT.YBF.1.150.170.11.2.50.380.DS.LCD", 1500),
    ("DT.YBF.1.200.170.15.2.50.380.DS.LCD", 2000),
    ("DT.YBF.2.300.170.22.2.50.380.DS.LCD", 3000),
    ("DT.YBF.2.350.170.30.2.50.380.DS.LCD", 3500),
    ("DT.YBF.2.400.170.30.2.50.380.DS.LCD", 4000),
    ("DT.YBF.2.450.170.40.2.50.380.DS.LCD", 4500),
    ("DT.YBF.2.500.170.40.2.50.380.DS.LCD", 5000),
]

DAVLUMBAZ_TYPE_FACTORS = {
    "Duz Davlumbaz": 1.0,
    "Flansli Davlumbaz": 0.75,
}

DAVLUMBAZ_CAPTURE_OPTIONS = {
    "Hafif yayilim (0,3 - 0,5 m/s)": 0.4,
    "Dusuk hizli yayilim (0,5 - 1,0 m/s)": 0.75,
    "Aktif olusum (1,0 - 2,5 m/s)": 1.75,
    "Yuksek hizli yayilim (2,5 - 10 m/s)": 4.0,
}

WELDING_FUME_MATRIX = {
    "SC Ar/CO2": 2.3,
    "SC CO2": 4.5,
    "STT Ar/CO2": 3.6,
    "RMD Ar/CO2": 3.3,
    "CMT Ar/CO2": 0.9,
    "GT Ar/CO2": 3.5,
    "GT CO2": 4.3,
    "AXS Ar/CO2": 6.4,
    "AX-P Ar/CO2": 1.5,
    "FCAW Ar/CO2": 4.7,
    "SMAW": 12.5,
}


def parse_float(value: str) -> float | None:
    raw = str(value).strip()
    if not raw:
        return None

    normalized = raw.replace(" ", "")
    if normalized.count(",") == 1 and "." not in normalized:
        normalized = normalized.replace(",", ".")
    elif normalized.count(",") > 1 and "." not in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", "")

    try:
        return float(normalized)
    except Exception:
        return None


def format_number(value: float, digits: int = 2) -> str:
    formatted = f"{value:,.{digits}f}"
    formatted = formatted.replace(",", "_").replace(".", ",").replace("_", ".")
    if digits > 0:
        formatted = formatted.rstrip("0").rstrip(",")
    return formatted


def create_labeled_entry(
    parent: ctk.CTkFrame,
    row: int,
    column: int,
    label: str,
    variable: ctk.StringVar,
) -> None:
    cell = ctk.CTkFrame(parent, fg_color="white")
    cell.grid(row=row, column=column, sticky="ew", padx=14, pady=(0, 12))
    cell.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        cell,
        text=label,
        font=ctk.CTkFont(size=13),
        text_color="#555555",
    ).grid(row=0, column=0, sticky="w", pady=(0, 6))

    ctk.CTkEntry(
        cell,
        textvariable=variable,
        height=38,
        corner_radius=10,
        fg_color="#fafafa",
        border_color="#d6d6d6",
        text_color="#212121",
    ).grid(row=1, column=0, sticky="ew")


def create_labeled_combobox(
    parent: ctk.CTkFrame,
    row: int,
    column: int,
    label: str,
    variable: ctk.StringVar,
    values: list[str],
    command=None,
) -> None:
    cell = ctk.CTkFrame(parent, fg_color="white")
    cell.grid(row=row, column=column, sticky="ew", padx=14, pady=(0, 12))
    cell.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        cell,
        text=label,
        font=ctk.CTkFont(size=13),
        text_color="#555555",
    ).grid(row=0, column=0, sticky="w", pady=(0, 6))

    ctk.CTkComboBox(
        cell,
        variable=variable,
        values=values,
        height=38,
        corner_radius=10,
        fg_color="#fafafa",
        border_color="#d6d6d6",
        button_color="#d32f2f",
        button_hover_color="#c62828",
        dropdown_fg_color="white",
        dropdown_hover_color="#f5f5f5",
        dropdown_text_color="#212121",
        text_color="#212121",
        command=command,
    ).grid(row=1, column=0, sticky="ew")


def create_labeled_entry_with_unit(
    parent: ctk.CTkFrame,
    row: int,
    column: int,
    label: str,
    variable: ctk.StringVar,
    unit: str,
) -> None:
    cell = ctk.CTkFrame(parent, fg_color="white")
    cell.grid(row=row, column=column, sticky="ew", padx=14, pady=(0, 12))
    cell.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        cell,
        text=label,
        font=ctk.CTkFont(size=13),
        text_color="#555555",
    ).grid(row=0, column=0, sticky="w", pady=(0, 6))

    input_row = ctk.CTkFrame(cell, fg_color="white")
    input_row.grid(row=1, column=0, sticky="ew")
    input_row.grid_columnconfigure(0, weight=1)

    ctk.CTkEntry(
        input_row,
        textvariable=variable,
        height=38,
        corner_radius=10,
        fg_color="#fafafa",
        border_color="#d6d6d6",
        text_color="#212121",
    ).grid(row=0, column=0, sticky="ew")

    ctk.CTkLabel(
        input_row,
        text=unit,
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color="#777777",
    ).grid(row=0, column=1, sticky="w", padx=(10, 0))


def open_isil_kesim_submodule(
    module: dict[str, object],
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    defaults = module["defaults"]
    pencere = ctk.CTkToplevel(parent)
    pencere.title(str(module["title"]))
    pencere.geometry("1180x860")
    pencere.minsize(1000, 760)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1000, min_height=760)

    cutting_power_var = ctk.StringVar(value=str(defaults["cutting_power"]))
    valve_tightness_var = ctk.StringVar(value=str(defaults["valve_tightness"]))
    valve_count_var = ctk.StringVar(value=str(defaults["valve_count"]))
    module_width_var = ctk.StringVar(value=str(defaults["module_width_mm"]))
    module_length_var = ctk.StringVar(value=str(defaults["module_length_mm"]))
    reserve_var = ctk.StringVar(value=str(defaults["reserve"]))

    base_emission_factor_var = ctk.StringVar(value="-")
    valve_leakage_var = ctk.StringVar(value="-")
    module_area_var = ctk.StringVar(value="-")
    process_flow_var = ctk.StringVar(value="-")
    leakage_flow_var = ctk.StringVar(value="-")
    required_capacity_var = ctk.StringVar(value="-")
    reserved_capacity_var = ctk.StringVar(value="-")

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 14))
    ctk.CTkLabel(
        header,
        text=str(module["title"]),
        font=ctk.CTkFont(family="Inter", size=26, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text=str(module["description"]),
        font=ctk.CTkFont(family="Inter", size=13),
        text_color="#666666",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w")

    content = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    content.pack(fill="both", expand=True)
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)

    inputs_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    inputs_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    inputs_card.grid_columnconfigure(0, weight=1)
    inputs_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        inputs_card,
        text="Girdiler",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def update_lookup_preview(_selected: str | None = None) -> None:
        selected_power = cutting_power_var.get()
        selected_tightness = valve_tightness_var.get()
        base_emission_factor_var.set(
            format_number(float(THERMAL_CUTTING_POWER_OPTIONS.get(selected_power, 0)), 0)
        )
        valve_leakage_var.set(
            format_number(float(VALVE_TIGHTNESS_OPTIONS.get(selected_tightness, 0)), 0)
        )

    create_labeled_combobox(
        inputs_card,
        1,
        0,
        "Kesim gucu",
        cutting_power_var,
        list(THERMAL_CUTTING_POWER_OPTIONS.keys()),
        command=update_lookup_preview,
    )
    create_labeled_combobox(
        inputs_card,
        1,
        1,
        "Klape sizdirmazligi",
        valve_tightness_var,
        list(VALVE_TIGHTNESS_OPTIONS.keys()),
        command=update_lookup_preview,
    )
    create_labeled_entry(inputs_card, 2, 0, "Valf adedi", valve_count_var)
    create_labeled_entry(inputs_card, 2, 1, "Modul genisligi (mm)", module_width_var)
    create_labeled_entry(inputs_card, 3, 0, "Emniyet / rezerv payi (%)", reserve_var)
    create_labeled_entry(inputs_card, 3, 1, "Modul uzunlugu (mm)", module_length_var)

    factors_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    factors_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    factors_card.grid_columnconfigure(0, weight=1)
    factors_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        factors_card,
        text="Secime Bagli Katsayilar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_factor_row(row: int, label: str, variable: ctk.StringVar, suffix: str) -> None:
        ctk.CTkLabel(
            factors_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            factors_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))
        ctk.CTkLabel(
            factors_card,
            text=suffix,
            font=ctk.CTkFont(size=12),
            text_color="#777777",
        ).grid(row=row, column=2, sticky="w", padx=(0, 16), pady=(0, 10))

    add_factor_row(1, "Temel emis katsayisi", base_emission_factor_var, "m3/m2/h")
    add_factor_row(2, "Klape sizintisi", valve_leakage_var, "m3/h / valf")

    results_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    results_card.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(8, 0), pady=(0, 14))
    results_card.grid_columnconfigure(0, weight=1)
    results_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        results_card,
        text="Sonuclar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_result_row(row: int, label: str, variable: ctk.StringVar) -> None:
        ctk.CTkLabel(
            results_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            results_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))

    add_result_row(1, "Modul alani", module_area_var)
    add_result_row(2, "Proses debisi", process_flow_var)
    add_result_row(3, "Klape sizinti debisi", leakage_flow_var)
    add_result_row(4, "Gerekli kapasite", required_capacity_var)
    add_result_row(5, "Rezerve edilmis kapasite", reserved_capacity_var)

    def hesapla() -> None:
        selected_power = cutting_power_var.get()
        selected_tightness = valve_tightness_var.get()
        base_emission_factor = float(THERMAL_CUTTING_POWER_OPTIONS.get(selected_power, 0))
        valve_leakage = float(VALVE_TIGHTNESS_OPTIONS.get(selected_tightness, 0))
        valve_count = parse_float(valve_count_var.get())
        module_width = parse_float(module_width_var.get())
        module_length = parse_float(module_length_var.get())
        reserve = parse_float(reserve_var.get())

        values = [valve_count, module_width, module_length, reserve]
        if any(value is None for value in values):
            messagebox.showerror(
                "Gecersiz veri",
                "Tum alanlara sayisal bir deger girin.",
                parent=pencere,
            )
            return

        assert valve_count is not None
        assert module_width is not None
        assert module_length is not None
        assert reserve is not None

        if min(base_emission_factor, valve_leakage, valve_count, module_width, module_length, reserve) < 0:
            messagebox.showerror(
                "Gecersiz veri",
                "Negatif deger kullanilamaz.",
                parent=pencere,
            )
            return

        module_area_m2 = (module_width * module_length) / 1_000_000.0
        process_flow = base_emission_factor * module_area_m2
        leakage_flow = valve_count * valve_leakage
        required_capacity = process_flow + leakage_flow
        reserved_capacity = required_capacity * (1 + reserve / 100.0)

        update_lookup_preview()
        module_area_var.set(f"{format_number(module_area_m2, 3)} m2")
        process_flow_var.set(f"{format_number(process_flow)} m3/h")
        leakage_flow_var.set(f"{format_number(leakage_flow)} m3/h")
        required_capacity_var.set(f"{format_number(required_capacity)} m3/h")
        reserved_capacity_var.set(f"{format_number(reserved_capacity)} m3/h")

    actions = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    actions.pack(fill="x", pady=(12, 0))

    ctk.CTkButton(
        actions,
        text="Kapat",
        width=120,
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=pencere.destroy,
    ).pack(side="right")

    update_lookup_preview()
    hesapla()


def open_cnc_torna_submodule(
    module: dict[str, object],
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    defaults = module["defaults"]
    pencere = ctk.CTkToplevel(parent)
    pencere.title(str(module["title"]))
    pencere.geometry("1240x860")
    pencere.minsize(1040, 760)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1040, min_height=760)

    coolant_type_var = ctk.StringVar(value=str(defaults["coolant_type"]))
    spindle_speed_var = ctk.StringVar(value=str(defaults["max_spindle_speed"]))
    pump_pressure_var = ctk.StringVar(value=str(defaults["coolant_pump_pressure"]))
    processing_space_var = ctk.StringVar(value=str(defaults["processing_space"]))
    door_type_var = ctk.StringVar(value=str(defaults["door_type"]))
    door_opening_area_var = ctk.StringVar(value=str(defaults["door_opening_area"]))
    cycle_time_var = ctk.StringVar(value=str(defaults["cycle_time"]))
    viscosity_var = ctk.StringVar(value=str(defaults["viscosity"]))

    required_airflow_var = ctk.StringVar(value="-")

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 14))
    ctk.CTkLabel(
        header,
        text=str(module["title"]),
        font=ctk.CTkFont(family="Inter", size=26, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text=str(module["description"]),
        font=ctk.CTkFont(family="Inter", size=13),
        text_color="#666666",
        wraplength=1100,
        justify="left",
    ).pack(anchor="w")

    content = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    content.pack(fill="both", expand=True)
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)

    inputs_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    inputs_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    inputs_card.grid_columnconfigure(0, weight=1)
    inputs_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        inputs_card,
        text="Girdiler",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def update_default_viscosity(_selected: str | None = None) -> None:
        viscosity_var.set(CNC_DEFAULT_VISCOSITY.get(coolant_type_var.get(), "1"))

    create_labeled_combobox(
        inputs_card,
        1,
        0,
        "Sogutma sivisi tipi",
        coolant_type_var,
        CNC_COOLANT_TYPE_OPTIONS,
        command=update_default_viscosity,
    )
    create_labeled_entry(inputs_card, 1, 1, "Maks. is mili devri (rpm)", spindle_speed_var)
    create_labeled_entry(inputs_card, 2, 0, "Sogutma pompasi basinci (bar)", pump_pressure_var)
    create_labeled_entry_with_unit(inputs_card, 2, 1, "Isleme hacmi", processing_space_var, "m3")
    create_labeled_combobox(
        inputs_card,
        3,
        0,
        "Kapi tipi",
        door_type_var,
        CNC_DOOR_TYPE_OPTIONS,
    )
    create_labeled_entry_with_unit(inputs_card, 3, 1, "Kapi aciklik alani (toplam)", door_opening_area_var, "m2")
    create_labeled_entry(inputs_card, 4, 0, "Cevrim suresi (dk)", cycle_time_var)
    create_labeled_entry_with_unit(inputs_card, 4, 1, "Viskozite", viscosity_var, "cSt")

    info_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    info_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    ctk.CTkLabel(
        info_card,
        text="Aciklamalar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 10))
    ctk.CTkLabel(
        info_card,
        text="1. Isleme Hacmi:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="Islem yapilan tezgahin ic hacmi.",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 8))
    ctk.CTkLabel(
        info_card,
        text="2. Kapi Aciklik Alani:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="Toplam kapi aciklik alani",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 8))
    ctk.CTkLabel(
        info_card,
        text="3. Viskozite:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text=(
            "Kullanilan sivi tipine gore degiskenlik gosterir. "
            "Girilen deger, sogutucu tipine gore standart olarak verilmistir. "
            "Kullanilan sogutucunun MSDS degerinden kontrol ediniz."
        ),
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    results_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    results_card.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(8, 0), pady=(0, 14))
    results_card.grid_columnconfigure(0, weight=1)
    results_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        results_card,
        text="Sonuclar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_result_row(row: int, label: str, variable: ctk.StringVar) -> None:
        ctk.CTkLabel(
            results_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            results_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
            wraplength=360,
            justify="right",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))

    add_result_row(1, "Gerekli hava debisi", required_airflow_var)

    def hesapla() -> None:
        spindle_speed = parse_float(spindle_speed_var.get())
        pump_pressure = parse_float(pump_pressure_var.get())
        processing_space = parse_float(processing_space_var.get())
        door_opening_area = parse_float(door_opening_area_var.get())
        cycle_time = parse_float(cycle_time_var.get())
        viscosity = parse_float(viscosity_var.get())

        values = [
            spindle_speed,
            pump_pressure,
            processing_space,
            door_opening_area,
            cycle_time,
            viscosity,
        ]
        if any(value is None for value in values):
            messagebox.showerror(
                "Gecersiz veri",
                "Tum alanlara sayisal bir deger girin.",
                parent=pencere,
            )
            return

        assert spindle_speed is not None
        assert pump_pressure is not None
        assert processing_space is not None
        assert door_opening_area is not None
        assert cycle_time is not None
        assert viscosity is not None

        if min(spindle_speed, pump_pressure, processing_space, door_opening_area, cycle_time, viscosity) <= 0:
            messagebox.showerror(
                "Gecersiz veri",
                "Tum sayisal alanlar sifirdan buyuk olmalidir.",
                parent=pencere,
            )
            return

        from math import log

        spindle_pressure_factor = (0.22 * log(spindle_speed * pump_pressure) - 2.4) * 2
        processing_space_multiplier = 1 + (processing_space - 1) * 0.5

        door_type = door_type_var.get()
        if door_type == "Sliding Door, Vertical":
            door_type_adjustment = -100.0
        elif door_type == "Sliding Door, Horizontal":
            door_type_adjustment = 0.0
        else:
            door_type_adjustment = 250.0

        door_opening_multiplier = ((door_opening_area - 0.5) * 0.08) * 2
        cycle_factor_raw = -0.0653 * log(cycle_time) + 0.1419
        cycle_factor = max(0.0, cycle_factor_raw)
        viscosity_factor = max(1.0, -0.00002 * (viscosity**2) + 0.0313 * viscosity + 0.7349)

        airflow_before_safety = (
            (
                processing_space_multiplier
                * 3600
                * 0.2
                * (1 + spindle_pressure_factor)
                * (1 + cycle_factor)
                * (1 + door_opening_multiplier)
                + door_type_adjustment
            )
            / viscosity_factor
        )
        required_airflow = max(150.0, airflow_before_safety) / 1.3

        coolant_type = coolant_type_var.get()
        use_hepa = coolant_type == "Straight Oil"
        product_list = CNC_HEPA_PRODUCTS if use_hepa else CNC_STANDARD_PRODUCTS

        for product_name, threshold in product_list:
            normalized_threshold = threshold / viscosity_factor
            if required_airflow <= normalized_threshold:
                break

        required_airflow_var.set(f"{format_number(required_airflow)} m3/h")

    actions = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    actions.pack(fill="x", pady=(12, 0))

    ctk.CTkButton(
        actions,
        text="Kapat",
        width=120,
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=pencere.destroy,
    ).pack(side="right")

    update_default_viscosity()
    hesapla()


def open_capacity_submodule(
    module: dict[str, object],
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    if module.get("key") == "isil_kesim":
        open_isil_kesim_submodule(module, parent)
        return
    if module.get("key") == "cnc_torna":
        open_cnc_torna_submodule(module, parent)
        return
    if module.get("key") == "davlumbaz":
        open_davlumbaz_submodule(module, parent)
        return
    if module.get("key") == "genel_hol":
        open_kaynak_hol_submodule(module, parent)
        return

    defaults = module["defaults"]
    pencere = ctk.CTkToplevel(parent)
    pencere.title(str(module["title"]))
    pencere.geometry("1180x860")
    pencere.minsize(1000, 760)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1000, min_height=760)

    source_count_var = ctk.StringVar(value=str(defaults["source_count"]))
    unit_airflow_var = ctk.StringVar(value=str(defaults["unit_airflow"]))
    simultaneity_var = ctk.StringVar(value=str(defaults["simultaneity"]))
    area_var = ctk.StringVar(value=str(defaults["area"]))
    height_var = ctk.StringVar(value=str(defaults["height"]))
    air_change_var = ctk.StringVar(value=str(defaults["air_change"]))
    reserve_var = ctk.StringVar(value=str(defaults["reserve"]))
    hours_var = ctk.StringVar(value=str(defaults["hours"]))

    local_result_var = ctk.StringVar(value="-")
    general_result_var = ctk.StringVar(value="-")
    base_total_var = ctk.StringVar(value="-")
    reserve_total_var = ctk.StringVar(value="-")
    daily_capacity_var = ctk.StringVar(value="-")
    summary_var = ctk.StringVar(value="Degerleri girip Hesapla butonuna basin.")

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 14))
    ctk.CTkLabel(
        header,
        text=str(module["title"]),
        font=ctk.CTkFont(family="Inter", size=26, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text=str(module["description"]),
        font=ctk.CTkFont(family="Inter", size=13),
        text_color="#666666",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w")

    content = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    content.pack(fill="both", expand=True)
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)

    info_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    info_card.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 14))
    ctk.CTkLabel(
        info_card,
        text="Hesap yaklasimi",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 6))
    ctk.CTkLabel(
        info_card,
        text=(
            "Toplam kapasite = (Lokal emis debisi + genel hacim havalandirmasi) x (1 + emniyet payi). "
            "Bu ekran, teklif ve on boyutlandirma icin hizli kapasite tahmini verir."
        ),
        font=ctk.CTkFont(size=13),
        text_color="#666666",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    inputs_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    inputs_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    inputs_card.grid_columnconfigure(0, weight=1)
    inputs_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        inputs_card,
        text="Girdiler",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    create_labeled_entry(inputs_card, 1, 0, "Kaynak / ekipman adedi", source_count_var)
    create_labeled_entry(inputs_card, 1, 1, "Birim lokal debi (m3/h)", unit_airflow_var)
    create_labeled_entry(inputs_card, 2, 0, "Eszamanlilik katsayisi", simultaneity_var)
    create_labeled_entry(inputs_card, 2, 1, "Alan (m2)", area_var)
    create_labeled_entry(inputs_card, 3, 0, "Yukseklik (m)", height_var)
    create_labeled_entry(inputs_card, 3, 1, "Saatlik hava degisim sayisi (ACH)", air_change_var)
    create_labeled_entry(inputs_card, 4, 0, "Emniyet / rezerv payi (%)", reserve_var)
    create_labeled_entry(inputs_card, 4, 1, "Gunluk calisma suresi (saat)", hours_var)

    results_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    results_card.grid(row=1, column=1, sticky="nsew", padx=(8, 0), pady=(0, 14))
    results_card.grid_columnconfigure(0, weight=1)
    results_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        results_card,
        text="Sonuclar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_result_row(row: int, label: str, variable: ctk.StringVar) -> None:
        ctk.CTkLabel(
            results_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            results_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))

    add_result_row(1, "Lokal emis debisi", local_result_var)
    add_result_row(2, "Genel hacim debisi", general_result_var)
    add_result_row(3, "Baz toplam debi", base_total_var)
    add_result_row(4, "Rezerve edilmis toplam", reserve_total_var)
    add_result_row(5, "Gunluk hava hacmi", daily_capacity_var)

    summary_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    summary_card.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 14))
    ctk.CTkLabel(
        summary_card,
        text="Yorum",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 6))
    ctk.CTkLabel(
        summary_card,
        textvariable=summary_var,
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    def hesapla() -> None:
        source_count = parse_float(source_count_var.get())
        unit_airflow = parse_float(unit_airflow_var.get())
        simultaneity = parse_float(simultaneity_var.get())
        area = parse_float(area_var.get())
        height = parse_float(height_var.get())
        air_change = parse_float(air_change_var.get())
        reserve = parse_float(reserve_var.get())
        hours = parse_float(hours_var.get())

        values = [
            source_count,
            unit_airflow,
            simultaneity,
            area,
            height,
            air_change,
            reserve,
            hours,
        ]
        if any(value is None for value in values):
            messagebox.showerror(
                "Gecersiz veri",
                "Tum alanlara sayisal bir deger girin.",
                parent=pencere,
            )
            return

        assert source_count is not None
        assert unit_airflow is not None
        assert simultaneity is not None
        assert area is not None
        assert height is not None
        assert air_change is not None
        assert reserve is not None
        assert hours is not None

        if min(source_count, unit_airflow, simultaneity, area, height, air_change, reserve, hours) < 0:
            messagebox.showerror(
                "Gecersiz veri",
                "Negatif deger kullanilamaz.",
                parent=pencere,
            )
            return

        local_flow = source_count * unit_airflow * simultaneity
        general_flow = area * height * air_change
        base_total = local_flow + general_flow
        reserved_total = base_total * (1 + reserve / 100.0)
        daily_capacity = reserved_total * hours

        local_result_var.set(f"{format_number(local_flow)} m3/h")
        general_result_var.set(f"{format_number(general_flow)} m3/h")
        base_total_var.set(f"{format_number(base_total)} m3/h")
        reserve_total_var.set(f"{format_number(reserved_total)} m3/h")
        daily_capacity_var.set(f"{format_number(daily_capacity)} m3/gun")

        summary_var.set(
            f"Lokal ihtiyac {format_number(local_flow)} m3/h ve mahal ihtiyaci "
            f"{format_number(general_flow)} m3/h olarak bulundu. "
            f"%{format_number(reserve)} rezerv ile onerilen fan/sistem kapasitesi "
            f"{format_number(reserved_total)} m3/h seviyesindedir."
        )

    actions = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    actions.pack(fill="x", pady=(12, 0))

    ctk.CTkButton(
        actions,
        text="Kapat",
        width=120,
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=pencere.destroy,
    ).pack(side="right")

    hesapla()


def open_kaynak_hol_submodule(
    module: dict[str, object],
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    defaults = module["defaults"]
    pencere = ctk.CTkToplevel(parent)
    pencere.title(str(module["title"]))
    pencere.geometry("1200x860")
    pencere.minsize(1020, 760)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1020, min_height=760)

    welding_type_var = ctk.StringVar(value=str(defaults["welding_type"]))
    manual_welder_count_var = ctk.StringVar(value=str(defaults["manual_welder_count"]))
    manual_wire_consumption_var = ctk.StringVar(value=str(defaults["manual_wire_consumption"]))
    robot_count_var = ctk.StringVar(value=str(defaults["robot_count"]))
    robot_wire_consumption_var = ctk.StringVar(value=str(defaults["robot_wire_consumption"]))
    hall_depth_var = ctk.StringVar(value=str(defaults["hall_depth"]))
    hall_width_var = ctk.StringVar(value=str(defaults["hall_width"]))
    hall_height_var = ctk.StringVar(value=str(defaults["height"]))
    particle_limit_var = ctk.StringVar(value=str(defaults["particle_limit"]))
    requested_ach_var = ctk.StringVar(value=str(defaults["requested_ach"]))

    fume_factor_var = ctk.StringVar(value="-")
    total_wire_var = ctk.StringVar(value="-")
    total_particle_var = ctk.StringVar(value="-")
    hall_volume_var = ctk.StringVar(value="-")
    particle_ratio_var = ctk.StringVar(value="-")
    minimum_ach_var = ctk.StringVar(value="-")
    particle_value_var = ctk.StringVar(value="-")
    required_capacity_var = ctk.StringVar(value="-")
    summary_var = ctk.StringVar(value="Temel verileri girin. Minimum ACH otomatik hesaplanir.")
    current_result: dict[str, float | str] = {}

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 14))
    ctk.CTkLabel(
        header,
        text=str(module["title"]),
        font=ctk.CTkFont(family="Inter", size=26, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text=str(module["description"]),
        font=ctk.CTkFont(family="Inter", size=13),
        text_color="#666666",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w")

    content = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    content.pack(fill="both", expand=True)
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)

    info_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    info_card.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 14))
    ctk.CTkLabel(
        info_card,
        text="Hesap yaklasimi",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 6))
    ctk.CTkLabel(
        info_card,
        text=(
            "Toplam tel tuketimi kaynak tipine ait referans duman katsayisi ile carpilarak "
            "saatlik partikul yuku bulunur. Bu yuk hol hacmine bolunur, limit deger ile "
            "minimum ACH hesaplanir ve secilen ACH ile gerekli kapasite elde edilir."
        ),
        font=ctk.CTkFont(size=13),
        text_color="#666666",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    ctk.CTkLabel(
        info_card,
        text="Kaynak tipi kisaltmalari",
        font=ctk.CTkFont(size=14, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 10))

    abbreviation_items = [
        ("SC Ar/CO2", "Kisa devre MIG/MAG kaynagi (argon + CO2 gazli). Halk arasinda kisa devre gazalti."),
        ("SC CO2", "Kisa devre CO2 gazalti kaynagi. Genelde CO2 gazalti, saf CO2 ile gazalti."),
        ("STT Ar/CO2", "STT kontrollu kisa devre gazalti kaynagi. Surface Tension Transfer, sahada STT kaynak."),
        ("RMD Ar/CO2", "RMD kontrollu gazalti kaynagi. Regulated Metal Deposition, sahada RMD kaynak."),
        ("CMT Ar/CO2", "CMT dusuk isi girdili gazalti kaynagi. Cold Metal Transfer, sahada CMT kaynak."),
        ("GT Ar/CO2", "Globuler transfer gazalti kaynagi (argon/CO2 karisim gazli). Sahada globuler mod."),
        ("GT CO2", "Globuler transfer CO2 gazalti kaynagi. Sahada CO2 globuler."),
        ("AXS Ar/CO2", "Aksiyel sprey transfer gazalti kaynagi (spray arc). Sahada sprey transfer, spray ark."),
        ("AX-P Ar/CO2", "Pulse sprey (darbeli sprey) gazalti kaynagi. Sahada pulse MIG/MAG, darbeli gazalti."),
        ("FCAW Ar/CO2", "Ozlu tel gazalti kaynagi (flux-cored wire). Sahada ozlu tel kaynagi."),
        ("SMAW", "Ortulu elektrod kaynagi. Sahada elektrod kaynagi, ark kaynagi."),
    ]

    for short_name, description in abbreviation_items:
        item_frame = ctk.CTkFrame(info_card, fg_color="white")
        item_frame.pack(fill="x", padx=16, pady=(0, 10))
        ctk.CTkLabel(
            item_frame,
            text=short_name,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#212121",
        ).pack(anchor="w")
        ctk.CTkLabel(
            item_frame,
            text=description,
            font=ctk.CTkFont(size=12),
            text_color="#555555",
            wraplength=1040,
            justify="left",
        ).pack(anchor="w", pady=(2, 0))

    inputs_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    inputs_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    inputs_card.grid_columnconfigure(0, weight=1)
    inputs_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        inputs_card,
        text="Girdiler",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    create_labeled_combobox(
        inputs_card,
        1,
        0,
        "Kaynak tipi",
        welding_type_var,
        list(WELDING_FUME_MATRIX.keys()),
    )
    create_labeled_entry(inputs_card, 1, 1, "Ayni anda calisan manuel kaynakci", manual_welder_count_var)
    create_labeled_entry_with_unit(
        inputs_card, 2, 0, "Manuel kaynakci tel tuketimi", manual_wire_consumption_var, "kg/h"
    )
    create_labeled_entry(inputs_card, 2, 1, "Ayni anda calisan robot sayisi", robot_count_var)
    create_labeled_entry_with_unit(
        inputs_card, 3, 0, "Robot tel tuketimi", robot_wire_consumption_var, "kg/h"
    )
    create_labeled_entry_with_unit(inputs_card, 3, 1, "Partikul limit degeri", particle_limit_var, "mg/Nm3")
    create_labeled_entry_with_unit(inputs_card, 4, 0, "Gerekli ACH", requested_ach_var, "1/h")

    hall_cell = ctk.CTkFrame(inputs_card, fg_color="white")
    hall_cell.grid(row=4, column=1, sticky="ew", padx=14, pady=(0, 12))
    for column_index in range(3):
        hall_cell.grid_columnconfigure(column_index, weight=1)
    ctk.CTkLabel(
        hall_cell,
        text="Hol olculeri",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
    ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))

    for idx, (label, variable) in enumerate(
        [
            ("Derinlik", hall_depth_var),
            ("Genislik", hall_width_var),
            ("Yukseklik", hall_height_var),
        ]
    ):
        item = ctk.CTkFrame(hall_cell, fg_color="white")
        item.grid(row=1, column=idx, sticky="ew", padx=(0 if idx == 0 else 8, 0))
        item.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            item,
            text=label,
            font=ctk.CTkFont(size=12),
            text_color="#777777",
        ).grid(row=0, column=0, sticky="w", pady=(0, 4))
        input_row = ctk.CTkFrame(item, fg_color="white")
        input_row.grid(row=1, column=0, sticky="ew")
        input_row.grid_columnconfigure(0, weight=1)
        ctk.CTkEntry(
            input_row,
            textvariable=variable,
            height=38,
            corner_radius=10,
            fg_color="#fafafa",
            border_color="#d6d6d6",
            text_color="#212121",
        ).grid(row=0, column=0, sticky="ew")
        ctk.CTkLabel(
            input_row,
            text="m",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#777777",
        ).grid(row=0, column=1, sticky="w", padx=(10, 0))

    results_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    results_card.grid(row=1, column=1, sticky="nsew", padx=(8, 0), pady=(0, 14))
    results_card.grid_columnconfigure(0, weight=1)
    results_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        results_card,
        text="Sonuclar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_result_row(row: int, label: str, variable: ctk.StringVar) -> None:
        ctk.CTkLabel(
            results_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            results_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))

    add_result_row(1, "Referans duman katsayisi", fume_factor_var)
    add_result_row(2, "Toplam tel tuketimi", total_wire_var)
    add_result_row(3, "Toplam partikul", total_particle_var)
    add_result_row(4, "Hol hacmi", hall_volume_var)
    add_result_row(5, "Partikul orani", particle_ratio_var)
    add_result_row(6, "Minimum ACH", minimum_ach_var)
    add_result_row(7, "Secilen ACH'te partikul", particle_value_var)
    add_result_row(8, "Hesaplanan kapasite", required_capacity_var)

    summary_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    summary_card.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 14))
    ctk.CTkLabel(
        summary_card,
        text="Yorum",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 6))
    ctk.CTkLabel(
        summary_card,
        textvariable=summary_var,
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=1040,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    def clear_results(message: str = "Temel verileri girin. Minimum ACH otomatik hesaplanir.") -> None:
        for variable in (
            fume_factor_var,
            total_wire_var,
            total_particle_var,
            hall_volume_var,
            particle_ratio_var,
            minimum_ach_var,
            particle_value_var,
            required_capacity_var,
        ):
            variable.set("-")
        current_result.clear()
        summary_var.set(message)

    def export_pdf() -> None:
        if not current_result.get("is_complete"):
            messagebox.showwarning(
                "PDF Hazir Degil",
                "PDF olusturmak icin once gecerli bir Gerekli ACH degeri girin.",
                parent=pencere,
            )
            return

        project_name = simpledialog.askstring(
            "Ilgili Proje Adi",
            "Lutfen ilgili proje adini girin:",
            parent=pencere,
        )
        if project_name is None:
            return
        project_name = project_name.strip()
        if not project_name:
            messagebox.showwarning(
                "Proje Adi Gerekli",
                "PDF olusturmak icin ilgili proje adini girmeniz gerekiyor.",
                parent=pencere,
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Kaynak Hol Havalandirma Raporu Kaydet",
            initialfile="kaynak_hol_havalandirma_raporu.pdf",
        )
        if not path:
            return

        try:
            import os
            from datetime import datetime
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen import canvas
        except ImportError:
            messagebox.showerror("PDF Disa Aktarma", "PDF olusturmak icin reportlab kutuphanesi bulunamadi.")
            return

        export_user = str(get_username() or os.environ.get("USERNAME") or "Bilinmeyen Kullanici").strip()

        try:
            regular_font_name = "Helvetica"
            bold_font_name = "Helvetica-Bold"
            possible_font_pairs = [
                (
                    os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arial.ttf"),
                    os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "arialbd.ttf"),
                ),
                (
                    os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "DejaVuSans.ttf"),
                    os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "DejaVuSans-Bold.ttf"),
                ),
            ]
            for regular_path, bold_path in possible_font_pairs:
                if os.path.exists(regular_path) and os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont("KaynakHolPdfFont", regular_path))
                    pdfmetrics.registerFont(TTFont("KaynakHolPdfFontBold", bold_path))
                    regular_font_name = "KaynakHolPdfFont"
                    bold_font_name = "KaynakHolPdfFontBold"
                    break

            pdf = canvas.Canvas(path, pagesize=A4)
            width, height = A4
            y = height - 50

            def ensure_space(lines_needed: int = 2) -> None:
                nonlocal y
                if y < 60 + (lines_needed * 16):
                    pdf.showPage()
                    y = height - 50

            pdf.setTitle("Kaynak Hol Havalandirma Raporu")
            pdf.setFont(bold_font_name, 16)
            pdf.drawString(40, y, "Kaynak Hol Havalandirma Raporu")
            y -= 26
            pdf.setFont(regular_font_name, 11)
            pdf.drawString(40, y, f"Ilgili Proje Adi: {project_name}")
            y -= 16
            pdf.drawString(40, y, f"Ciktiyi Alan Kullanici: {export_user}")
            y -= 16
            pdf.drawString(40, y, f"Olusturma Zamani: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
            y -= 24

            sections = [
                (
                    "Girdiler",
                    [
                        ("Kaynak tipi", str(current_result["welding_type"])),
                        ("Manuel kaynakci", str(current_result["manual_welder_count_display"])),
                        ("Manuel tel tuketimi", str(current_result["manual_wire_display"])),
                        ("Robot sayisi", str(current_result["robot_count_display"])),
                        ("Robot tel tuketimi", str(current_result["robot_wire_display"])),
                        ("Hol olculeri", str(current_result["hall_dimensions_display"])),
                        ("Partikul limit degeri", str(current_result["particle_limit_display"])),
                        ("Gerekli ACH", str(current_result["requested_ach_display"])),
                    ],
                ),
                (
                    "Sonuclar",
                    [
                        ("Referans duman katsayisi", str(current_result["fume_factor_display"])),
                        ("Toplam tel tuketimi", str(current_result["total_wire_display"])),
                        ("Toplam partikul", str(current_result["total_particle_display"])),
                        ("Hol hacmi", str(current_result["hall_volume_display"])),
                        ("Partikul orani", str(current_result["particle_ratio_display"])),
                        ("Minimum ACH", str(current_result["minimum_ach_display"])),
                        ("Secilen ACH'te partikul", str(current_result["particle_value_display"])),
                        ("Hesaplanan kapasite", str(current_result["required_capacity_display"])),
                    ],
                ),
            ]

            for section_title, rows in sections:
                ensure_space(len(rows) + 3)
                pdf.setFont(bold_font_name, 12)
                pdf.drawString(40, y, section_title)
                y -= 18
                for label, value in rows:
                    ensure_space()
                    pdf.setFont(bold_font_name, 11)
                    pdf.drawString(50, y, f"{label}:")
                    label_width = pdf.stringWidth(f"{label}:", bold_font_name, 11)
                    pdf.setFont(regular_font_name, 11)
                    pdf.drawString(56 + label_width, y, value)
                    y -= 15
                y -= 8

            ensure_space(4)
            pdf.setFont(bold_font_name, 12)
            pdf.drawString(40, y, "Degerlendirme")
            y -= 18
            pdf.setFont(regular_font_name, 11)
            for line in str(summary_var.get()).split(". "):
                text = line.strip()
                if not text:
                    continue
                ensure_space()
                pdf.drawString(50, y, text if text.endswith(".") else f"{text}.")
                y -= 15

            pdf.save()
            messagebox.showinfo("PDF Hazir", f"Rapor kaydedildi:\n{path}", parent=pencere)
        except Exception as exc:
            messagebox.showerror("PDF Disa Aktarma", f"PDF olusturulurken hata olustu:\n{exc}", parent=pencere)

    def hesapla() -> bool:
        manual_welder_count = parse_float(manual_welder_count_var.get())
        manual_wire_consumption = parse_float(manual_wire_consumption_var.get())
        robot_count = parse_float(robot_count_var.get())
        robot_wire_consumption = parse_float(robot_wire_consumption_var.get())
        hall_depth = parse_float(hall_depth_var.get())
        hall_width = parse_float(hall_width_var.get())
        hall_height = parse_float(hall_height_var.get())
        particle_limit = parse_float(particle_limit_var.get())
        requested_ach = parse_float(requested_ach_var.get())

        base_values = [
            manual_welder_count,
            manual_wire_consumption,
            robot_count,
            robot_wire_consumption,
            hall_depth,
            hall_width,
            hall_height,
            particle_limit,
        ]
        if any(value is None for value in base_values):
            clear_results()
            return False

        assert manual_welder_count is not None
        assert manual_wire_consumption is not None
        assert robot_count is not None
        assert robot_wire_consumption is not None
        assert hall_depth is not None
        assert hall_width is not None
        assert hall_height is not None
        assert particle_limit is not None

        if min(base_values) < 0:
            clear_results()
            return False

        if min(hall_depth, hall_width, hall_height, particle_limit) <= 0:
            clear_results()
            return False

        fume_factor = float(WELDING_FUME_MATRIX.get(welding_type_var.get(), 0.0))
        total_wire_consumption = (manual_welder_count * manual_wire_consumption) + (
            robot_count * robot_wire_consumption
        )
        total_particle = total_wire_consumption * fume_factor * 1000.0
        hall_volume = hall_depth * hall_width * hall_height
        particle_ratio = total_particle / hall_volume
        minimum_ach = particle_ratio / particle_limit

        fume_factor_var.set(f"{format_number(fume_factor, 1)} mg/g")
        total_wire_var.set(f"{format_number(total_wire_consumption, 2)} kg/h")
        total_particle_var.set(f"{format_number(total_particle, 0)} mg/h")
        hall_volume_var.set(f"{format_number(hall_volume, 2)} m3")
        particle_ratio_var.set(f"{format_number(particle_ratio, 2)} mg/(m3xh)")
        minimum_ach_var.set(f"{format_number(minimum_ach, 2)} 1/h")
        current_result.update(
            {
                "welding_type": welding_type_var.get(),
                "manual_welder_count_display": format_number(manual_welder_count, 0),
                "manual_wire_display": f"{format_number(manual_wire_consumption, 2)} kg/h",
                "robot_count_display": format_number(robot_count, 0),
                "robot_wire_display": f"{format_number(robot_wire_consumption, 2)} kg/h",
                "hall_dimensions_display": (
                    f"{format_number(hall_depth, 2)} m x {format_number(hall_width, 2)} m x "
                    f"{format_number(hall_height, 2)} m"
                ),
                "particle_limit_display": f"{format_number(particle_limit, 2)} mg/Nm3",
                "fume_factor_display": f"{format_number(fume_factor, 1)} mg/g",
                "total_wire_display": f"{format_number(total_wire_consumption, 2)} kg/h",
                "total_particle_display": f"{format_number(total_particle, 0)} mg/h",
                "hall_volume_display": f"{format_number(hall_volume, 2)} m3",
                "particle_ratio_display": f"{format_number(particle_ratio, 2)} mg/(m3xh)",
                "minimum_ach_display": f"{format_number(minimum_ach, 2)} 1/h",
                "is_complete": False,
            }
        )

        if not str(requested_ach_var.get()).strip():
            particle_value_var.set("-")
            required_capacity_var.set("-")
            summary_var.set(
                f"Minimum ACH {format_number(minimum_ach, 2)} 1/h olarak hesaplandi. "
                "Bu degerden buyuk veya esit bir Gerekli ACH girdiginizde kapasite otomatik hesaplanir."
            )
            return False

        if requested_ach is None or requested_ach <= 0:
            particle_value_var.set("-")
            required_capacity_var.set("-")
            summary_var.set(
                f"Minimum ACH {format_number(minimum_ach, 2)} 1/h olarak hesaplandi. "
                "Lutfen sifirdan buyuk bir Gerekli ACH girin."
            )
            return False

        if requested_ach < minimum_ach:
            particle_value_var.set("-")
            required_capacity_var.set("-")
            summary_var.set(
                f"Minimum ACH {format_number(minimum_ach, 2)} 1/h. Girilen Gerekli ACH bu degerden kucuk oldugu icin kapasite hesabi yapilmadi."
            )
            current_result["requested_ach_display"] = f"{format_number(requested_ach, 2)} 1/h"
            return False

        particle_value = particle_ratio / requested_ach
        required_capacity = hall_volume * requested_ach

        particle_value_var.set(f"{format_number(particle_value, 2)} mg/Nm3")
        required_capacity_var.set(f"{format_number(required_capacity, 0)} m3/h")

        current_result.update(
            {
                "requested_ach_display": f"{format_number(requested_ach, 2)} 1/h",
                "particle_value_display": f"{format_number(particle_value, 2)} mg/Nm3",
                "required_capacity_display": f"{format_number(required_capacity, 0)} m3/h",
                "is_complete": True,
            }
        )

        summary_var.set(
            f"{welding_type_var.get()} icin referans duman katsayisi {format_number(fume_factor, 1)} mg/g alindi. "
            f"Toplam tel tuketimi {format_number(total_wire_consumption, 2)} kg/h ve toplam partikul yuku "
            f"{format_number(total_particle, 0)} mg/h bulundu. Minimum ACH {format_number(minimum_ach, 2)} 1/h, "
            f"girilen ACH {format_number(requested_ach, 2)} 1/h ve onerilen kapasite {format_number(required_capacity, 0)} m3/h."
        )
        return True

    actions = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    actions.pack(fill="x", pady=(12, 0))

    ctk.CTkButton(
        actions,
        text="PDF Cikti Al",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        command=export_pdf,
    ).pack(side="right", padx=(0, 16))

    ctk.CTkButton(
        actions,
        text="Kapat",
        width=120,
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=pencere.destroy,
    ).pack(side="right", padx=(10, 0))

    for variable in (
        welding_type_var,
        manual_welder_count_var,
        manual_wire_consumption_var,
        robot_count_var,
        robot_wire_consumption_var,
        hall_depth_var,
        hall_width_var,
        hall_height_var,
        particle_limit_var,
        requested_ach_var,
    ):
        variable.trace_add("write", lambda *_args: hesapla())

    hesapla()


def open_davlumbaz_submodule(
    module: dict[str, object],
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    defaults = module["defaults"]
    pencere = ctk.CTkToplevel(parent)
    pencere.title(str(module["title"]))
    pencere.geometry("1200x840")
    pencere.minsize(1020, 740)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1020, min_height=740)

    hood_type_var = ctk.StringVar(value=str(defaults["hood_type"]))
    capture_profile_var = ctk.StringVar(value=str(defaults["capture_profile"]))
    capture_velocity_var = ctk.StringVar(value=str(defaults["capture_velocity"]))
    distance_var = ctk.StringVar(value=str(defaults["distance"]))
    width_var = ctk.StringVar(value=str(defaults["width"]))
    length_var = ctk.StringVar(value=str(defaults["length"]))
    reserve_var = ctk.StringVar(value=str(defaults["reserve"]))

    hood_factor_var = ctk.StringVar(value="-")
    area_var = ctk.StringVar(value="-")
    airflow_m3s_var = ctk.StringVar(value="-")
    airflow_m3h_var = ctk.StringVar(value="-")
    reserve_airflow_var = ctk.StringVar(value="-")
    quantity_var = ctk.StringVar(value="1")

    current_result: dict[str, object] = {}
    hood_rows: list[dict[str, str]] = []

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=24, pady=20)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 14))
    ctk.CTkLabel(
        header,
        text=str(module["title"]),
        font=ctk.CTkFont(family="Inter", size=26, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text=str(module["description"]),
        font=ctk.CTkFont(family="Inter", size=13),
        text_color="#666666",
        wraplength=1080,
        justify="left",
    ).pack(anchor="w")

    content = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    content.pack(fill="both", expand=True)
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)

    inputs_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    inputs_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    inputs_card.grid_columnconfigure(0, weight=1)
    inputs_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        inputs_card,
        text="Girdiler",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def update_capture_velocity(_selected: str | None = None) -> None:
        profile = capture_profile_var.get()
        capture_velocity_var.set(format_number(DAVLUMBAZ_CAPTURE_OPTIONS.get(profile, 0.75), 2))

    def update_hood_factor() -> None:
        hood_factor_var.set(format_number(DAVLUMBAZ_TYPE_FACTORS.get(hood_type_var.get(), 1.0), 2))

    create_labeled_combobox(
        inputs_card,
        1,
        0,
        "Davlumbaz tipi",
        hood_type_var,
        list(DAVLUMBAZ_TYPE_FACTORS.keys()),
        command=lambda _value: update_hood_factor(),
    )
    create_labeled_combobox(
        inputs_card,
        1,
        1,
        "Yakalama hizi onerisi",
        capture_profile_var,
        list(DAVLUMBAZ_CAPTURE_OPTIONS.keys()),
        command=update_capture_velocity,
    )
    create_labeled_entry_with_unit(inputs_card, 2, 0, "Yakalama hizi", capture_velocity_var, "m/s")
    create_labeled_entry_with_unit(inputs_card, 2, 1, "Mesafe - x", distance_var, "m")
    create_labeled_entry_with_unit(inputs_card, 3, 0, "Davlumbaz genisligi", width_var, "m")
    create_labeled_entry_with_unit(inputs_card, 3, 1, "Davlumbaz uzunlugu", length_var, "m")
    create_labeled_entry_with_unit(inputs_card, 4, 0, "Guvenlik payi", reserve_var, "%")

    info_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    info_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    ctk.CTkLabel(
        info_card,
        text="Aciklamalar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(16, 10))
    ctk.CTkLabel(
        info_card,
        text="1. Yakalama Hizi:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="Proses tipine gore secilir. Oneri listesinden gelen deger kullanici tarafindan degistirilebilir.",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 8))
    ctk.CTkLabel(
        info_card,
        text="2. Mesafe - x:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="Kirlilik kaynagi ile davlumbaz arasindaki mesafedir. Debiyi en cok etkileyen parametrelerden biridir.",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 8))
    ctk.CTkLabel(
        info_card,
        text="3. Alan:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="Davlumbazin emise baslayan kesit alanidir. Genislik x uzunluk olarak hesaplanir.",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 8))
    ctk.CTkLabel(
        info_card,
        text="4. Not:",
        font=ctk.CTkFont(size=13, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        info_card,
        text="PDF dokumanindaki tavsiye dogrultusunda sonuca %10-%15 guvenlik payi eklenmesi onerilir.",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
        wraplength=500,
        justify="left",
    ).pack(anchor="w", padx=16, pady=(0, 16))

    results_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    results_card.grid(row=2, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    results_card.grid_columnconfigure(0, weight=1)
    results_card.grid_columnconfigure(1, weight=1)

    ctk.CTkLabel(
        results_card,
        text="Sonuclar",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(16, 12))

    def add_result_row(row: int, label: str, variable: ctk.StringVar) -> None:
        ctk.CTkLabel(
            results_card,
            text=label,
            font=ctk.CTkFont(size=13),
            text_color="#555555",
        ).grid(row=row, column=0, sticky="w", padx=(16, 8), pady=(0, 10))
        ctk.CTkLabel(
            results_card,
            textvariable=variable,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#d32f2f",
        ).grid(row=row, column=1, sticky="e", padx=(8, 16), pady=(0, 10))

    add_result_row(1, "Davlumbaz katsayisi", hood_factor_var)
    add_result_row(2, "Kesit alani", area_var)
    add_result_row(3, "Hava debisi", airflow_m3s_var)
    add_result_row(4, "Hava debisi", airflow_m3h_var)
    add_result_row(5, "Guvenlik payli debi", reserve_airflow_var)

    add_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    add_card.grid(row=3, column=0, sticky="nsew", padx=(0, 8), pady=(0, 14))
    add_card.grid_columnconfigure(0, weight=1)
    add_card.grid_columnconfigure(1, weight=0)
    add_card.grid_columnconfigure(2, weight=0)

    ctk.CTkLabel(
        add_card,
        text="Listeye Ekle",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=16, pady=(16, 12))

    qty_cell = ctk.CTkFrame(add_card, fg_color="white")
    qty_cell.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 16))
    qty_cell.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        qty_cell,
        text="Adet",
        font=ctk.CTkFont(size=13),
        text_color="#555555",
    ).grid(row=0, column=0, sticky="w", pady=(0, 6))
    ctk.CTkEntry(
        qty_cell,
        textvariable=quantity_var,
        height=38,
        corner_radius=10,
        fg_color="#fafafa",
        border_color="#d6d6d6",
        text_color="#212121",
        width=120,
    ).grid(row=1, column=0, sticky="w")

    table_card = ctk.CTkFrame(content, fg_color="white", corner_radius=14)
    table_card.grid(row=0, column=1, rowspan=4, sticky="nsew", padx=(8, 0), pady=(0, 14))
    table_card.grid_columnconfigure(0, weight=1)
    table_card.grid_rowconfigure(1, weight=1)

    table_header = ctk.CTkFrame(table_card, fg_color="white")
    table_header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 12))
    table_header.grid_columnconfigure(0, weight=1)
    table_header.grid_columnconfigure(1, weight=0)
    table_header.grid_columnconfigure(2, weight=0)

    ctk.CTkLabel(
        table_header,
        text="Davlumbaz Listesi",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color="#212121",
    ).grid(row=0, column=0, sticky="w")

    table_frame = ctk.CTkFrame(table_card, fg_color="white")
    table_frame.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 12))
    table_frame.grid_columnconfigure(0, weight=1)
    table_frame.grid_rowconfigure(0, weight=1)

    columns = (
        "hood_id",
        "hood_type",
        "quantity",
        "width",
        "length",
        "capture_velocity",
        "distance",
        "reserve",
        "total_airflow",
    )
    tree = ttk.Treeview(
        table_frame,
        columns=columns,
        show="headings",
        selectmode="extended",
        height=14,
    )
    headings = {
        "hood_id": "Davlumbaz ID",
        "hood_type": "Davlumbaz Tipi",
        "quantity": "Adeti",
        "width": "Genisligi",
        "length": "Uzunlugu",
        "capture_velocity": "Yakalama Hizi",
        "distance": "Mesafe",
        "reserve": "Guvenlik Payi",
        "total_airflow": "Toplam Debi",
    }
    widths = {
        "hood_id": 140,
        "hood_type": 150,
        "quantity": 70,
        "width": 85,
        "length": 85,
        "capture_velocity": 110,
        "distance": 80,
        "reserve": 100,
        "total_airflow": 120,
    }
    for col in columns:
        tree.heading(col, text=headings[col])
        tree.column(col, width=widths[col], anchor="center", stretch=False)

    y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
    tree.grid(row=0, column=0, sticky="nsew")
    y_scroll.grid(row=0, column=1, sticky="ns")
    x_scroll.grid(row=1, column=0, sticky="ew")

    def clear_results() -> None:
        current_result.clear()
        hood_factor_var.set("-")
        area_var.set("-")
        airflow_m3s_var.set("-")
        airflow_m3h_var.set("-")
        reserve_airflow_var.set("-")

    def hesapla(show_errors: bool = False) -> bool:
        capture_velocity = parse_float(capture_velocity_var.get())
        distance = parse_float(distance_var.get())
        width = parse_float(width_var.get())
        length = parse_float(length_var.get())
        reserve = parse_float(reserve_var.get())

        values = [capture_velocity, distance, width, length, reserve]
        if any(value is None for value in values):
            clear_results()
            if show_errors:
                messagebox.showerror(
                    "Gecersiz veri",
                    "Tum alanlara sayisal bir deger girin.",
                    parent=pencere,
                )
            return False

        assert capture_velocity is not None
        assert distance is not None
        assert width is not None
        assert length is not None
        assert reserve is not None

        if min(capture_velocity, distance, width, length, reserve) < 0:
            clear_results()
            if show_errors:
                messagebox.showerror(
                    "Gecersiz veri",
                    "Negatif deger kullanilamaz.",
                    parent=pencere,
                )
            return False

        hood_factor = DAVLUMBAZ_TYPE_FACTORS.get(hood_type_var.get(), 1.0)
        area = width * length
        airflow_m3s = hood_factor * capture_velocity * (10 * (distance**2) + area)
        airflow_m3h = airflow_m3s * 3600.0
        airflow_with_reserve = airflow_m3h * (1 + reserve / 100.0)

        update_hood_factor()
        area_var.set(f"{format_number(area, 3)} m2")
        airflow_m3s_var.set(f"{format_number(airflow_m3s, 3)} m3/s")
        airflow_m3h_var.set(f"{format_number(airflow_m3h)} m3/h")
        reserve_airflow_var.set(f"{format_number(airflow_with_reserve)} m3/h")
        current_result.clear()
        current_result.update(
            {
                "hood_type": hood_type_var.get(),
                "capture_velocity": capture_velocity,
                "distance": distance,
                "width": width,
                "length": length,
                "reserve": reserve,
                "required_airflow_m3h": airflow_with_reserve,
            }
        )
        return True

    def build_hood_id(width_m: float, length_m: float) -> str:
        width_cm = int(round(width_m * 100))
        length_cm = int(round(length_m * 100))
        return f"DAV.{width_cm}x{length_cm}"

    def add_row() -> None:
        if not hesapla(show_errors=True):
            if not current_result:
                return

        quantity = parse_float(quantity_var.get())
        if quantity is None or quantity <= 0 or int(quantity) != quantity:
            messagebox.showerror(
                "Gecersiz adet",
                "Adet alani pozitif tam sayi olmalidir.",
                parent=pencere,
            )
            return

        quantity_int = int(quantity)
        width = float(current_result["width"])
        length = float(current_result["length"])
        total_airflow = float(current_result["required_airflow_m3h"]) * quantity_int

        row = {
            "hood_id": build_hood_id(width, length),
            "hood_type": str(current_result["hood_type"]),
            "quantity": str(quantity_int),
            "width": f"{format_number(width, 2)} m",
            "length": f"{format_number(length, 2)} m",
            "capture_velocity": f"{format_number(float(current_result['capture_velocity']), 2)} m/s",
            "distance": f"{format_number(float(current_result['distance']), 2)} m",
            "reserve": f"%{format_number(float(current_result['reserve']), 0)}",
            "total_airflow": f"{format_number(total_airflow)} m3/h",
            "total_airflow_value": total_airflow,
        }
        hood_rows.append(row)
        tree.insert("", "end", values=tuple(row[col] for col in columns))

    def delete_selected() -> None:
        selected_items = tree.selection()
        if not selected_items:
            return

        selected_indices = {tree.index(item) for item in selected_items}
        for item in selected_items:
            tree.delete(item)
        hood_rows[:] = [row for idx, row in enumerate(hood_rows) if idx not in selected_indices]

    def export_to_excel() -> None:
        if not hood_rows:
            messagebox.showwarning(
                "Liste bos",
                "Excel'e aktarmak icin once listeye en az bir davlumbaz ekleyin.",
                parent=pencere,
            )
            return

        output_path = filedialog.asksaveasfilename(
            parent=pencere,
            title="Davlumbaz Listesini Excel'e Aktar",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyasi", "*.xlsx")],
            initialfile="davlumbaz_listesi.xlsx",
        )
        if not output_path:
            return

        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Davlumbaz Listesi"
            header_labels = [headings[col] for col in columns]
            ws.append(header_labels)
            for row in hood_rows:
                ws.append(
                    [
                        row["hood_id"],
                        row["hood_type"],
                        row["quantity"],
                        row["width"],
                        row["length"],
                        row["capture_velocity"],
                        row["distance"],
                        row["reserve"],
                        row["total_airflow_value"],
                    ]
                )

            total_airflow_column = header_labels.index("Toplam Debi") + 1
            for excel_row in range(2, ws.max_row + 1):
                ws.cell(row=excel_row, column=total_airflow_column).number_format = '#.##0 "m³/h"'

            for column_cells in ws.columns:
                max_len = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 28)

            wb.save(output_path)
            messagebox.showinfo(
                "Excel'e Aktarildi",
                f"Liste basariyla disa aktarildi:\n{output_path}",
                parent=pencere,
            )
        except Exception as exc:
            messagebox.showerror(
                "Excel Aktarma Hatasi",
                f"Liste aktarilirken hata olustu:\n{exc}",
                parent=pencere,
            )

    actions = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    actions.pack(fill="x", pady=(12, 0))

    ctk.CTkButton(
        actions,
        text="Hesapla",
        width=140,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        command=hesapla,
    ).pack(side="right", padx=(12, 0))

    ctk.CTkButton(
        actions,
        text="Kapat",
        width=120,
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=pencere.destroy,
    ).pack(side="right")

    ctk.CTkButton(
        add_card,
        text="Ekle",
        width=120,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        command=add_row,
    ).grid(row=1, column=2, sticky="e", padx=(12, 16), pady=(18, 16))

    ctk.CTkButton(
        table_header,
        text="Excel'e Aktar",
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        command=export_to_excel,
    ).grid(row=0, column=2, sticky="e")

    ctk.CTkButton(
        table_header,
        text="Secili Ogeyi Sil",
        fg_color="#e0e0e0",
        hover_color="#d5d5d5",
        text_color="#424242",
        command=delete_selected,
    ).grid(row=0, column=1, sticky="e", padx=(0, 10))

    update_capture_velocity()
    update_hood_factor()
    for variable in (
        hood_type_var,
        capture_profile_var,
        capture_velocity_var,
        distance_var,
        width_var,
        length_var,
        reserve_var,
    ):
        variable.trace_add("write", lambda *_args: hesapla())
    hesapla()


def kapasite_hesaplama_ekrani_ac(
    parent: Optional[ctk.CTk] | Optional[ctk.CTkToplevel] = None,
) -> None:
    pencere = ctk.CTkToplevel(parent)
    pencere.title("Kapasite Hesaplama")
    pencere.geometry("1100x760")
    pencere.minsize(1000, 720)
    pencere.configure(fg_color="#f5f5f5")

    def _bring_to_front() -> None:
        try:
            pencere.lift()
            pencere.attributes("-topmost", True)
            pencere.after(250, lambda: pencere.attributes("-topmost", False))
            pencere.focus_force()
        except Exception:
            pass

    pencere.after(10, _bring_to_front)
    open_window_zoomed(pencere, min_width=1000, min_height=720)

    main_container = ctk.CTkFrame(pencere, fg_color="#f5f5f5")
    main_container.pack(fill="both", expand=True, padx=32, pady=24)

    header = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    header.pack(fill="x", pady=(0, 20))

    ctk.CTkLabel(
        header,
        text="Kapasite Hesaplama",
        font=ctk.CTkFont(family="Inter", size=28, weight="bold"),
        text_color="#212121",
    ).pack(anchor="w")
    ctk.CTkLabel(
        header,
        text="Alt moduller uzerinden proses ve mahal bazli kapasite hesaplamalari yapin.",
        font=ctk.CTkFont(family="Inter", size=14),
        text_color="#666666",
    ).pack(anchor="w")

    cards = ctk.CTkScrollableFrame(main_container, fg_color="#f5f5f5")
    cards.pack(fill="both", expand=True)
    for column in range(3):
        cards.grid_columnconfigure(column, weight=1)

    for idx, module in enumerate(MODULE_CONFIGS):
        row_idx = idx // 3
        col_idx = idx % 3
        cards.grid_rowconfigure(row_idx, weight=1)

        card = ctk.CTkFrame(
            cards,
            fg_color="white",
            corner_radius=15,
            width=290,
            height=230,
        )
        card.grid(row=row_idx, column=col_idx, padx=10, pady=10, sticky="nsew")
        card.grid_propagate(False)

        ctk.CTkLabel(
            card,
            text=str(module["icon"]),
            font=ctk.CTkFont(family="Inter", size=28, weight="bold"),
            text_color="#d32f2f",
        ).pack(anchor="nw", padx=20, pady=(20, 10))

        ctk.CTkLabel(
            card,
            text=str(module["title"]),
            font=ctk.CTkFont(family="Inter", size=17, weight="bold"),
            text_color="#212121",
            wraplength=230,
            justify="left",
        ).pack(anchor="nw", padx=20, pady=(0, 8))

        ctk.CTkLabel(
            card,
            text=str(module["description"]),
            font=ctk.CTkFont(family="Inter", size=12),
            text_color="#666666",
            wraplength=230,
            justify="left",
        ).pack(anchor="nw", padx=20, pady=(0, 16))

        link_button = ctk.CTkButton(
            card,
            text="Modulu Ac ->",
            font=ctk.CTkFont(family="Inter", size=12, weight="bold"),
            fg_color="white",
            hover_color="#d32f2f",
            text_color="#d32f2f",
            border_color="#d32f2f",
            border_width=2,
            height=30,
            corner_radius=15,
            command=lambda m=module: open_capacity_submodule(m, pencere),
        )

        def _on_enter(_event, button=link_button):
            button.configure(fg_color="#d32f2f", text_color="white")

        def _on_leave(_event, button=link_button):
            button.configure(fg_color="white", text_color="#d32f2f")

        link_button.bind("<Enter>", _on_enter)
        link_button.bind("<Leave>", _on_leave)
        link_button.pack(anchor="w", padx=20, pady=(0, 20))

    footer = ctk.CTkFrame(main_container, fg_color="#f5f5f5")
    footer.pack(fill="x", pady=(16, 0))

    ctk.CTkButton(
        footer,
        text="Kapat",
        width=100,
        fg_color="#d32f2f",
        hover_color="#c62828",
        text_color="white",
        command=pencere.destroy,
    ).pack(side="right")
